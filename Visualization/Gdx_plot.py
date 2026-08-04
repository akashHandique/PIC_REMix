"""
Pacific Island Countries (PICs) — Levelized Cost Assessment
============================================================
Extracts commodity balance and cost indicators from a GAMS GDX file
and computes levelized costs (LCO) for:

    Electricity · Heat · Water · Hydrogen · CO₂ ·
    Ammonia · Methanol · e-Kerosene · System Energy

Each LCO is computed as:
    LCO = (CAPEX + OPEX + Input energy costs) / Total output demand

Results are written to a single Excel file with one sheet per carrier.

Usage
-----
    python pic_lco_assessment.py
    python pic_lco_assessment.py --gdx path/to/results.gdx --out results.xlsx
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-8s %(message)s"
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

GDX_PATH = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S1_minload.gdx"

data = gdxpds.to_dataframes(GDX_PATH)
YEARS = ["2020", "2030", "2040", "2050"]
OUTPUT_PATH = "LCO_results_IP_2050_Final_S23.xlsx"
ISLANDS = [
    "CI_model", "FJ_model", "FSM_model", "KB_model", "MI_model",
    "NU_model", "NE_model", "PU_model", "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model", "VU_model",
]

# ── Population projections (*1000 persons) ────────────────────────────────────
# Source: user-provided projections; years 2020/2030/2040/2050 used directly.
# Keys are (island_model_code, year_int) -> population in thousands.
POPULATION = {
    # Cook Islands
    ("CI_model",  2020): 14,   ("CI_model",  2030): 14,
    ("CI_model",  2040): 18,   ("CI_model",  2050): 26,
    # FSM
    ("FSM_model", 2020): 113,  ("FSM_model", 2030): 126,
    ("FSM_model", 2040): 156,  ("FSM_model", 2050): 195,
    # Fiji
    ("FJ_model",  2020): 924,  ("FJ_model",  2030): 998,
    ("FJ_model",  2040): 1155, ("FJ_model",  2050): 1321,
    # Kiribati
    ("KB_model",  2020): 128,  ("KB_model",  2030): 159,
    ("KB_model",  2040): 211,  ("KB_model",  2050): 275,
    # Marshall Islands
    ("MI_model",  2020): 41,   ("MI_model",  2030): 36,
    ("MI_model",  2040): 41,   ("MI_model",  2050): 52,
    # Nauru
    ("NU_model",  2020): 12,   ("NU_model",  2030): 14,
    ("NU_model",  2040): 20,   ("NU_model",  2050): 28,
    # Niue
    ("NE_model",  2020): 1.7,  ("NE_model",  2030): 2,
    ("NE_model",  2040): 3,    ("NE_model",  2050): 5,
    # Palau
    ("PU_model",  2020): 18,   ("PU_model",  2030): 19,
    ("PU_model",  2040): 23,   ("PU_model",  2050): 27,
    # Papua New Guinea
    ("PNG_model", 2020): 9950, ("PNG_model", 2030): 12366,
    ("PNG_model", 2040): 15626,("PNG_model", 2050): 19450,
    # Samoa
    ("SA_model",  2020): 218,  ("SA_model",  2030): 242,
    ("SA_model",  2040): 307,  ("SA_model",  2050): 404,
    # Solomon Islands
    ("SI_model",  2020): 708,  ("SI_model",  2030): 998,
    ("SI_model",  2040): 1354, ("SI_model",  2050): 1773,
    # Tonga
    ("TA_model",  2020): 105,  ("TA_model",  2030): 111,
    ("TA_model",  2040): 133,  ("TA_model",  2050): 168,
    # Tuvalu
    ("TU_model",  2020): 10,   ("TU_model",  2030): 10,
    ("TU_model",  2040): 13,   ("TU_model",  2050): 18,
    # Vanuatu
    ("VU_model",  2020): 319,  ("VU_model",  2030): 391,
    ("VU_model",  2040): 513,  ("VU_model",  2050): 666,
}

# Human-readable labels for the chart legend
ISLAND_LABELS = {
    "CI_model":  "Cook Islands",
    "FJ_model":  "Fiji",
    "FSM_model": "FSM",
    "KB_model":  "Kiribati",
    "MI_model":  "Marshall Islands",
    "NU_model":  "Nauru",
    "NE_model":  "Niue",
    "PU_model":  "Palau",
    "PNG_model": "Papua New Guinea",
    "SA_model":  "Samoa",
    "SI_model":  "Solomon Islands",
    "TA_model":  "Tonga",
    "TU_model":  "Tuvalu",
    "VU_model":  "Vanuatu",
}

# ── Technology groups ─────────────────────────────────────────────────────────
TECHS = {
    "el_production":   ["DG", "PV_B", "NG_plant", "BG_B", "WindOnshore_B",
                        "Hydro_B", "BG_N", "PV_N", "WindOnshore_N",
                        "Wave_N", "WindOffshore_N", "Hydro_N", "Geothermal_B"],
    "el_storage":      ["Battery"],
    "heat_production": ["cook_b", "Industry", "DW_LPG_converter",
                        "DW_Electric_converter", "ST_N", "HP",
                        "cook_el", "cook_LPG", "Industry_EL"],
    "heat_storage":    ["THSS"],
    "water_production":["RO"],
    "water_storage":   ["H20_storage"],
    "h2_production":   ["AEL"],
    "h2_storage":      ["H2_storage"],
    "co2_production":  ["DAC"],
    "co2_storage":     ["co2_storage"],
    "ammonia_production": ["Ammonia_synthesis"],
    "ammonia_storage":    ["Ammonia_storage"],
    "methanol_production": ["Methanol_synthesis"],
    "methanol_storage":    ["Methanol_storage"],
    "ekerosene_production": ["FTL"],
    "ekerosene_storage":    ["eKerosene_storage"],
}

# Converter techs for LCOE electricity demand (unchanged from original)
ALL_CONVERTER_TECHS = [
    "BG_N", "PV_N", "WindOnshore_N", "Wave_N", "WindOffshore_N", "ST_N",
    "Industry_EL", "LDV_BF", "RO", "Ammonia_synthesis", "DAC",
    "Methanol_synthesis", "HP", "FTL", "AEL", "LDV_el", "HDV_el",
    "HDV_BF", "MDV_el", "MDV_BF", "Two_wheel_el", "Bus_el", "Marine_e",
    "Aviation_el", "Aviation_e", "cook_el", "cook_LPG", "Industry_EH",
    "DW_heat", "Dummy_Ammonia", "Dummy_Methanol",
    "DW_Electric_converter", "Ship_BEV", "Battery"
]

# Converter techs for LCOEnergy system demand
All_CONVERTERS = [
    "DW_Electric_converter",
    "DW_LPG_converter",
    "FTL",
    "Marine",
    "cook_b",
    "Industry",
    "Methanol_synthesis",
    "Ammonia_synthesis",
    "Aviation_el",
    "Aviation",
    "Bus_el",
    "Bus",
    "cook_el",
    "cook_LPG",
    "HDV_BF",
    "HDV_el",
    "HDV",
    "Industry_EL",
    "Industry_EH",
    "LDV_BF",
    "LDV_el",
    "LDV",
    "HFO",
    "MDV_BF",
    "MDV_el",
    "MDV",
    "Ship_BEV",
    "Two_wheel_el",
    "Two_wheel", "FuelImport"
]

# Final commodities for LCOEnergy — Elec excluded here; handled separately
ALL_FINAL_COMMODITIES = [
    "T_Two_wheel_th", "T_Two_wheel_el", "T_MDV_th", "T_MDV_el", "T_MDV_BF",
    "T_Marine_f_th", "T_LDV_th", "T_LDV_el", "T_LDV_BF", "T_Industry_EH",
    "T_HDV_th", "T_HDV_el", "T_HDV_BF", "T_Bus_th", "T_Bus_el",
    "T_Aviation_th", "T_Aviation_el", "Heat_industry",
    "eKerosene", "DHW_LPG", "DHW_el", "T_Marine_th",
    "Heat_cooking", "T_cook_LPG", "T_cook_el", "Methanol", "Ammonia", "T_ship_el"
]

HEAT_DEMAND_COMMODITIES = [
    "Heat_cooking", "T_cook_LPG", "T_cook_el",
    "Heat_industry", "T_Industry_EH", "DHW_el", "DHW_LPG", "Heat",
]

FUEL_COSTS = {
    "Diesel":   0.090,
    "LPG":      0.065,
    "Biomass":  0.032,
    "NG":       0.025,
}

EL_FUEL_MAP = {
    "DG":       "Diesel",
    "NG_plant": "NG",
    "BG_B":     "Biomass",
    "BG_N":     "Biomass",
}

HEAT_FUEL_MAP = {
    "cook_b":                "Biomass",
    "Industry":              "Diesel",
    "cook_LPG":              "LPG",
    "DW_LPG_converter":      "LPG",
    "DW_Electric_converter": None,
    "HP":                    None,
    "cook_el":               None,
    "Industry_EL":           None,
}

CAPEX_OPEX_INDICATORS = ["Invest", "OMFix", "OMVar"]


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Symbols loaded: %s", list(data.keys()))
    return data


# ─────────────────────────────────────────────────────────────────────────────
# QUERY HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def capex_opex(ind: pd.DataFrame, island: str, year: str,
               techs: list[str]) -> float:
    mask = (
        (ind["nodesModel"] == island) &
        (ind["years"] == year) &
        (ind["techs"].isin(techs)) &
        (ind["indicator"].isin(CAPEX_OPEX_INDICATORS))
    )
    return float(ind.loc[mask, "Value"].sum())


def commodity_flow(cb: pd.DataFrame, island: str, year: str,
                   techs: list[str], commodity: str,
                   direction: str = "input") -> float:
    mask = (
        (cb["accNodesModel"] == island) &
        (cb["accYears"] == year) &
        (cb["balanceType"] == "net") &
        (cb["techs"].isin(techs)) &
        (cb["commodity"] == commodity)
    )
    subset = cb.loc[mask, "Value"]
    if direction == "input":
        return float(abs(subset[subset < 0].sum()))
    else:
        return float(subset[subset > 0].sum())


def demand_flow(cb: pd.DataFrame, island: str, year: str,
                commodities: list[str],
                demand_tech: str = "Demand") -> float:
    mask = (
        (cb["accNodesModel"] == island) &
        (cb["accYears"] == year) &
        (cb["balanceType"] == "net") &
        (cb["techs"] == demand_tech) &
        (cb["commodity"].isin(commodities)) &
        (cb["Value"] < 0)
    )
    return float(abs(cb.loc[mask, "Value"].sum()))


# ─────────────────────────────────────────────────────────────────────────────
# LCO COMPUTATION FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def lco_electricity(data: dict, island: str, year: str) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    el_techs = TECHS["el_production"] + TECHS["el_storage"]
    cost = capex_opex(ind, island, year, el_techs)

    for tech, fuel in EL_FUEL_MAP.items():
        use = commodity_flow(cb, island, year, [tech], fuel, "input")
        cost += use * FUEL_COSTS.get(fuel, 0.0)

    direct   = demand_flow(cb, island, year, ["Elec"])
    via_conv = commodity_flow(cb, island, year, ALL_CONVERTER_TECHS, "Elec", "input")
    demand   = direct + via_conv

    return _lco_result("LCOE_Electricity", cost, demand)


def lco_heat(data: dict, island: str, year: str,
             lcoe_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    heat_techs = TECHS["heat_production"] + TECHS["heat_storage"]
    cost = capex_opex(ind, island, year, heat_techs)

    for tech, fuel in HEAT_FUEL_MAP.items():
        if fuel is None:
            elec_use = commodity_flow(cb, island, year, [tech], "Elec", "input")
            cost += elec_use * _safe_lookup(lcoe_lookup, island, year, "LCOE")
        else:
            fuel_use = commodity_flow(cb, island, year, [tech], fuel, "input")
            cost += fuel_use * FUEL_COSTS.get(fuel, 0.0)

    demand = (
        demand_flow(cb, island, year, HEAT_DEMAND_COMMODITIES)
        + commodity_flow(cb, island, year, TECHS["co2_production"], "Heat", "input")
    )

    return _lco_result("LCOHeat", cost, demand)


def lco_water(data: dict, island: str, year: str,
              lcoe_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    water_techs = TECHS["water_production"] + TECHS["water_storage"]
    cost = capex_opex(ind, island, year, water_techs)

    elec_use = commodity_flow(cb, island, year, TECHS["water_production"], "Elec", "input")
    cost += elec_use * _safe_lookup(lcoe_lookup, island, year, "LCOE")

    demand = commodity_flow(cb, island, year, TECHS["water_production"],
                            "Pure_water", "output")

    return _lco_result("LCOWater", cost, demand)


def lco_hydrogen(data: dict, island: str, year: str,
                 lcoe_lookup: dict, lco_water_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    h2_techs = TECHS["h2_production"] + TECHS["h2_storage"]
    cost = capex_opex(ind, island, year, h2_techs)

    elec_use  = commodity_flow(cb, island, year, TECHS["h2_production"], "Elec", "input")
    water_use = commodity_flow(cb, island, year, TECHS["h2_production"], "Pure_water", "input")

    cost += elec_use  * _safe_lookup(lcoe_lookup,      island, year, "LCOE")
    cost += water_use * _safe_lookup(lco_water_lookup, island, year, "LCOWater")

    demand = commodity_flow(cb, island, year, TECHS["h2_production"], "Hydrogen", "output")

    return _lco_result("LCOH2", cost, demand)


def lco_co2(data: dict, island: str, year: str,
            lcoe_lookup: dict, lco_heat_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    co2_techs = TECHS["co2_production"] + TECHS["co2_storage"]
    cost = capex_opex(ind, island, year, co2_techs)

    elec_use = commodity_flow(cb, island, year, TECHS["co2_production"], "Elec", "input")
    heat_use = commodity_flow(cb, island, year, TECHS["co2_production"], "Heat", "input")

    cost += elec_use * _safe_lookup(lcoe_lookup,     island, year, "LCOE")
    cost += heat_use * _safe_lookup(lco_heat_lookup, island, year, "LCOHeat")

    demand = commodity_flow(cb, island, year, TECHS["co2_production"], "co", "output")

    return _lco_result("LCOCO2", cost, demand)


def lco_ammonia(data: dict, island: str, year: str,
                lcoe_lookup: dict, lco_h2_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    ammo_techs = TECHS["ammonia_production"] + TECHS["ammonia_storage"]
    cost = capex_opex(ind, island, year, ammo_techs)

    elec_use = commodity_flow(cb, island, year, TECHS["ammonia_production"], "Elec", "input")
    h2_use   = commodity_flow(cb, island, year, TECHS["ammonia_production"], "Hydrogen", "input")

    cost += elec_use * _safe_lookup(lcoe_lookup,   island, year, "LCOE")
    cost += h2_use   * _safe_lookup(lco_h2_lookup, island, year, "LCOH2")

    demand = commodity_flow(cb, island, year,
                            TECHS["ammonia_production"], "Ammonia", "output")

    return _lco_result("LCOAmmonia", cost, demand)


def lco_methanol(data: dict, island: str, year: str,
                 lco_h2_lookup: dict, lco_co2_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    meoh_techs = TECHS["methanol_production"] + TECHS["methanol_storage"]
    cost = capex_opex(ind, island, year, meoh_techs)

    h2_use = commodity_flow(cb, island, year, TECHS["methanol_production"], "Hydrogen", "input")
    co_use = commodity_flow(cb, island, year, TECHS["methanol_production"], "co", "input")

    cost += h2_use * _safe_lookup(lco_h2_lookup,  island, year, "LCOH2")
    cost += co_use * _safe_lookup(lco_co2_lookup, island, year, "LCOCO2")

    demand = commodity_flow(cb, island, year,
                            TECHS["methanol_production"], "Methanol", "output")

    return _lco_result("LCOMethanol", cost, demand)


def lco_ekerosene(data: dict, island: str, year: str,
                  lco_h2_lookup: dict, lco_co2_lookup: dict) -> dict:
    ind = data["indicator_accounting_detailed"]
    cb  = data["commodity_balance_annual"]

    ekero_techs = TECHS["ekerosene_production"] + TECHS["ekerosene_storage"]
    cost = capex_opex(ind, island, year, ekero_techs)

    h2_use = commodity_flow(cb, island, year, TECHS["ekerosene_production"], "Hydrogen", "input")
    co_use = commodity_flow(cb, island, year, TECHS["ekerosene_production"], "co", "input")

    cost += h2_use * _safe_lookup(lco_h2_lookup,  island, year, "LCOH2")
    cost += co_use * _safe_lookup(lco_co2_lookup, island, year, "LCOCO2")

    demand = commodity_flow(cb, island, year, TECHS["ekerosene_production"],
                            "eKerosene", "output")

    return _lco_result("LCOeKerosene", cost, demand)


def lco_system_energy(data: dict, island: str, year: str) -> dict:
    """LCOEnergy — total system cost / total final energy demand.

    Demand has two components added together:

    1. Electricity demand — read via the Demand pseudo-tech for commodity "Elec".
       Negative values (consumption) are taken as absolute.
       This correctly captures only local electricity consumption and avoids
       counting inter-island exports that would appear in the generation-side
       balance.

    2. All other final energy — read directly from commodity_balance_annual
       using All_CONVERTERS (excludes electricity generation techs) and
       ALL_FINAL_COMMODITIES (excludes "Elec").
       Negative values (consumption by converter techs) are taken as absolute.

    The two components are summed to give total final energy demand.
    System cost is unchanged (total SystemCost from indicator_accounting).
    """
    ind = data["indicator_accounting"]
    cb  = data["commodity_balance_annual"]

    # ── System cost (unchanged) ───────────────────────────────────────────────
    mask_cost = (
        (ind["accNodesModel"] == island) &
        (ind["accYears"]      == year)   &
        (ind["indicator"]     == "SystemCost")
    )
    cost = float(ind.loc[mask_cost, "Value"].sum())

    # ── Component 1: electricity demand via Demand pseudo-tech ───────────────
    mask_elec = (
        (cb["accNodesModel"] == island) &
        (cb["accYears"]      == year)   &
        (cb["balanceType"]   == "net")  &
        (cb["techs"]         == "Demand") &
        (cb["commodity"]     == "Elec") &
        (cb["Value"] < 0)
    )
    elec_demand = float(abs(cb.loc[mask_elec, "Value"].sum()))

    # ── Component 2: all other final energy via converter techs ──────────────
    mask_other = (
        (cb["accNodesModel"] == island)                  &
        (cb["accYears"]      == year)                    &
        (cb["balanceType"]   == "net")                   &
        (cb["techs"].isin(All_CONVERTERS))               &
        (cb["commodity"].isin(ALL_FINAL_COMMODITIES))    &
        (cb["Value"] > 0)
    )
    other_demand = float(abs(cb.loc[mask_other, "Value"].sum()))

    demand = elec_demand + other_demand

    log.debug(
        "LCOEnergy [%s, %s]: elec_demand=%.2f  other_demand=%.2f  total=%.2f",
        island, year, elec_demand, other_demand, demand
    )

    return _lco_result("LCOEnergy", cost, demand)


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _lco_result(label: str, cost: float, demand: float) -> dict:
    lco = (cost / demand) if demand > 0 else None
    if demand == 0:
        log.debug("Zero demand for %s — LCO set to None.", label)
    return {label: lco, "Total_Cost": cost, "Total_Demand": demand}


def _safe_lookup(lookup: dict, island: str, year: str, label: str) -> float:
    val = lookup.get((island, int(year)))
    if val is None or (isinstance(val, float) and pd.isna(val)):
        log.warning("Missing %s for (%s, %s) — using 0.", label, island, year)
        return 0.0
    return float(val)


def build_lookup(df: pd.DataFrame, value_col: str) -> dict:
    return {(row["Island"], row["Year"]): row[value_col]
            for _, row in df.iterrows()}


def run_all(data: dict, lco_fn, *lookup_args, label: str,
            extra_cols: dict | None = None) -> pd.DataFrame:
    rows = []
    for island in ISLANDS:
        for year in YEARS:
            out = lco_fn(data, island, year, *lookup_args)
            row = {"Island": island, "Year": int(year)}
            row.update(out)
            if extra_cols:
                row.update(extra_cols)
            rows.append(row)
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# PER-CAPITA ELECTRICITY DEMAND CHART  (NEW)
# ─────────────────────────────────────────────────────────────────────────────

def plot_electricity_demand_per_capita(lcoe_df: pd.DataFrame,
                                       output_png: str = "electricity_demand_per_capita.png") -> None:
    """
    Plot per-capita electricity demand (MWh/person/year) for all 14 islands
    across 2020, 2030, 2040, 2050 using dotted lines with markers.

    The electricity demand column used is 'Total_Electricity_Demand' from the
    LCOE sheet — it already combines direct demand (via Demand pseudo-tech)
    and via-converter demand, exactly mirroring the lco_electricity() logic.

    Population is sourced from the POPULATION dict (in *1000 persons), so the
    demand (assumed to be in GWh or the model's native energy unit — here kept
    as-is and labelled accordingly) is divided by population * 1000 to get the
    per-person value.  Units on the y-axis label should be adjusted if the
    model energy unit differs from MWh.
    """
    years_int = [int(y) for y in YEARS]

    # Use a qualitatively distinct colour palette for 14 lines
    cmap = plt.get_cmap("tab20")
    colors = [cmap(i) for i in range(len(ISLANDS))]

    fig, ax = plt.subplots(figsize=(12, 7))

    for idx, island in enumerate(ISLANDS):
        per_capita_values = []
        valid = True

        for year in YEARS:
            row = lcoe_df[
                (lcoe_df["Island"] == island) & (lcoe_df["Year"] == int(year))
            ]
            if row.empty:
                log.warning("No LCOE row for (%s, %s) — skipping island in chart.", island, year)
                valid = False
                break

            total_demand = float(row["Total_Electricity_Demand"].iloc[0])
            pop_thousands = POPULATION.get((island, int(year)))

            if pop_thousands is None or pop_thousands == 0:
                log.warning("No population for (%s, %s) — skipping island in chart.", island, year)
                valid = False
                break

            # population in thousands → multiply by 1000 for actual headcount
            pop = pop_thousands 
            per_capita_values.append(total_demand / pop)

        if not valid:
            continue

        label = ISLAND_LABELS.get(island, island)
        ax.plot(
            years_int,
            per_capita_values,
            linestyle=":",          # dotted line
            linewidth=2.0,
            marker="o",
            markersize=6,
            color=colors[idx],
            label=label,
        )

        # Annotate the 2050 end-point with the island name for readability
        ax.annotate(
            label,
            xy=(years_int[-1], per_capita_values[-1]),
            xytext=(5, 0),
            textcoords="offset points",
            fontsize=7,
            color=colors[idx],
            va="center",
        )

    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Electricity Demand per Capita\n(MWh / person)", fontsize=11)
    ax.set_title("",
                 fontsize=13, fontweight="bold")
    ax.set_xticks(years_int)
    ax.xaxis.set_minor_locator(mticker.NullLocator())
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.4f"))
    ax.legend(
        loc="upper left",
        fontsize=8,
        framealpha=0.7,
        ncol=2,
        title="Island / Country",
        title_fontsize=9,
    )
    ax.grid(True, linestyle="--", alpha=0.4)
    fig.tight_layout()
    fig.savefig(output_png, dpi=150, bbox_inches="tight")
    log.info("Per-capita electricity demand chart saved to: %s", output_png)
    plt.show()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, output_path: str,preloaded_data=None) -> None:


    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_gdx(gdx_path)

    log.info("Computing LCOE Electricity …")
    lcoe_df = run_all(data, lco_electricity, label="LCOE_Electricity")
    lcoe_df.rename(columns={"Total_Cost": "Total_Electricity_Cost",
                             "Total_Demand": "Total_Electricity_Demand"}, inplace=True)
    lcoe_lookup = build_lookup(lcoe_df, "LCOE_Electricity")

    log.info("Computing LCOHeat …")
    heat_df = run_all(data, lco_heat, lcoe_lookup, label="LCOHeat")
    heat_df.rename(columns={"Total_Cost": "Total_Heat_Cost",
                             "Total_Demand": "Total_Heat_Demand"}, inplace=True)
    lco_heat_lookup = build_lookup(heat_df, "LCOHeat")

    log.info("Computing LCOWater …")
    water_df = run_all(data, lco_water, lcoe_lookup, label="LCOWater")
    water_df.rename(columns={"Total_Cost": "Total_Water_Cost",
                              "Total_Demand": "Total_Water_Demand"}, inplace=True)
    lco_water_lookup = build_lookup(water_df, "LCOWater")

    log.info("Computing LCOH2 …")
    h2_df = run_all(data, lco_hydrogen, lcoe_lookup, lco_water_lookup, label="LCOH2")
    h2_df.rename(columns={"Total_Cost": "Total_H2_Cost",
                           "Total_Demand": "Total_H2_Demand"}, inplace=True)
    lco_h2_lookup = build_lookup(h2_df, "LCOH2")

    log.info("Computing LCOCO2 …")
    co_df = run_all(data, lco_co2, lcoe_lookup, lco_heat_lookup, label="LCOCO2")
    co_df.rename(columns={"Total_Cost": "Total_CO2_Cost",
                           "Total_Demand": "Total_CO2_Demand"}, inplace=True)
    lco_co2_lookup = build_lookup(co_df, "LCOCO2")

    log.info("Computing LCOAmmonia …")
    ammonia_df = run_all(data, lco_ammonia, lcoe_lookup, lco_h2_lookup,
                         label="LCOAmmonia")
    ammonia_df.rename(columns={"Total_Cost": "Total_Ammonia_Cost",
                                "Total_Demand": "Total_Ammonia_Demand"}, inplace=True)

    log.info("Computing LCOMethanol …")
    methanol_df = run_all(data, lco_methanol, lco_h2_lookup, lco_co2_lookup,
                          label="LCOMethanol")
    methanol_df.rename(columns={"Total_Cost": "Total_Methanol_Cost",
                                 "Total_Demand": "Total_Methanol_Demand"}, inplace=True)

    log.info("Computing LCOeKerosene …")
    ekerosene_df = run_all(data, lco_ekerosene, lco_h2_lookup, lco_co2_lookup,
                           label="LCOeKerosene")
    ekerosene_df.rename(columns={"Total_Cost": "Total_eKerosene_Cost",
                                  "Total_Demand": "Total_eKerosene_Demand"}, inplace=True)

    log.info("Computing LCOEnergy (system) …")
    system_df = run_all(data, lco_system_energy, label="LCOEnergy")
    system_df.rename(columns={"Total_Cost": "Total_Energy_Cost",
                               "Total_Demand": "Total_Energy_Demand"}, inplace=True)

    sheets = {
        "LCOE_Electricity": lcoe_df,
        "LCOHeat":          heat_df,
        "LCOWater":         water_df,
        "LCOH2":            h2_df,
        "LCOCO2":           co_df,
        "LCOAmmonia":       ammonia_df,
        "LCOMethanol":      methanol_df,
        "LCOeKerosene":     ekerosene_df,
        "LCOEnergy":        system_df,
    }

    log.info("Writing results to: %s", output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    log.info("Done. %d sheets written.", len(sheets))

    print("\n── Summary (non-null LCO counts per carrier) ──────────────────")
    for sheet, df in sheets.items():
        col = sheet
        if col in df.columns:
            n_valid = df[col].notna().sum()
            n_total = len(df)
            print(f"  {sheet:<22} {n_valid}/{n_total} island-year pairs have results")
    print()

    # ── NEW: per-capita electricity demand chart ──────────────────────────────
    log.info("Generating per-capita electricity demand chart …")
    chart_path = Path(output_path).stem + "_electricity_per_capita.png"
    plot_electricity_demand_per_capita(lcoe_df, output_png=chart_path)


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--gdx", default=GDX_PATH,
                        help="Path to GAMS GDX results file")
    parser.add_argument("--out", default=OUTPUT_PATH,
                        help="Output Excel file path")
    args = parser.parse_args()

    main(args.gdx, args.out)
##########################all commodity values are written to an excel###############################
######now for plotiing these results###############################################################
# Default — PNG at 300 dpi, saved to figures/
"""
Pacific Island Countries (PICs) — LCO Results Visualisation
============================================================
Reads the LCO Excel output and produces publication-quality bar charts
for each energy carrier. All plots share a consistent style suitable
for scientific journals (Nature Energy, Applied Energy, etc.).

Usage
-----
    python pic_lco_plots.py
    python pic_lco_plots.py --file path/to/results.xlsx --dpi 600 --fmt pdf
"""

import argparse
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL PLOT STYLE  (journal-ready)
# ─────────────────────────────────────────────────────────────────────────────

rcParams.update({
    "font.family":        "sans-serif",
    "font.sans-serif":    ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":          10,
    "axes.titlesize":     11,
    "axes.labelsize":     10,
    "xtick.labelsize":    9,
    "ytick.labelsize":    9,
    "legend.fontsize":    9,
    "legend.title_fontsize": 9,
    "axes.linewidth":     0.8,
    "axes.spines.top":    False,
    "axes.spines.right":  False,
    "xtick.direction":    "out",
    "ytick.direction":    "out",
    "xtick.major.size":   3,
    "ytick.major.size":   3,
    "xtick.major.width":  0.8,
    "ytick.major.width":  0.8,
    "grid.linewidth":     0.5,
    "grid.alpha":         0.4,
    "grid.color":         "#888888",
    "figure.dpi":         150,
    "savefig.dpi":        300,
    "savefig.bbox":       "tight",
    "savefig.facecolor":  "white",
})

# ── Colour palette — colorblind-safe (Wong 2011, extended) ───────────────────
YEAR_COLORS = {
    2020: "#0072B2",   # blue
    2030: "#E69F00",   # amber
    2040: "#009E73",   # teal-green
    2050: "#D55E00",   # vermillion
}

YEARS      = [2020, 2030, 2040, 2050]
BAR_WIDTH  = 0.18
FIG_SIZE   = (14, 5)

# ── Clean island labels (strip "_model" suffix) ──────────────────────────────
def clean_labels(islands: np.ndarray) -> list[str]:
    return [s.replace("_model", "") for s in islands]


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION TABLE
# Each entry defines one chart completely.
# ─────────────────────────────────────────────────────────────────────────────

# reference_line: (y_value_in_MWh, label_text) or None
CHART_CONFIG = [
    {
        "sheet":      "LCOE_Electricity",
        "value_col":  "LCOE_Electricity",
        "ylabel":     "Levelized cost of electricity ($/MWh)",
        "title":      "Levelized Cost of Electricity by Island and Year",
        "filename":   "LCOE_Electricity",
        "reference":  None,
        "show_legend": True,
    },
    {
        "sheet":      "LCOHeat",
        "value_col":  "LCOHeat",
        "ylabel":     "Levelized cost of heat ($/MWh)",
        "title":      "Levelized Cost of Heat by Island and Year",
        "filename":   "LCOHeat",
        "reference":  None,
        "show_legend": False,
    },
    {
        "sheet":      "LCOEnergy",
        "value_col":  "LCOEnergy",
        "ylabel":     "Levelized cost of energy ($/MWh)",
        "title":      "System-Level Levelized Cost of Energy by Island and Year",
        "filename":   "LCOEnergy",
        "reference":  None,
        "show_legend": False,
    },
    {
        "sheet":      "LCOH2",
        "value_col":  "LCOH2",
        "ylabel":     "Levelized cost of hydrogen ($/MWh)",
        "title":      "Levelized Cost of Hydrogen by Island and Year",
        "filename":   "LCOH2",
        "reference":  (0,  ""),
        "show_legend": False,
    },
    {
        "sheet":      "LCOAmmonia",
        "value_col":  "LCOAmmonia",
        "ylabel":     "Levelized cost of ammonia ($/MWh)",
        "title":      "Levelized Cost of Ammonia  by Island and Year",
        "filename":   "LCOAmmonia",
        "reference":  (0, ""),
        "show_legend": False,
    },
    {
        "sheet":      "LCOMethanol",
        "value_col":  "LCOMethanol",
        "ylabel":     "Levelized cost of methanol ($/MWh)",
        "title":      "Levelized Cost of Methanol by Island and Year",
        "filename":   "LCOMethanol",
        "reference":  (0,  ""),
        "show_legend": False,
    },
    {
        "sheet":      "LCOeKerosene",
        "value_col":  "LCOeKerosene",
        "ylabel":     "Levelized cost of e-kerosene ($/MWh)",
        "title":      "Levelized Cost of e-Kerosene  by Island and Year",
        "filename":   "LCOeKerosene",
        "reference":  (0, ""),
        "show_legend": False,
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# CORE PLOT FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def plot_lco(
    df: pd.DataFrame,
    value_col: str,
    ylabel: str,
    title: str,
    reference: tuple | None,
    show_legend: bool,
    output_path: Path,
    fmt: str,
    dpi: int,
) -> None:
    """
    Render a grouped bar chart for one LCO carrier.

    Parameters
    ----------
    df          : DataFrame with columns [Island, Year, value_col]
    value_col   : name of the LCO column
    ylabel      : y-axis label
    title       : chart title
    reference   : (y_value_MWh, label_str) for an optional reference line, or None
    show_legend : whether to draw the year legend
    output_path : full save path (stem + suffix applied automatically)
    fmt         : file format ("png", "pdf", "svg", "tiff")
    dpi         : save resolution
    """
    df = df[["Island", "Year", value_col]].sort_values(["Island", "Year"])
    islands = df["Island"].unique()
    labels  = clean_labels(islands)
    x       = np.arange(len(islands))

    fig, ax = plt.subplots(figsize=FIG_SIZE)

    # ── Bars ─────────────────────────────────────────────────────────────────
    for i, year in enumerate(YEARS):
        values = (
            df[df["Year"] == year]
            .set_index("Island")
            .reindex(islands)[value_col]
            .values
        ) * 1000   # convert model units → USD/MWh

        bars = ax.bar(
            x + i * BAR_WIDTH,
            values,
            width=BAR_WIDTH,
            label=str(year),
            color=YEAR_COLORS[year],
            edgecolor="white",
            linewidth=0.4,
            zorder=3,
        )

    # ── Reference line (drawn once, outside the year loop) ───────────────────
    if reference is not None:
        ref_val, ref_label = reference
        ax.axhline(
            y=ref_val,
            color="#CC0000",
            linestyle="--",
            linewidth=1.2,
            zorder=4,
            label=ref_label,
        )
        # Annotate on the right margin
        ax.annotate(
            ref_label,
            xy=(1, ref_val),
            xycoords=("axes fraction", "data"),
            xytext=(4, 0),
            textcoords="offset points",
            va="center",
            fontsize=8,
            color="#CC0000",
        )

    # ── Axes formatting ───────────────────────────────────────────────────────
    ax.set_xlabel("Island", labelpad=6)
    ax.set_ylabel(ylabel, labelpad=6)
    ax.set_title(title, pad=10, fontweight="bold")

    ax.set_xticks(x + BAR_WIDTH * (len(YEARS) - 1) / 2)
    ax.set_xticklabels(labels, rotation=40, ha="right")

    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda val, _: f"{val:,.0f}"
    ))
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)

    # ── Legend ────────────────────────────────────────────────────────────────
    if show_legend:
        handles = [
            plt.Rectangle((0, 0), 1, 1, fc=YEAR_COLORS[y], ec="white", lw=0.4)
            for y in YEARS
        ]
        ax.legend(
            handles,
            [str(y) for y in YEARS],
            title="Year",
            ncol=4,
            frameon=False,
            loc="upper right",
        )

    fig.tight_layout()

    save_path = output_path.with_suffix(f".{fmt}")
    fig.savefig(save_path, dpi=dpi)
    # NOTE: fig is NOT closed here so plt.show() displays all figures together
    print(f"  Saved → {save_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(file_path: str, out_dir: str, fmt: str, dpi: int) -> None:
    xlsx    = Path(file_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Reading: {xlsx}")
    print(f"Saving to: {out_dir}/  [{fmt.upper()}, {dpi} dpi]\n")

    for cfg in CHART_CONFIG:
        sheet = cfg["sheet"]
        print(f"Plotting {sheet} …")

        try:
            df = pd.read_excel(xlsx, sheet_name=sheet)
        except Exception as e:
            print(f"  [SKIP] Could not read sheet '{sheet}': {e}")
            continue

        # Validate column exists
        if cfg["value_col"] not in df.columns:
            print(f"  [SKIP] Column '{cfg['value_col']}' not found in sheet '{sheet}'.")
            continue

        plot_lco(
            df          = df,
            value_col   = cfg["value_col"],
            ylabel      = cfg["ylabel"],
            title       = cfg["title"],
            reference   = cfg["reference"],
            show_legend = cfg["show_legend"],
            output_path = out_dir / cfg["filename"],
            fmt         = fmt,
            dpi         = dpi,
        )

    print("\nAll plots complete.")
    plt.show()   # display all figures; blocks until windows are closed


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--file", default="LCO_results_IP_2050_Final_S1.xlsx",
        help="Path to the LCO Excel results file (default: LCO_results_IP_2050_Final_S1.xlsx)"
    )
    parser.add_argument(
        "--out", default="figures/S_1/",
        help="Output directory for figures (default: figures/S_1/)"
    )
    parser.add_argument(
        "--fmt", default="png", choices=["png", "pdf", "svg", "tiff"],
        help="Output file format (default: png)"
    )
    parser.add_argument(
        "--dpi", default=600, type=int,
        help="Save resolution in DPI (default: 600)"
    )
    args = parser.parse_args()

    main(args.file, args.out, args.fmt, args.dpi)
##########################hourly elec generation profile#############################################################################################################################################################
"""
Pacific Island Countries (PICs) — Hourly Electricity Generation Profiles
=========================================================================
Opens the GAMS GDX file and reads the hourly commodity_balance table
(8760 timesteps × 14 islands × 4 years) to produce stacked-area generation
profiles for each island.

Layout
------
One figure per island  →  2 × 2 grid of subplots (one per target year).
Each subplot shows:
  • Stacked area chart  — hourly generation by technology (GWh)
  • Demand line         — total electricity demand (GWh, black)
  • Curtailment shade   — generation above demand, hatched on top (grey)
  • Curtailment label   — annual curtailment (GWh) and % of demand

Output
------
  figures/generation_profiles/
      CI_model.png
      FJ_model.png
      ...  (one file per island)

Usage
-----
    python pic_generation_profiles.py
    python pic_generation_profiles.py --gdx path/to/results.gdx
                                      --out figures/generation_profiles
                                      --fmt png --dpi 200
                                      --islands CI_model FJ_model
                                      --years 2020 2050
"""

"""
Pacific Island Countries (PICs) — Hourly Electricity Generation Profiles
=========================================================================
Opens the GAMS GDX file and reads the hourly commodity_balance table
(8760 timesteps × 14 islands × 4 years) to produce stacked-area generation
profiles for each island.

Layout
------
One figure per island  →  2 × 2 grid of subplots (one per target year).
Each subplot shows:
  • Stacked area chart  — 2-week centred rolling-average generation by
                          technology (MWh)
  • Demand line         — 2-week rolling-average total electricity demand
                          (MWh, black)
  • Curtailment shade   — rolling-smoothed excess generation above demand,
                          hatched grey on top
  • Top annotation      — annual totals: Gen / Demand / Curtailment (MWh)
                          and curtailment as % of demand

Rolling average
---------------
  window = 336 h (14 days), centred, min_periods=1.
  Annual totals in the annotation are computed from the raw hourly data
  before smoothing, so they are exact.

Output
------
  figures/generation_profiles/
      CI_model.png
      FJ_model.png
      ...  (one PNG per island)

Usage
-----
    python pic_generation_profiles.py
    python pic_generation_profiles.py --gdx path/to/results.gdx
                                      --out figures/generation_profiles
                                      --fmt png --dpi 200
                                      --islands CI_model FJ_model
                                      --years 2020 2050
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.patches import Patch

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/generation_profiles_minload_S23"
YEARS      = ["2020", "2030", "2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

# ── Electricity generation technologies (supply side, positive flows) ─────────
EL_PRODUCTION_TECHS = [
    "DG", "PV_B", "NG_plant", "BG_B", "WindOnshore_B", "Hydro_B",
    "BG_N", "PV_N", "WindOnshore_N", "Wave_N", "WindOffshore_N", "Hydro_N", "Geothermal_B", "Battery"
]

# ── Electricity demand techs (consumption side, negative flows in cb) ─────────
All_Elec_demand_tech = ['DW_Electric_converter',
                        'Ammonia_synthesis', 'Aviation_el',  'Bus_el', 'DAC',
                        'HDV_el', 'HP', 'Industry_EL', 'LDV_el', 'MDV_el', 'RO', 'Ship_BEV', 'Two_wheel_el',
                        'cook_el', 'AEL', 'Battery']

# ── Technology display names & colours ────────────────────────────────────────
# Colorblind-safe palette (Wong 2011 + extensions)
TECH_STYLE = {
    "PV_B":           ("Solar PV (behind-meter)", "#F0C234"),
    "PV_N":           ("Solar PV (new)",           "#E6A817"),
    "WindOnshore_B":  ("Wind onshore (existing)",  "#56B4E9"),
    "WindOnshore_N":  ("Wind onshore (new)",       "#0072B2"),
    "WindOffshore_N": ("Wind offshore",            "#004D80"),
    "Wave_N":         ("Wave",                     "#009E73"),
    "Hydro_B":        ("Hydro (existing)",         "#44AA99"),
    "Hydro_N":        ("Hydro (new)",              "#117733"),
    "BG_B":           ("Biogas (existing)",        "#CC79A7"),
    "BG_N":           ("Biogas (new)",             "#882255"),
    "NG_plant":       ("Natural gas",              "#E69F00"),
    "DG":             ("Diesel generator",         "#D55E00"),
    "Geothermal_B":   ("Geothermal",               "#8C564B"),
    "Battery":        ("Battery_dis",              "#B87333"),
}

FALLBACK_COLORS = ["#999999", "#BBBBBB", "#666666", "#444444", "#CCCCCC"]

# ── Rolling-average window ────────────────────────────────────────────────────
ROLLING_WINDOW = 14 * 24   # 336 hours = 2 weeks, centred

# ── Month layout helpers ──────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 744, 1416, 2160, 2880, 3624, 4344,
                    5088, 5832, 6552, 7296, 8016, 8760]
MONTH_LABELS     = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) // 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE  (journal-ready)
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "legend.fontsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "xtick.direction":   "out",
    "ytick.direction":   "out",
    "xtick.major.size":  3,
    "ytick.major.size":  3,
    "grid.linewidth":    0.4,
    "grid.alpha":        0.35,
    "grid.color":        "#888888",
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.bbox":      "tight",
    "savefig.facecolor": "white",
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_hourly_cb(data: dict) -> pd.DataFrame:
    """
    Pull the hourly commodity_balance table and pre-filter to Elec only.
    Converts timeModel strings ('tm1'...'tm8760') to 0-based integer hour index.

    Returns DataFrame with columns:
        hour (int 0-8759), island, year, tech, value_gwh
    Positive values = production; negative = consumption.
    """
    cb = data["commodity_balance"]
    log.info("commodity_balance rows: %d", len(cb))

    cb_elec = cb[cb["commodity"] == "Elec"].copy()
    cb_elec["hour"] = (
        cb_elec["timeModel"]
        .str.replace("tm", "", regex=False)
        .astype(int) - 1
    )
    cb_elec.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "techs":         "tech",
        "Value":         "value_gwh",
    }, inplace=True)

    return cb_elec[["hour", "island", "year", "tech", "value_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND / PER-YEAR DATA PREP
# ─────────────────────────────────────────────────────────────────────────────

def get_generation(cb_elec: pd.DataFrame, island: str, year: str) -> pd.DataFrame:
    """
    Returns an (8760 x n_techs) DataFrame of raw hourly generation (GWh >= 0)
    for electricity production technologies.
    """
    mask = (
        (cb_elec["island"] == island) &
        (cb_elec["year"]   == year) &
        (cb_elec["tech"].isin(EL_PRODUCTION_TECHS)) &
        (cb_elec["value_gwh"] > 0)
    )
    sub = cb_elec.loc[mask, ["hour", "tech", "value_gwh"]]

    pivot = (
        sub.pivot_table(index="hour", columns="tech",
                        values="value_gwh", aggfunc="sum")
           .reindex(range(8760), fill_value=0.0)
           .fillna(0.0)
    )
    # Drop techs with zero generation in this island-year
    pivot = pivot.loc[:, (pivot > 0).any(axis=0)]
    return pivot


def get_demand(cb_elec: pd.DataFrame, island: str, year: str) -> pd.Series:
    """
    Hourly electricity demand (GWh, positive), indexed 0-8759.
    = |Demand pseudo-tech| + |all converter techs consuming Elec|
    Mirrors the denominator logic from lco_electricity().
    """
    base_mask = (
        (cb_elec["island"] == island) &
        (cb_elec["year"]   == year) &
        (cb_elec["value_gwh"] < 0)
    )
    direct = (
        cb_elec.loc[base_mask & (cb_elec["tech"] == "Demand")]
        .groupby("hour")["value_gwh"].sum()
        .abs()
    )
    conv = (
        cb_elec.loc[base_mask & cb_elec["tech"].isin(All_Elec_demand_tech)]
        .groupby("hour")["value_gwh"].sum()
        .abs()
    )
    return direct.add(conv, fill_value=0.0).reindex(range(8760), fill_value=0.0)


def compute_curtailment(gen: pd.DataFrame, demand: pd.Series):
    """
    Compute curtailment as the hourly surplus of generation over demand.

    curtailment[h] = max(0,  total_gen[h] - demand[h])   [GWh]

    Annual curtailment (GWh) is the sum of these hourly values.
    Curtailment % is expressed relative to annual demand.

    Returns
    -------
    total_gen     : pd.Series  hourly total generation (GWh)
    curtailment   : pd.Series  hourly curtailed energy  (GWh, >= 0)
    curt_annual   : float      annual curtailment sum    (GWh)
    curt_pct      : float      curtailment as % of annual demand
    """
    total_gen     = gen.sum(axis=1)
    curtailment   = (total_gen - demand).clip(lower=0.0)   # surplus each hour
    curt_annual   = float(curtailment.sum())               # GWh
    demand_annual = float(demand.sum())
    curt_pct      = (curt_annual / demand_annual * 100) if demand_annual > 0 else 0.0
    return total_gen, curtailment, curt_annual, curt_pct


# ─────────────────────────────────────────────────────────────────────────────
# ROLLING AVERAGE
# ─────────────────────────────────────────────────────────────────────────────

def rolling_avg(arr: np.ndarray) -> np.ndarray:
    """
    Centred rolling average with window = ROLLING_WINDOW (336 h = 2 weeks).
    Returns a full-length 8760-point smoothed array.
    min_periods=1 prevents NaN at the first/last 168 hours where the full
    window is not available.
    """
    return (
        pd.Series(arr)
        .rolling(window=ROLLING_WINDOW, center=True, min_periods=1)
        .mean()
        .values
    )


# ─────────────────────────────────────────────────────────────────────────────
# COLOUR HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def assign_colors(techs: list) -> dict:
    """Return a {tech: hex_colour} map for every tech in the list."""
    color_map    = {}
    fallback_idx = 0
    for tech in techs:
        if tech in TECH_STYLE:
            color_map[tech] = TECH_STYLE[tech][1]
        else:
            color_map[tech] = FALLBACK_COLORS[fallback_idx % len(FALLBACK_COLORS)]
            fallback_idx   += 1
    return color_map


def tech_label(tech: str) -> str:
    return TECH_STYLE.get(tech, (tech, ""))[0]


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def draw_subplot(ax, gen: pd.DataFrame, demand: pd.Series,
                 year: str, color_map: dict, show_ylab: bool):
    """
    Draw one year panel into ax.

    Steps
    -----
    1. Compute exact annual totals from raw hourly GWh data.
    2. Apply 2-week centred rolling average to all series.
    3. Convert smoothed values GWh -> MWh for plotting.
    4. Stacked-area chart (techs) + demand line + curtailment hatch.
    5. Print annual summary above the subplot title.
       Curtailment is reported in GWh (summed from raw hourly values).
    """

    # ── Step 1: exact annual totals (raw hourly GWh) ──────────────────────────
    total_gen, curtailment, curt_annual_gwh, curt_pct = compute_curtailment(gen, demand)

    gen_annual_mwh    = float(total_gen.sum()) * 1e3
    demand_annual_mwh = float(demand.sum())    * 1e3

    if gen.empty:
        ax.text(0.5, 0.5, "No generation data",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return

    # ── Steps 2 & 3: rolling average + GWh -> MWh ────────────────────────────
    hours     = np.arange(8760)
    tech_cols = list(gen.columns)

    gen_smooth    = {t: rolling_avg(gen[t].values) * 1e3 for t in tech_cols}
    demand_smooth = rolling_avg(demand.values) * 1e3
    curt_smooth   = rolling_avg(curtailment.values) * 1e3

    # ── Step 4a: stacked area ─────────────────────────────────────────────────
    stack_data = [gen_smooth[t] for t in tech_cols]
    colors     = [color_map.get(t, "#999999") for t in tech_cols]

    ax.stackplot(hours, stack_data,
                 labels=[tech_label(t) for t in tech_cols],
                 colors=colors, linewidth=0, zorder=2)

    # ── Step 4b: curtailment hatch on top of demand ───────────────────────────
    curt_bottom = np.clip(demand_smooth, a_min=0, a_max=None)
    ax.fill_between(hours,
                    curt_bottom,
                    curt_bottom + curt_smooth,
                    color="#AAAAAA", alpha=0.55, hatch="////",
                    linewidth=0, label="Curtailment", zorder=3)

    # ── Step 4c: demand line ──────────────────────────────────────────────────
    ax.plot(hours, demand_smooth,
            color="black", linewidth=1.0, label="Demand", zorder=5)

    # ── Step 5: annual summary above subplot title ────────────────────────────
    # Curtailment is the raw hourly sum (GWh), not converted to MWh, per request
    ax.set_title(year, fontweight="bold", pad=16)
    ax.text(
        0.5, 1.02,
        (f"Gen: {gen_annual_mwh:,.0f} MWh  |  "
         f"Demand: {demand_annual_mwh:,.0f} MWh  |  "
         f"Curtailment: {curt_annual_gwh * 1e3:,.0f} MWh"),
        transform=ax.transAxes,
        ha="center", va="bottom",
        fontsize=7.2, color="#333333", fontstyle="italic",
    )

    # ── X-axis: month labels ──────────────────────────────────────────────────
    ax.set_xlim(0, 8759)
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=7.5)

    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="#CCCCCC", linewidth=0.4, zorder=1)

    # ── Y-axis: MWh, comma-formatted ─────────────────────────────────────────
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)

    if show_ylab:
        ax.set_ylabel("Generation - 2-week rolling avg (MWh)", labelpad=4)

    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v:,.0f}")
    )


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE  (2 x 2 subplots)
# ─────────────────────────────────────────────────────────────────────────────

def plot_island(cb_elec: pd.DataFrame, island: str,
                years: list, color_map: dict,
                out_dir: Path, fmt: str, dpi: int):
    """Create and save the 2x2 generation profile figure for one island."""

    island_label = island.replace("_model", "")
    fig, axes    = plt.subplots(2, 2, figsize=(18, 9), sharey=False)
    axes_flat    = axes.flatten()

    fig.suptitle(
        f"{island_label} - Electricity Generation Profile "
        f"(2-week centred rolling average)",
        fontsize=12, fontweight="bold", y=1.02,
    )

    # Collect legend handles across all subplots (deduplicated)
    legend_handles = {}

    for idx, year in enumerate(years):
        ax        = axes_flat[idx]
        show_ylab = (idx % 2 == 0)    # left column only

        gen    = get_generation(cb_elec, island, year)
        demand = get_demand(cb_elec, island, year)

        draw_subplot(ax, gen, demand, year, color_map, show_ylab)

        for handle, lbl in zip(*ax.get_legend_handles_labels()):
            if lbl not in legend_handles:
                legend_handles[lbl] = handle

    # ── Shared legend: generation techs first, then Demand & Curtailment ──────
    special     = {"Demand", "Curtailment"}
    gen_handles = {k: v for k, v in legend_handles.items() if k not in special}
    spc_handles = {k: v for k, v in legend_handles.items() if k in special}

    all_handles = list(gen_handles.values()) + list(spc_handles.values())
    all_labels  = list(gen_handles.keys())   + list(spc_handles.keys())

    fig.legend(
        all_handles, all_labels,
        loc="lower center",
        ncol=min(len(all_labels), 7),
        frameon=False,
        fontsize=8,
        bbox_to_anchor=(0.5, -0.04),
    )

    plt.tight_layout(rect=[0, 0.0, 1, 1])

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # # Load GDX and extract hourly Elec balance
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_gdx(gdx_path)
    cb_elec = extract_hourly_cb(data)

    log.info("Unique techs in hourly Elec balance: %s",
             sorted(cb_elec["tech"].unique()))

    # Build global colour map (RE first, thermal last)
    present_techs = (
        cb_elec.loc[
            cb_elec["tech"].isin(EL_PRODUCTION_TECHS) & (cb_elec["value_gwh"] > 0),
            "tech",
        ].unique().tolist()
    )
    ordered   = [t for t in EL_PRODUCTION_TECHS if t in present_techs]
    color_map = assign_colors(ordered)

    # Plot each island
    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island(cb_elec, island, years, color_map, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island figures written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH,
                        help="Path to GAMS GDX file")
    parser.add_argument("--out",     default=OUTPUT_DIR,
                        help="Output directory for figures")
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"],
                        help="Output file format (default: png)")
    parser.add_argument("--dpi",     default=200, type=int,
                        help="Save resolution in DPI (default: 200)")
    parser.add_argument("--islands", nargs="+", default=ISLANDS,
                        help="Subset of islands to plot (default: all 14)")
    parser.add_argument("--years",   nargs="+", default=YEARS,
                        help="Target years (default: 2020 2030 2040 2050)")
    args = parser.parse_args()

    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
    ################################################################################################################battery heatmap##############################
"""
Pacific Island Countries (PICs) — Battery State of Charge Heatmaps
====================================================================
Reads 'Storage_level_out' from the GAMS GDX file and produces
24 × 365 heatmaps of hourly battery State of Charge (SOC) expressed
as a percentage of the annual peak storage level.

Layout
------
One figure per island  →  1 × 3 row of subplots (one per target year).
Each subplot is a heatmap where:
    x-axis  : month labels (Jan–Dec)
    y-axis  : hour of day (0–23)
    colour  : SOC (%) = value / max(value for that island-year) × 100

SOC = 0 % → battery empty   |   SOC = 100 % → battery at peak capacity

Output
------
  figures/battery_soc/
      CI_model.png
      FJ_model.png
      ...  (one PNG per island)

Usage
-----
    python pic_battery_soc.py
    python pic_battery_soc.py --gdx path/to/results.gdx
                              --out figures/battery_soc
                              --fmt png --dpi 200
                              --islands CI_model FJ_model
                              --years 2030 2040 2050
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/battery_soc_minload"
YEARS      = ["2030", "2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

BATTERY_TECH = "Battery"

# ── Month axis ────────────────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR MAP
# YlGnBu-inspired: pale yellow (empty) → mint → teal → deep ocean blue (full)
# Perceptually uniform, colorblind-safe, prints well in greyscale too.
# ─────────────────────────────────────────────────────────────────────────────
SOC_CMAP = LinearSegmentedColormap.from_list(
    "soc",
    ["#FFFFCC", "#C7E9B4", "#41B6C4", "#1D91C0", "#0C2C84"],
    N=256,
)

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# IMPORTANT: savefig.bbox is intentionally NOT "tight".
# "tight" re-expands the canvas after layout and drags the colorbar axis
# back over the rightmost heatmap panel.  All margins are controlled via
# gridspec left/right/top/bottom fractions instead.
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
    # savefig.bbox deliberately omitted (defaults to "standard")
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_battery_soc(data: dict) -> pd.DataFrame:
    sl  = data["storage_level_out"]
    bat = sl[sl["techs"] == BATTERY_TECH].copy()

    if bat.empty:
        log.warning("No rows for tech='%s'.", BATTERY_TECH)
        return pd.DataFrame(columns=["hour", "island", "year", "soc_gwh"])

    bat["hour"] = (
        bat["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    bat.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "soc_gwh",
    }, inplace=True)

    bat = bat.groupby(["hour", "island", "year"], as_index=False)["soc_gwh"].sum()
    log.info("Battery rows: %d", len(bat))
    return bat[["hour", "island", "year", "soc_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# SOC MATRIX BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_soc_matrix(bat, island, year):
    mask = (bat["island"] == island) & (bat["year"] == year)
    sub  = bat.loc[mask, ["hour", "soc_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["soc_gwh"])

    peak_gwh = float(full.max())
    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    soc_pct        = (full / peak_gwh * 100).values
    matrix         = soc_pct.reshape(365, 24).T
    avg_soc        = float(soc_pct.mean())
    hours_above_80 = float((soc_pct > 80).mean() * 100)

    return matrix, peak_gwh, avg_soc, hours_above_80


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def draw_soc_heatmap(ax, matrix, year, peak_gwh, avg_soc,
                     hours_above_80, show_ylab):

    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, "No battery data", ha="center", va="center",
                transform=ax.transAxes, fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix,
        aspect="auto",
        origin="upper",
        cmap=SOC_CMAP,
        vmin=0, vmax=100,
        interpolation="nearest",
        extent=[1, 365, 23.5, -0.5],
    )

    # ── X-axis: month names, no overlapping day numbers ───────────────────────
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)          # labels only
    ax.tick_params(axis="x", which="minor", length=3,
                   color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)

    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    # ── Y-axis ────────────────────────────────────────────────────────────────
    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00", "06:00", "12:00", "18:00", "23:00"],
                       fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    # ── Title + stats ─────────────────────────────────────────────────────────
    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.3f} GWh  |  "
         f"Avg SOC: {avg_soc:.1f}%  |  "
         f"Hours >80%: {hours_above_80:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )

    return im


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island_soc(bat, island, years, out_dir, fmt, dpi):

    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5                              # inches per heatmap panel
    cbar_w  = 0.3                              # inches for colorbar column
    gap     = 0.55                             # right-side whitespace (inches)
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7   # 0.7 for left margin
    fig_h   = 5.2

    fig = plt.figure(figsize=(fig_w, fig_h))

    # right edge fraction: leaves exactly (cbar_w + gap) inches for cbar+space
    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06,
        left=0.07,
        right=right_frac,
        top=0.80,
        bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])

    for ax in axes[1:]:
        ax.sharey(axes[0])

    fig.suptitle(
        f"{island_label} — Battery State of Charge (% of annual peak)",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_soc, h80 = build_soc_matrix(bat, island, year)
        im = draw_soc_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_soc, h80,
            show_ylab=(idx == 0),
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    # ── Colorbar anchored to its own dedicated axis ───────────────────────────
    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("State of Charge (%)", fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)      # no bbox_inches="tight"
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path, out_dir, fmt, dpi, islands, years, preloaded_data=None):
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_gdx(gdx_path)

    bat  = extract_battery_soc(data)

    if bat.empty:
        log.error("No battery data extracted.")
        return

    log.info("Islands: %s", sorted(bat["island"].unique()))
    log.info("Years  : %s", sorted(bat["year"].unique()))

    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island_soc(bat, island, years, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island SOC figures written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--out",     default=OUTPUT_DIR)
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+", default=YEARS)
    parser.add_argument("--tech",    default=BATTERY_TECH)
    args = parser.parse_args()

    BATTERY_TECH = args.tech
    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
####################thermal storage SOC##################################################################
import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/thermal_soc_minload"
YEARS      = ["2030", "2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

STORAGE_TECH = "THSS"   # Thermal Heat Storage System

MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# Warm orange-red palette — visually distinct from the blue battery SOC map
SOC_CMAP = LinearSegmentedColormap.from_list(
    "thermal_soc",
    ["#FFFFB2", "#FECC5C", "#FD8D3C", "#E31A1C", "#800026"],
    N=256,
)

rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
})


def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_storage_soc(data: dict) -> pd.DataFrame:
    sl   = data["storage_level_out"]
    thss = sl[sl["techs"] == STORAGE_TECH].copy()   # commodity filter not needed

    if thss.empty:
        log.warning("No rows for tech='%s'.", STORAGE_TECH)
        return pd.DataFrame(columns=["hour", "island", "year", "soc_gwh"])

    thss["hour"] = (
        thss["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    thss.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "soc_gwh",
    }, inplace=True)

    thss = thss.groupby(["hour", "island", "year"], as_index=False)["soc_gwh"].sum()
    log.info("Thermal storage rows: %d", len(thss))
    return thss[["hour", "island", "year", "soc_gwh"]]


def build_soc_matrix(soc_df, island, year):
    mask = (soc_df["island"] == island) & (soc_df["year"] == year)
    sub  = soc_df.loc[mask, ["hour", "soc_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["soc_gwh"])

    peak_gwh = float(full.max())
    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    soc_pct        = (full / peak_gwh * 100).values
    matrix         = soc_pct.reshape(365, 24).T
    avg_soc        = float(soc_pct.mean())
    hours_above_80 = float((soc_pct > 80).mean() * 100)

    return matrix, peak_gwh, avg_soc, hours_above_80


def draw_soc_heatmap(ax, matrix, year, peak_gwh, avg_soc,
                     hours_above_80, show_ylab):
    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, "No thermal storage data",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix, aspect="auto", origin="upper",
        cmap=SOC_CMAP, vmin=0, vmax=100,
        interpolation="nearest",
        extent=[1, 365, 23.5, -0.5],
    )

    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)
    ax.tick_params(axis="x", which="minor", length=3, color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)

    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00","06:00","12:00","18:00","23:00"], fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.3f} GWh  |  "
         f"Avg SOC: {avg_soc:.1f}%  |  "
         f"Hours >80%: {hours_above_80:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )
    return im


def plot_island_soc(soc_df, island, years, out_dir, fmt, dpi):
    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5
    cbar_w  = 0.3
    gap     = 0.55
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7
    fig_h   = 5.2

    fig        = plt.figure(figsize=(fig_w, fig_h))
    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06, left=0.07, right=right_frac, top=0.80, bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])

    for ax in axes[1:]:
        ax.sharey(axes[0])

    fig.suptitle(
        f"{island_label} — Thermal Storage State of Charge (% of annual peak)",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_soc, h80 = build_soc_matrix(soc_df, island, year)
        im = draw_soc_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_soc, h80,
            show_ylab=(idx == 0),
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("State of Charge (%)", fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


def main(gdx_path, out_dir, fmt, dpi, islands, years, preloaded_data=None):
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     log.info("Using pre-loaded GDX data — skipping file read.")
    #     data = preloaded_data
    # else:
    #     data = load_gdx(gdx_path)

    soc_df = extract_storage_soc(data)

    if soc_df.empty:
        log.error("No thermal storage data extracted.")
        return

    log.info("Islands: %s", sorted(soc_df["island"].unique()))
    log.info("Years  : %s", sorted(soc_df["year"].unique()))

    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island_soc(soc_df, island, years, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island thermal SOC figures written to: {out_path}/")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--out",     default=OUTPUT_DIR)
    parser.add_argument("--fmt",     default="png", choices=["png","pdf","svg","tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+", default=YEARS)
    args = parser.parse_args()

    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
##########################################battery capacity and their generation######################


"""
battery_viz.py
--------------
For each of the 14 Pacific island models, produces a dual-axis 6-bar chart:
  LEFT  3 bars  — Battery storage capacity (GWh)   [2030, 2040, 2050]
  RIGHT 3 bars  — Power generation from battery (GWh) [2030, 2040, 2050]
  A dotted vertical line separates the two groups.

  Left  y-axis → capacity scale
  Right y-axis → generation scale

Source tables:
  storage_caps         → techs=Battery, capType=total, commodity=Elec_LiIon, +ve
  storage_flows_annual → techs=Battery, commodity=Elec_LiIon, balanceType=positive, +ve

Usage:
    python battery_viz.py --gdx path/to/results.gdx --out battery_charts/
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ── Configuration ─────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/battery_charts"

YEARS = ["2030", "2040", "2050"]

ISLANDS = [
    "CI_model", "FJ_model", "FSM_model", "KB_model", "MI_model",
    "NU_model", "NE_model", "PU_model", "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model", "VU_model",
]

ISLAND_LABELS = {
    "CI_model":  "Cook Islands",
    "FJ_model":  "Fiji",
    "FSM_model": "Federated States of Micronesia",
    "KB_model":  "Kiribati",
    "MI_model":  "Marshall Islands",
    "NU_model":  "Nauru",
    "NE_model":  "Niue",
    "PU_model":  "Palau",
    "PNG_model": "Papua New Guinea",
    "SA_model":  "Samoa",
    "SI_model":  "Solomon Islands",
    "TA_model":  "Tonga",
    "TU_model":  "Tuvalu",
    "VU_model":  "Vanuatu",
}

# Capacity bars (left axis) — solid colours
CAP_COLORS  = {"2030": "#4472C4", "2040": "#ED7D31", "2050": "#70AD47"}
# Generation bars (right axis) — hatched, slightly lighter tones
GEN_COLORS  = {"2030": "#A9C0E8", "2040": "#F7C09A", "2050": "#B8DDA0"}

BAR_W = 0.28   # width of every bar

# X-positions: capacity group centred at 0, generation group centred at 1
# Each group has 3 bars offset by BAR_W around the group centre
GROUP_GAP = 1.0          # distance between the two group centres
CAP_CTR   = 0.0
GEN_CTR   = CAP_CTR + GROUP_GAP

OFFSETS = {              # offset within a group for each year
    "2030": -BAR_W,
    "2040":  0.0,
    "2050":  BAR_W,
}

CAP_POSITIONS  = {y: CAP_CTR + OFFSETS[y] for y in YEARS}
GEN_POSITIONS  = {y: GEN_CTR + OFFSETS[y] for y in YEARS}

# Dotted separator x position — midway between the two groups
SEP_X = (CAP_CTR + BAR_W/2 + GEN_CTR - BAR_W/2) / 2   # ≈ 0.5

# Unit conversion: storage_caps values come out of the GDX in GWh.
# Capacities are displayed in MWh, generation stays in GWh.
CAP_UNIT_SCALE = 1000.0   # GWh -> MWh


# ── Data loading ──────────────────────────────────────────────────────────────

def load_data(gdx_path: str):
    log.info("Loading GDX: %s", gdx_path)
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)

    caps = data["storage_caps"].copy()
    log.info("storage_caps columns: %s", caps.columns.tolist())
    caps = caps[
        (caps["techs"]        == "Battery") &
        (caps["capType"]      == "total") &
        (caps["commodity"]    == "Elec_LiIon") &
        (caps["accYears"].isin(YEARS)) &
        (caps["accNodesModel"].isin(ISLANDS)) &
        (caps["Value"]        >  0)
    ].copy()
    log.info("storage_caps after filter: %d rows", len(caps))

    flows = data["storage_flows_annual"].copy()
    log.info("storage_flows_annual columns: %s", flows.columns.tolist())
    flows = flows[
        (flows["techs"]        == "Battery") &
        (flows["commodity"]    == "Elec_LiIon") &
        (flows["balanceType"]  == "positive") &
        (flows["accYears"].isin(YEARS)) &
        (flows["accNodesModel"].isin(ISLANDS)) &
        (flows["Value"]        >  0)
    ].copy()
    log.info("storage_flows_annual after filter: %d rows", len(flows))

    return caps, flows


def get_value(df, island, year):
    mask = (df["accNodesModel"] == island) & (df["accYears"] == year)
    return float(df.loc[mask, "Value"].sum())


# ── Single-island plot ────────────────────────────────────────────────────────

def plot_island(island, caps, flows, output_dir):
    label    = ISLAND_LABELS.get(island, island)
    cap_vals = {y: get_value(caps,  island, y) * CAP_UNIT_SCALE for y in YEARS}
    gen_vals = {y: get_value(flows, island, y) for y in YEARS}

    if all(v == 0 for v in cap_vals.values()) and all(v == 0 for v in gen_vals.values()):
        log.info("No battery data for %s — skipping.", label)
        return

    fig, ax1 = plt.subplots(figsize=(9, 5))
    ax2 = ax1.twinx()

    cap_max  = max(cap_vals.values())  or 1
    gen_max  = max(gen_vals.values())  or 1

    # ── 3 capacity bars (left axis) ───────────────────────────────────────────
    for year in YEARS:
        x = CAP_POSITIONS[year]
        v = cap_vals[year]
        ax1.bar(x, v, width=BAR_W * 0.90,
                color=CAP_COLORS[year], alpha=0.88, zorder=3,
                label=f"Capacity {year}")
        if v > 0:
            ax1.text(x, v + cap_max * 0.025, f"{v:,.1f}",
                     ha="center", va="bottom", fontsize=7.5,
                     color=CAP_COLORS[year], fontweight="bold")

    # ── 3 generation bars (right axis) ────────────────────────────────────────
    for year in YEARS:
        x = GEN_POSITIONS[year]
        v = gen_vals[year]
        ax2.bar(x, v, width=BAR_W * 0.90,
                color=GEN_COLORS[year], alpha=0.88, zorder=3,
                hatch="//", edgecolor="grey", linewidth=0.5,
                label=f"Generation {year}")
        if v > 0:
            ax2.text(x, v + gen_max * 0.025, f"{v:.1f}",
                     ha="center", va="bottom", fontsize=7.5,
                     color="dimgrey", fontweight="bold")

    # ── Dotted separator ──────────────────────────────────────────────────────
    ax1.axvline(x=SEP_X, color="grey", linestyle=":", linewidth=1.5, alpha=0.8, zorder=5)

    # ── Group header labels ───────────────────────────────────────────────────
    y_header = cap_max * 1.18
    ax1.text(CAP_CTR, y_header, "Battery Capacity",
             ha="center", va="bottom", fontsize=9, color="#2F4F8F",
             fontweight="bold", transform=ax1.transData)
    # For generation label use ax1 transform but position on gen side
    ax1.text(GEN_CTR, y_header, "Power Generation",
             ha="center", va="bottom", fontsize=9, color="#555555",
             fontweight="bold", transform=ax1.transData)

    # ── Year tick labels (shared x-axis) ─────────────────────────────────────
    tick_positions = (
        [CAP_POSITIONS[y] for y in YEARS] +
        [GEN_POSITIONS[y] for y in YEARS]
    )
    tick_labels = YEARS + YEARS
    ax1.set_xticks(tick_positions)
    ax1.set_xticklabels(tick_labels, fontsize=8, color="dimgrey")

    # ── Axis limits & labels ──────────────────────────────────────────────────
    ax1.set_xlim(CAP_CTR - BAR_W * 2, GEN_CTR + BAR_W * 2)
    ax1.set_ylim(0, cap_max * 1.30)
    ax2.set_ylim(0, gen_max * 1.30)

    ax1.set_ylabel("Battery Storage Capacity (MWh)", fontsize=10, color="#2F4F8F")
    ax2.set_ylabel("Power Output from Battery (GWh)", fontsize=10, color="#555555")
    ax1.tick_params(axis="y", labelcolor="#2F4F8F")
    ax2.tick_params(axis="y", labelcolor="#555555")

    ax1.set_title(f"{label}",
                  fontsize=12, fontweight="bold", pad=12)

    # ── Legend ────────────────────────────────────────────────────────────────
    cap_patches = [mpatches.Patch(facecolor=CAP_COLORS[y], label=f"Capacity {y}",
                                  alpha=0.88) for y in YEARS]
    gen_patches = [mpatches.Patch(facecolor=GEN_COLORS[y], label=f"Generation {y}",
                                  alpha=0.88, hatch="//", edgecolor="grey") for y in YEARS]
    ax1.legend(handles=cap_patches + gen_patches,
               loc="upper left", fontsize=7.5, framealpha=0.8, ncol=2)

    ax1.yaxis.grid(True, linestyle="--", alpha=0.35, zorder=0)
    ax1.set_axisbelow(True)

    fig.tight_layout()
    out_file = output_dir / f"battery_{island}.png"
    fig.savefig(out_file, dpi=150, bbox_inches="tight")
    plt.close(fig)
    log.info("Saved: %s", out_file)


# ── Summary: all 14 islands, 4×4 subplots ────────────────────────────────────

def plot_all_islands_summary(caps, flows, output_dir):
    fig, axes = plt.subplots(4, 4, figsize=(22, 18))
    axes_flat = axes.flatten()

    for idx, island in enumerate(ISLANDS):
        ax1 = axes_flat[idx]
        ax2 = ax1.twinx()
        label    = ISLAND_LABELS.get(island, island)
        cap_vals = {y: get_value(caps,  island, y) * CAP_UNIT_SCALE for y in YEARS}
        gen_vals = {y: get_value(flows, island, y) for y in YEARS}

        cap_max = max(cap_vals.values()) or 1
        gen_max = max(gen_vals.values()) or 1

        for year in YEARS:
            ax1.bar(CAP_POSITIONS[year], cap_vals[year], width=BAR_W * 0.88,
                    color=CAP_COLORS[year], alpha=0.88, zorder=3)
            ax2.bar(GEN_POSITIONS[year], gen_vals[year], width=BAR_W * 0.88,
                    color=GEN_COLORS[year], alpha=0.88, zorder=3,
                    hatch="//", edgecolor="grey", linewidth=0.4)

        ax1.axvline(x=SEP_X, color="grey", linestyle=":", linewidth=1.2, alpha=0.7)

        tick_positions = [CAP_POSITIONS[y] for y in YEARS] + [GEN_POSITIONS[y] for y in YEARS]
        tick_labels    = [y[-2:] for y in YEARS] + [y[-2:] for y in YEARS]
        ax1.set_xticks(tick_positions)
        ax1.set_xticklabels(tick_labels, fontsize=5.5, color="dimgrey")

        ax1.set_xlim(CAP_CTR - BAR_W * 2, GEN_CTR + BAR_W * 2)
        ax1.set_ylim(0, cap_max * 1.30)
        ax2.set_ylim(0, gen_max * 1.30)

        ax1.set_title(label, fontsize=7.5, fontweight="bold", pad=3)
        ax1.tick_params(axis="y", labelsize=5.5, labelcolor="#2F4F8F")
        ax2.tick_params(axis="y", labelsize=5.5, labelcolor="#555555")
        ax1.yaxis.grid(True, linestyle="--", alpha=0.3, zorder=0)
        ax1.set_axisbelow(True)

    for idx in range(len(ISLANDS), len(axes_flat)):
        axes_flat[idx].set_visible(False)

    # Shared legend
    cap_patches = [mpatches.Patch(facecolor=CAP_COLORS[y], label=f"Capacity {y}",
                                  alpha=0.88) for y in YEARS]
    gen_patches = [mpatches.Patch(facecolor=GEN_COLORS[y], label=f"Generation {y}",
                                  alpha=0.88, hatch="//", edgecolor="grey") for y in YEARS]
    fig.legend(handles=cap_patches + gen_patches,
               loc="lower center", ncol=6, fontsize=9,
               framealpha=0.85, bbox_to_anchor=(0.5, 0.005))

    fig.suptitle(
        "Battery Storage Capacity (solid bars, left axis, MWh)  |  "
        "Power Generation from Battery (hatched bars, right axis, GWh)\n"
        "Pacific Island Countries — 2030 / 2040 / 2050",
        fontsize=12, fontweight="bold", y=0.998)

    fig.tight_layout(rect=[0, 0.055, 1, 0.998])
    out_file = output_dir / "battery_ALL_ISLANDS_summary.png"
    fig.savefig(out_file, dpi=150, bbox_inches="tight")
    plt.close(fig)
    log.info("Summary chart saved: %s", out_file)


# ── Entry point ───────────────────────────────────────────────────────────────

def main(gdx_path, output_dir_str):
    output_dir = Path(output_dir_str)
    output_dir.mkdir(parents=True, exist_ok=True)
    caps, flows = load_data(gdx_path)
    for island in ISLANDS:
        plot_island(island, caps, flows, output_dir)
    plot_all_islands_summary(caps, flows, output_dir)
    log.info("All charts written to: %s", output_dir.resolve())


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gdx", default=GDX_PATH)
    parser.add_argument("--out", default=OUTPUT_DIR)
    args = parser.parse_args()
    main(args.gdx, args.out)
    ##################################H2_storage##########################################
"""
Pacific Island Countries (PICs) — Hydrogen Storage Level Heatmaps
==================================================================
Reads 'storage_level_out' from the GAMS GDX file and produces
24 x 365 heatmaps of hydrogen storage level expressed as a percentage
of the annual peak storage level.

    Storage level (%) = value / max(value for that island-year) x 100

    0 %   -> H2 storage empty
    100 % -> H2 storage at annual peak capacity

Layout
------
One figure per island  ->  1 x 3 row of subplots (2030 | 2040 | 2050).
Each subplot is a heatmap where:
    x-axis  : month labels (Jan-Dec)
    y-axis  : hour of day (0-23, midnight at top)
    colour  : storage level (%)

Source
------
    Symbol    : storage_level_out
    Tech      : H2_storage
    Commodity : Hydrogen_T

Output
------
  figures/hydrogen_storage/
      CI_model.png
      FJ_model.png
      ...  (one PNG per island)

Usage
-----
    python pic_h2_storage.py
    python pic_h2_storage.py --gdx path/to/results.gdx
                              --out figures/hydrogen_storage
                              --fmt png --dpi 200
                              --islands CI_model FJ_model
                              --years 2030 2040 2050
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH    = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR  = "figures/S_23/hydrogen_storage_minload"
YEARS       = ["2030", "2040", "2050"]
ISLANDS     = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

H2_TECH      = "H2_storage"
H2_COMMODITY = "Hydrogen_T"

# ── Month axis ────────────────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR MAP
# Warm yellow-green -> teal -> deep blue-green (same family as battery SOC
# but shifted toward green to visually distinguish H2 from battery plots)
# ─────────────────────────────────────────────────────────────────────────────
H2_CMAP = LinearSegmentedColormap.from_list(
    "h2_storage",
    ["#FFFFCC", "#C7E9B4", "#41B6C4", "#1D91C0", "#0C2C84"],
    N=256,
)

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# NOTE: savefig.bbox is intentionally NOT "tight" — margins are controlled
# explicitly via gridspec to prevent the colorbar from overlapping panels.
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
    # savefig.bbox deliberately omitted
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_h2_storage(data: dict) -> pd.DataFrame:
    """
    Pull storage_level_out, filter to H2_storage tech + Hydrogen_T commodity.

    Columns returned:
        hour   (int 0-8759)
        island (str)
        year   (str)
        h2_gwh (float, GWh)
    """
    sl   = data["storage_level_out"]
    mask = (sl["techs"] == H2_TECH) & (sl["commodity"] == H2_COMMODITY)
    h2   = sl.loc[mask].copy()

    if h2.empty:
        log.warning("No rows for tech='%s', commodity='%s'.", H2_TECH, H2_COMMODITY)
        return pd.DataFrame(columns=["hour", "island", "year", "h2_gwh"])

    h2["hour"] = (
        h2["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    h2.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "h2_gwh",
    }, inplace=True)

    h2 = h2.groupby(["hour", "island", "year"], as_index=False)["h2_gwh"].sum()

    log.info("H2 storage rows: %d", len(h2))
    log.info("Islands: %s", sorted(h2["island"].unique()))
    log.info("Years  : %s", sorted(h2["year"].unique()))

    return h2[["hour", "island", "year", "h2_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# STORAGE LEVEL MATRIX BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_h2_matrix(h2: pd.DataFrame, island: str, year: str):
    mask = (h2["island"] == island) & (h2["year"] == year)
    sub  = h2.loc[mask, ["hour", "h2_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["h2_gwh"])

    peak_gwh = float(full.max())
    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    level_pct      = (full / peak_gwh * 100).values
    matrix         = level_pct.reshape(365, 24).T
    avg_level      = float(level_pct.mean())
    hours_above_80 = float((level_pct > 80).mean() * 100)

    return matrix, peak_gwh, avg_level, hours_above_80


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def draw_h2_heatmap(ax, matrix: np.ndarray, year: str,
                    peak_gwh: float, avg_level: float,
                    hours_above_80: float, show_ylab: bool):

    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, "No H2 storage data\n(not deployed)",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix,
        aspect="auto",
        origin="upper",
        cmap=H2_CMAP,
        vmin=0, vmax=100,
        interpolation="nearest",
        extent=[1, 365, 23.5, -0.5],
    )

    # ── X-axis: month names, no overlapping day numbers ───────────────────────
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)
    ax.tick_params(axis="x", which="minor", length=3, color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)

    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    # ── Y-axis ────────────────────────────────────────────────────────────────
    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00", "06:00", "12:00", "18:00", "23:00"], fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    # ── Title + stats ─────────────────────────────────────────────────────────
    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.3f} GWh  |  "
         f"Avg level: {avg_level:.1f}%  |  "
         f"Hours >80%: {hours_above_80:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )

    return im


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island_h2(h2: pd.DataFrame, island: str,
                   years: list, out_dir: Path, fmt: str, dpi: int):

    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5
    cbar_w  = 0.3
    gap     = 0.55
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7
    fig_h   = 5.2

    fig = plt.figure(figsize=(fig_w, fig_h))

    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06,
        left=0.07,
        right=right_frac,
        top=0.80,
        bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])

    for ax in axes[1:]:
        ax.sharey(axes[0])

    fig.suptitle(
        f"{island_label} — Hydrogen Storage Level (% of annual peak)",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_level, h80 = build_h2_matrix(h2, island, year)
        im = draw_h2_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_level, h80,
            show_ylab=(idx == 0),
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    # ── Colorbar anchored to its own dedicated axis ───────────────────────────
    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("Storage level (%)", fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)      # no bbox_inches="tight"
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)
    h2   = extract_h2_storage(data)

    if h2.empty:
        log.error("No H2 storage data. Check tech='%s', commodity='%s'.",
                  H2_TECH, H2_COMMODITY)
        return

    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island_h2(h2, island, years, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island H2 storage figures written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",       default=GDX_PATH)
    parser.add_argument("--out",       default=OUTPUT_DIR)
    parser.add_argument("--fmt",       default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",       default=200, type=int)
    parser.add_argument("--islands",   nargs="+", default=ISLANDS)
    parser.add_argument("--years",     nargs="+", default=YEARS)
    parser.add_argument("--tech",      default=H2_TECH)
    parser.add_argument("--commodity", default=H2_COMMODITY)
    args = parser.parse_args()

    H2_TECH      = args.tech
    H2_COMMODITY = args.commodity

    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)

##################################Sustainable fuel storage SOC##########################################
##################################Sustainable fuel storage SOC##########################################
"""
Pacific Island Countries (PICs) — Sustainable-Fuel Storage Level Heatmaps
=========================================================================
Mirrors the Hydrogen storage heatmaps, applied to the three synthetic-fuel
storages defined in TECHS:

    Ammonia    (tech: Ammonia_storage,   commodity: Ammonia_T)
    Methanol   (tech: Methanol_storage,  commodity: Methanol_T)
    eKerosene  (tech: eKerosene_storage, commodity: eKerosene_T)

    Storage level (%) = value / max(value for that island-year) x 100
        0 %   -> store empty
        100 % -> store at annual peak

Layout (identical to the H2 figures)
------------------------------------
One figure per island per fuel -> 1 x 3 row of subplots (2030 | 2040 | 2050).
    x-axis : month labels (Jan-Dec)
    y-axis : hour of day (0-23, midnight at top)
    colour : storage level (%)

Source
------
    Symbol : storage_level_out
    Tech / commodity per fuel (see FUEL_STORAGES below)

Output
------
  figures/S_23/ammonia_storage_minload/   CI_model.png ...
  figures/S_23/methanol_storage_minload/  ...
  figures/S_23/ekerosene_storage_minload/ ...
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S1_minload.gdx"
YEARS    = ["2030", "2040", "2050"]
ISLANDS  = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

# ── Per-fuel configuration ────────────────────────────────────────────────────
# Each fuel gets its own tech, commodity, output folder and (distinct) colormap.
# `unit` is only used for the "Peak: … " annotation; storage_level_out is assumed
# to be in GWh (energy), matching the H2 figures — change if your model differs.
FUEL_STORAGES = [
    {
        "name":      "Ammonia",
        "tech":      "Ammonia_storage",
        "commodity": "Ammonia_T",
        "out_dir":   "figures/S_1/ammonia_storage_minload",
        "unit":      "GWh",
        "cmap":      LinearSegmentedColormap.from_list(
                         "ammonia_storage",
                         ["#FFFFE5", "#D9F0A3", "#78C679", "#238443", "#004529"],
                         N=256),
    },
    {
        "name":      "Methanol",
        "tech":      "Methanol_storage",
        "commodity": "Methanol_T",
        "out_dir":   "figures/S_1/methanol_storage_minload",
        "unit":      "GWh",
        "cmap":      LinearSegmentedColormap.from_list(
                         "methanol_storage",
                         ["#FFF7EC", "#FDD49E", "#FC8D59", "#D7301F", "#7F0000"],
                         N=256),
    },
    {
        "name":      "eKerosene",
        "tech":      "eKerosene_storage",
        "commodity": "eKerosene_T",
        "out_dir":   "figures/S_1/ekerosene_storage_minload",
        "unit":      "GWh",
        "cmap":      LinearSegmentedColormap.from_list(
                         "ekerosene_storage",
                         ["#F7FCFD", "#BFD3E6", "#8C96C6", "#88419D", "#4D004B"],
                         N=256),
    },
]

# ── Month axis ────────────────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE  (savefig.bbox intentionally NOT "tight" — margins via gridspec)
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA EXTRACTION  (one fuel at a time)
# ─────────────────────────────────────────────────────────────────────────────
def extract_fuel_storage(data: dict, tech: str, commodity: str) -> pd.DataFrame:
    """storage_level_out filtered to one storage tech + commodity.

    Returns columns: hour (0-8759), island, year, level_gwh.
    """
    sl   = data["storage_level_out"]
    mask = (sl["techs"] == tech) & (sl["commodity"] == commodity)
    df   = sl.loc[mask].copy()

    if df.empty:
        log.warning("No rows for tech='%s', commodity='%s'.", tech, commodity)
        return pd.DataFrame(columns=["hour", "island", "year", "level_gwh"])

    df["hour"] = (
        df["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "level_gwh",
    }, inplace=True)

    df = df.groupby(["hour", "island", "year"], as_index=False)["level_gwh"].sum()

    log.info("[%s] rows: %d | islands: %s | years: %s",
             tech, len(df), sorted(df["island"].unique()),
             sorted(df["year"].unique()))
    return df[["hour", "island", "year", "level_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# STORAGE-LEVEL MATRIX BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def build_fuel_matrix(df: pd.DataFrame, island: str, year: str):
    mask = (df["island"] == island) & (df["year"] == year)
    sub  = df.loc[mask, ["hour", "level_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["level_gwh"])

    peak_gwh = float(full.max())
    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0

    level_pct      = (full / peak_gwh * 100).values
    matrix         = level_pct.reshape(365, 24).T
    avg_level      = float(level_pct.mean())
    hours_above_80 = float((level_pct > 80).mean() * 100)

    return matrix, peak_gwh, avg_level, hours_above_80


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────
def draw_fuel_heatmap(ax, matrix, year, peak_gwh, avg_level, hours_above_80,
                      show_ylab, cmap, fuel_name, unit):

    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, f"No {fuel_name} storage data\n(not deployed)",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix, aspect="auto", origin="upper",
        cmap=cmap, vmin=0, vmax=100,
        interpolation="nearest", extent=[1, 365, 23.5, -0.5],
    )

    # ── X-axis: month names ───────────────────────────────────────────────────
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)
    ax.tick_params(axis="x", which="minor", length=3, color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)
    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    # ── Y-axis ────────────────────────────────────────────────────────────────
    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00", "06:00", "12:00", "18:00", "23:00"], fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    # ── Title + stats ─────────────────────────────────────────────────────────
    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.3f} {unit}  |  "
         f"Avg level: {avg_level:.1f}%  |  "
         f"Hours >80%: {hours_above_80:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )
    return im


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────
def plot_island_fuel(df, island, years, out_dir, fmt, dpi, cmap, fuel_name, unit):
    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5
    cbar_w  = 0.3
    gap     = 0.55
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7
    fig_h   = 5.2

    fig = plt.figure(figsize=(fig_w, fig_h))
    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06, left=0.07, right=right_frac, top=0.80, bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])
    for ax in axes[1:]:
        ax.sharey(axes[0])

    fig.suptitle(
        f"{island_label} — {fuel_name} Storage Level (% of annual peak)",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_level, h80 = build_fuel_matrix(df, island, year)
        im = draw_fuel_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_level, h80,
            show_ylab=(idx == 0), cmap=cmap, fuel_name=fuel_name, unit=unit,
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("Storage level (%)", fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)      # no bbox_inches="tight"
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN — loops over the three fuels
# ─────────────────────────────────────────────────────────────────────────────
def main(gdx_path: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    # Uses the shared global `data` loaded once at the top of Gdx_plot.py
    # (same pattern as the H2 storage block).
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)

    for fuel in FUEL_STORAGES:
        log.info("=== %s storage (tech=%s, commodity=%s) ===",
                 fuel["name"], fuel["tech"], fuel["commodity"])

        df = extract_fuel_storage(data, fuel["tech"], fuel["commodity"])
        if df.empty:
            log.error("No %s storage data — skipping. "
                      "Check tech='%s', commodity='%s'.",
                      fuel["name"], fuel["tech"], fuel["commodity"])
            continue

        out_path = Path(fuel["out_dir"])
        out_path.mkdir(parents=True, exist_ok=True)

        for island in islands:
            log.info("Plotting %s %s ...", island, fuel["name"])
            plot_island_fuel(df, island, years, out_path, fmt, dpi,
                             fuel["cmap"], fuel["name"], fuel["unit"])

        print(f"All {len(islands)} island {fuel['name']} storage figures "
              f"written to: {out_path}/")

    log.info("Done — all sustainable-fuel storage figures generated.")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+", default=YEARS)
    args = parser.parse_args()

    main(args.gdx, args.fmt, args.dpi, args.islands, args.years)
    ######################################Efuel synthesis activity##########################
######################e-Fuel synthesis activity####################################
"""
Pacific Island Countries (PICs) — e-Fuel Synthesis Activity Heatmaps
====================================================================
24 x 365 heatmaps of synthetic-fuel synthesis activity, expressed as a
percentage of the annual peak hourly production.  Flows are read from
'commodity_balance'.

    Activity (%) = flow(hour) / max(flow over the year) x 100
        0 %   -> converter off
        100 % -> converter at peak throughput

Converters (see CONVERTERS below) — output flow, Value > 0
----------------------------------------------------------
    Ammonia_synthesis  -> Ammonia produced
    Methanol_synthesis -> Methanol produced
    FTL                -> eKerosene produced

NOTE: the synthesis commodities use the bare name (Ammonia / Methanol /
eKerosene) — no "_T" suffix.

Layout
------
One figure per island per converter -> 1 x 2 row (2040 | 2050).
    x-axis : month labels (Jan-Dec)
    y-axis : hour of day (0-23, 00:00 at top)
    colour : activity (%)

Output
------
  figures/S_1/ammonia_synthesis_activity_minload/   CI_model.png ...
  figures/S_1/methanol_synthesis_activity_minload/  ...
  figures/S_1/ekerosene_synthesis_activity_minload/ ...
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S1_minload.gdx"
YEARS    = ["2040", "2050"]
ISLANDS  = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

# ── Colour maps ───────────────────────────────────────────────────────────────
# Distinct hues per fuel (matching the fuel-storage figures) so they are easy
# to tell apart.
CMAP_AMMONIA = LinearSegmentedColormap.from_list(
    "ammonia_act", ["#FFFFE5", "#D9F0A3", "#78C679", "#238443", "#004529"], N=256)
CMAP_METHANOL = LinearSegmentedColormap.from_list(
    "methanol_act", ["#FFF7EC", "#FDD49E", "#FC8D59", "#D7301F", "#7F0000"], N=256)
CMAP_EKEROSENE = LinearSegmentedColormap.from_list(
    "ekerosene_act", ["#F7FCFD", "#BFD3E6", "#8C96C6", "#88419D", "#4D004B"], N=256)

# ── Per-converter configuration ───────────────────────────────────────────────
# flow: "in"  -> electricity/feedstock consumed (filter Value < 0, take abs)
#       "out" -> product generated              (filter Value > 0)
CONVERTERS = [
    {
        "name":       "Ammonia synthesis",
        "tech":       "Ammonia_synthesis",
        "commodity":  "Ammonia",
        "flow":       "out",
        "out_dir":    "figures/S_1/ammonia_synthesis_activity_minload",
        "cmap":       CMAP_AMMONIA,
        "cbar_label": "Ammonia synthesis activity (%)",
    },
    {
        "name":       "Methanol synthesis",
        "tech":       "Methanol_synthesis",
        "commodity":  "Methanol",
        "flow":       "out",
        "out_dir":    "figures/S_1/methanol_synthesis_activity_minload",
        "cmap":       CMAP_METHANOL,
        "cbar_label": "Methanol synthesis activity (%)",
    },
    {
        "name":       "eKerosene synthesis (FTL)",
        "tech":       "FTL",
        "commodity":  "eKerosene",
        "flow":       "out",
        "out_dir":    "figures/S_1/ekerosene_synthesis_activity_minload",
        "cmap":       CMAP_EKEROSENE,
        "cbar_label": "eKerosene synthesis activity (%)",
    },
]

# ── Month axis ────────────────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE  (savefig.bbox intentionally NOT "tight")
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA EXTRACTION  (one converter at a time)
# ─────────────────────────────────────────────────────────────────────────────
def extract_converter_flow(data: dict, tech: str, commodity: str,
                           flow: str) -> pd.DataFrame:
    """commodity_balance filtered to one tech + commodity + flow direction.

    flow="in"  -> consumption (Value < 0), returned as +ve magnitude
    flow="out" -> production  (Value > 0)

    Returns columns: hour (0-8759), island, year, value_gwh (>= 0).
    """
    cb = data["commodity_balance"]

    if flow == "in":
        mask = ((cb["techs"] == tech) & (cb["commodity"] == commodity) &
                (cb["Value"] < 0))
    else:  # "out"
        mask = ((cb["techs"] == tech) & (cb["commodity"] == commodity) &
                (cb["Value"] > 0))

    df = cb.loc[mask].copy()

    if df.empty:
        log.warning("No rows for tech='%s', commodity='%s', flow='%s'.",
                    tech, commodity, flow)
        return pd.DataFrame(columns=["hour", "island", "year", "value_gwh"])

    df["hour"] = (
        df["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "value_gwh",
    }, inplace=True)

    df["value_gwh"] = df["value_gwh"].abs()
    df = df.groupby(["hour", "island", "year"], as_index=False)["value_gwh"].sum()

    log.info("[%s] rows: %d | islands: %s | years: %s",
             tech, len(df), sorted(df["island"].unique()),
             sorted(df["year"].unique()))
    return df[["hour", "island", "year", "value_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# ACTIVITY MATRIX BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def build_activity_matrix(df: pd.DataFrame, island: str, year: str):
    mask = (df["island"] == island) & (df["year"] == year)
    sub  = df.loc[mask, ["hour", "value_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["value_gwh"])

    peak_gwh  = float(full.max())
    total_gwh = float(full.sum())
    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0, 0.0

    activity_pct = (full / peak_gwh * 100).values
    matrix       = activity_pct.reshape(365, 24).T
    avg_activity = float(activity_pct.mean())
    hours_active = float((activity_pct > 0).mean() * 100)

    return matrix, peak_gwh, avg_activity, hours_active, total_gwh


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────
def draw_activity_heatmap(ax, matrix, year, peak_gwh, avg_activity,
                          hours_active, total_gwh, show_ylab, cmap, name):

    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, f"No {name} data\n(not deployed)",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix, aspect="auto", origin="upper",
        cmap=cmap, vmin=0, vmax=100,
        interpolation="nearest", extent=[1, 365, 23.5, -0.5],
    )

    # ── X-axis: month names ───────────────────────────────────────────────────
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)
    ax.tick_params(axis="x", which="minor", length=3, color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)
    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    # ── Y-axis ────────────────────────────────────────────────────────────────
    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00", "06:00", "12:00", "18:00", "23:00"], fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    # ── Title + stats ─────────────────────────────────────────────────────────
    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.4f} GWh/h  |  "
         f"Total: {total_gwh:.2f} GWh/yr  |  "
         f"Avg activity: {avg_activity:.1f}%  |  "
         f"Hours active: {hours_active:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )
    return im


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────
def plot_island_converter(df, island, years, out_dir, fmt, dpi,
                          cmap, name, cbar_label, flow):
    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5
    cbar_w  = 0.3
    gap     = 0.55
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7
    fig_h   = 5.2

    fig = plt.figure(figsize=(fig_w, fig_h))
    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06, left=0.07, right=right_frac, top=0.80, bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])
    for ax in axes[1:]:
        ax.sharey(axes[0])

    flow_word = "consumption" if flow == "in" else "production"
    fig.suptitle(
        f"{island_label} — {name} Activity  "
        f"(% of annual peak hourly {flow_word})",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_act, h_active, total_gwh = \
            build_activity_matrix(df, island, year)
        im = draw_activity_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_act, h_active, total_gwh,
            show_ylab=(idx == 0), cmap=cmap, name=name,
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label(cbar_label, fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)      # no bbox_inches="tight"
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN — loops over the four converters
# ─────────────────────────────────────────────────────────────────────────────
def main(gdx_path: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    # Uses the shared global `data` loaded once at the top of Gdx_plot.py
    # (same pattern as the original activity block).
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)

    for conv in CONVERTERS:
        log.info("=== %s (tech=%s, commodity=%s, flow=%s) ===",
                 conv["name"], conv["tech"], conv["commodity"], conv["flow"])

        df = extract_converter_flow(
            data, conv["tech"], conv["commodity"], conv["flow"])
        if df.empty:
            log.error("No %s data — skipping. Check tech='%s', commodity='%s'.",
                      conv["name"], conv["tech"], conv["commodity"])
            continue

        out_path = Path(conv["out_dir"])
        out_path.mkdir(parents=True, exist_ok=True)

        for island in islands:
            log.info("Plotting %s %s ...", island, conv["name"])
            plot_island_converter(df, island, years, out_path, fmt, dpi,
                                   conv["cmap"], conv["name"],
                                   conv["cbar_label"], conv["flow"])

        print(f"All {len(islands)} island {conv['name']} figures "
              f"written to: {out_path}/")

    log.info("Done — all converter activity figures generated.")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+", default=YEARS)
    args = parser.parse_args()

    main(args.gdx, args.fmt, args.dpi, args.islands, args.years)
    
    ######################AEL activity####################################################
"""
Pacific Island Countries (PICs) — Electrolyzer (AEL) Activity Heatmaps
=======================================================================
Reads hourly electricity consumption of the AEL (alkaline electrolyser)
from 'commodity_balance' in the GAMS GDX file and produces 24 × 365
heatmaps of electrolyser activity expressed as a percentage of the
annual peak consumption.

    Activity (%) = |AEL Elec consumption (hour)| /
                   max(|AEL Elec consumption|) over the year  × 100

    0 %   → electrolyser off
    100 % → electrolyser running at peak rated load

Layout
------
One figure per island  →  1 × 2 row of subplots (2040 | 2050).
Each subplot is a heatmap where:
    x-axis  : month labels (Jan-Dec)
    y-axis  : hour of day (0-23,  00:00 at top)
    colour  : activity (%)

Output
------
  figures/electrolyzer_activity/
      CI_model.png
      FJ_model.png
      ...  (one PNG per island)

Usage
-----
    python pic_electrolyzer_activity.py
    python pic_electrolyzer_activity.py --gdx path/to/results.gdx
                                        --out figures/electrolyzer_activity
                                        --fmt png --dpi 200
                                        --islands CI_model FJ_model
                                        --years 2040 2050
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/electrolyzer_activity_minload"
YEARS      = ["2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

AEL_TECH      = "AEL"
AEL_COMMODITY = "Elec"

# ── Month axis ────────────────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR MAP
# YlGnBu-inspired (matching battery SOC design):
# pale yellow (idle) → mint → teal → deep ocean blue (full load)
# Perceptually uniform, colorblind-safe, prints well in greyscale too.
# ─────────────────────────────────────────────────────────────────────────────
AEL_CMAP = LinearSegmentedColormap.from_list(
    "soc",
    ["#FFFFCC", "#C7E9B4", "#41B6C4", "#1D91C0", "#0C2C84"],
    N=256,
)

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# IMPORTANT: savefig.bbox is intentionally NOT "tight".
# "tight" re-expands the canvas after layout and drags the colorbar axis
# back over the rightmost heatmap panel.  All margins are controlled via
# gridspec left/right/top/bottom fractions instead.
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
    # savefig.bbox deliberately omitted (defaults to "standard")
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_ael_consumption(data: dict) -> pd.DataFrame:
    """
    Pull commodity_balance, filter to AEL tech + Elec commodity.
    Take absolute value of the negative consumption rows.

    Returns DataFrame with columns:
        hour      (int 0-8759)
        island    (str)
        year      (str)
        elec_gwh  (float, GWh, always >= 0)
    """
    cb   = data["commodity_balance"]
    mask = (
        (cb["techs"]     == AEL_TECH) &
        (cb["commodity"] == AEL_COMMODITY) &
        (cb["Value"]     <  0)
    )
    ael = cb.loc[mask].copy()

    if ael.empty:
        log.warning("No rows for tech='%s', commodity='%s', Value<0.",
                    AEL_TECH, AEL_COMMODITY)
        return pd.DataFrame(columns=["hour", "island", "year", "elec_gwh"])

    ael["hour"] = (
        ael["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    ael.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "elec_gwh",
    }, inplace=True)

    ael["elec_gwh"] = ael["elec_gwh"].abs()
    ael = ael.groupby(["hour", "island", "year"], as_index=False)["elec_gwh"].sum()

    log.info("AEL rows: %d", len(ael))
    log.info("Islands: %s", sorted(ael["island"].unique()))
    log.info("Years  : %s", sorted(ael["year"].unique()))

    return ael[["hour", "island", "year", "elec_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# ACTIVITY MATRIX BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_activity_matrix(ael: pd.DataFrame, island: str, year: str):
    mask = (ael["island"] == island) & (ael["year"] == year)
    sub  = ael.loc[mask, ["hour", "elec_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["elec_gwh"])

    peak_gwh  = float(full.max())
    total_gwh = float(full.sum())

    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0, 0.0

    activity_pct = (full / peak_gwh * 100).values
    matrix       = activity_pct.reshape(365, 24).T

    avg_activity = float(activity_pct.mean())
    hours_active = float((activity_pct > 0).mean() * 100)

    return matrix, peak_gwh, avg_activity, hours_active, total_gwh


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def draw_activity_heatmap(ax, matrix: np.ndarray, year: str,
                          peak_gwh: float, avg_activity: float,
                          hours_active: float, total_gwh: float,
                          show_ylab: bool):

    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, "No AEL data\n(electrolyser not deployed)",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix,
        aspect="auto",
        origin="upper",
        cmap=AEL_CMAP,
        vmin=0, vmax=100,
        interpolation="nearest",
        extent=[1, 365, 23.5, -0.5],
    )

    # ── X-axis: month names, no overlapping day numbers ───────────────────────
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)          # labels only
    ax.tick_params(axis="x", which="minor", length=3,
                   color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)

    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    # ── Y-axis ────────────────────────────────────────────────────────────────
    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00", "06:00", "12:00", "18:00", "23:00"],
                       fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    # ── Title + stats ─────────────────────────────────────────────────────────
    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.4f} GWh/h  |  "
         f"Total: {total_gwh:.2f} GWh/yr  |  "
         f"Avg activity: {avg_activity:.1f}%  |  "
         f"Hours active: {hours_active:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )

    return im


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island_ael(ael: pd.DataFrame, island: str,
                    years: list, out_dir: Path, fmt: str, dpi: int):

    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5                              # inches per heatmap panel
    cbar_w  = 0.3                              # inches for colorbar column
    gap     = 0.55                             # right-side whitespace (inches)
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7   # 0.7 for left margin
    fig_h   = 5.2

    fig = plt.figure(figsize=(fig_w, fig_h))

    # right edge fraction: leaves exactly (cbar_w + gap) inches for cbar+space
    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06,
        left=0.07,
        right=right_frac,
        top=0.80,
        bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])

    for ax in axes[1:]:
        ax.sharey(axes[0])

    fig.suptitle(
        f"{island_label} — Electrolyser (AEL) Activity  "
        f"(% of annual peak hourly consumption)",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_act, h_active, total_gwh = \
            build_activity_matrix(ael, island, year)
        im = draw_activity_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_act, h_active, total_gwh,
            show_ylab=(idx == 0),
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    # ── Colorbar anchored to its own dedicated axis ───────────────────────────
    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("Electrolyser activity (%)", fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)      # no bbox_inches="tight"
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)
    ael  = extract_ael_consumption(data)

    if ael.empty:
        log.error("No AEL data. Verify tech='%s', commodity='%s' in commodity_balance.",
                  AEL_TECH, AEL_COMMODITY)
        return

    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island_ael(ael, island, years, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island figures written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--out",     default=OUTPUT_DIR)
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+", default=YEARS)
    parser.add_argument("--tech",    default=AEL_TECH)
    args = parser.parse_args()

    AEL_TECH = args.tech
    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
##############Heat pump activity ###############################################
import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.colors import LinearSegmentedColormap

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/heat_pump_activity_minload"
YEARS      = ["2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

HP_TECH      = "HP"
HP_COMMODITY = "Elec"

# ── Month axis ────────────────────────────────────────────────────────────────
MONTH_BOUNDARIES = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
MONTH_LABELS     = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_MIDPOINTS  = [(MONTH_BOUNDARIES[i] + MONTH_BOUNDARIES[i + 1]) / 2
                    for i in range(12)]

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR MAP
# YlGnBu-inspired (matching battery SOC design):
# pale yellow (idle) → mint → teal → deep ocean blue (full load)
# Perceptually uniform, colorblind-safe, prints well in greyscale too.
# ─────────────────────────────────────────────────────────────────────────────
HP_CMAP = LinearSegmentedColormap.from_list(
    "soc",
    ["#FFFFCC", "#C7E9B4", "#41B6C4", "#1D91C0", "#0C2C84"],
    N=256,
)

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# IMPORTANT: savefig.bbox is intentionally NOT "tight".
# "tight" re-expands the canvas after layout and drags the colorbar axis
# back over the rightmost heatmap panel.  All margins are controlled via
# gridspec left/right/top/bottom fractions instead.
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.facecolor": "white",
    # savefig.bbox deliberately omitted (defaults to "standard")
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_hp_consumption(data: dict) -> pd.DataFrame:
    """
    Pull commodity_balance, filter to HP tech + Elec commodity.
    Take absolute value of the negative consumption rows.

    Returns DataFrame with columns:
        hour      (int 0-8759)
        island    (str)
        year      (str)
        elec_gwh  (float, GWh, always >= 0)
    """
    cb   = data["commodity_balance"]
    mask = (
        (cb["techs"]     == HP_TECH) &
        (cb["commodity"] == HP_COMMODITY) &
        (cb["Value"]     <  0)
    )
    hp = cb.loc[mask].copy()

    if hp.empty:
        log.warning("No rows for tech='%s', commodity='%s', Value<0.",
                    HP_TECH, HP_COMMODITY)
        return pd.DataFrame(columns=["hour", "island", "year", "elec_gwh"])

    hp["hour"] = (
        hp["timeModel"].str.replace("tm", "", regex=False).astype(int) - 1
    )
    hp.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "Value":         "elec_gwh",
    }, inplace=True)

    hp["elec_gwh"] = hp["elec_gwh"].abs()
    hp = hp.groupby(["hour", "island", "year"], as_index=False)["elec_gwh"].sum()

    log.info("HP rows: %d", len(hp))
    log.info("Islands: %s", sorted(hp["island"].unique()))
    log.info("Years  : %s", sorted(hp["year"].unique()))

    return hp[["hour", "island", "year", "elec_gwh"]]


# ─────────────────────────────────────────────────────────────────────────────
# ACTIVITY MATRIX BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_activity_matrix(hp: pd.DataFrame, island: str, year: str):
    mask = (hp["island"] == island) & (hp["year"] == year)
    sub  = hp.loc[mask, ["hour", "elec_gwh"]].copy()

    if sub.empty:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0, 0.0

    full = pd.Series(0.0, index=range(8760))
    full.update(sub.set_index("hour")["elec_gwh"])

    peak_gwh  = float(full.max())
    total_gwh = float(full.sum())

    if peak_gwh == 0:
        return np.full((24, 365), np.nan), 0.0, 0.0, 0.0, 0.0

    activity_pct = (full / peak_gwh * 100).values
    matrix       = activity_pct.reshape(365, 24).T

    avg_activity = float(activity_pct.mean())
    hours_active = float((activity_pct > 0).mean() * 100)

    return matrix, peak_gwh, avg_activity, hours_active, total_gwh


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE SUBPLOT RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def draw_activity_heatmap(ax, matrix: np.ndarray, year: str,
                          peak_gwh: float, avg_activity: float,
                          hours_active: float, total_gwh: float,
                          show_ylab: bool):

    if np.all(np.isnan(matrix)):
        ax.text(0.5, 0.5, "No HP data\n(heat pump not deployed)",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=9, color="#888888")
        ax.set_title(year, fontweight="bold")
        return None

    im = ax.imshow(
        matrix,
        aspect="auto",
        origin="upper",
        cmap=HP_CMAP,
        vmin=0, vmax=100,
        interpolation="nearest",
        extent=[1, 365, 23.5, -0.5],
    )

    # ── X-axis: month names, no overlapping day numbers ───────────────────────
    ax.set_xticks(MONTH_MIDPOINTS)
    ax.set_xticklabels(MONTH_LABELS, fontsize=8)
    ax.set_xticks(MONTH_BOUNDARIES, minor=True)
    ax.tick_params(axis="x", which="major", length=0)          # labels only
    ax.tick_params(axis="x", which="minor", length=3,
                   color="#aaaaaa", width=0.5)
    ax.set_xlim(1, 365)
    ax.set_xlabel("Month", labelpad=4)

    for mb in MONTH_BOUNDARIES[1:-1]:
        ax.axvline(mb, color="white", linewidth=0.4, alpha=0.55)

    # ── Y-axis ────────────────────────────────────────────────────────────────
    ax.set_yticks([0, 6, 12, 18, 23])
    ax.set_yticklabels(["00:00", "06:00", "12:00", "18:00", "23:00"],
                       fontsize=7.5)
    ax.set_ylim(23.5, -0.5)
    if show_ylab:
        ax.set_ylabel("Hour of day", labelpad=4)

    # ── Title + stats ─────────────────────────────────────────────────────────
    ax.set_title(year, fontweight="bold", pad=14)
    ax.text(
        0.5, 1.01,
        (f"Peak: {peak_gwh:.4f} GWh/h  |  "
         f"Total: {total_gwh:.2f} GWh/yr  |  "
         f"Avg activity: {avg_activity:.1f}%  |  "
         f"Hours active: {hours_active:.1f}%"),
        transform=ax.transAxes, ha="center", va="bottom",
        fontsize=7.2, color="#444444", fontstyle="italic",
    )

    return im


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island_hp(hp: pd.DataFrame, island: str,
                   years: list, out_dir: Path, fmt: str, dpi: int):

    island_label = island.replace("_model", "")
    n_years      = len(years)

    panel_w = 5.5                              # inches per heatmap panel
    cbar_w  = 0.3                              # inches for colorbar column
    gap     = 0.55                             # right-side whitespace (inches)
    fig_w   = panel_w * n_years + cbar_w + gap + 0.7   # 0.7 for left margin
    fig_h   = 5.2

    fig = plt.figure(figsize=(fig_w, fig_h))

    # right edge fraction: leaves exactly (cbar_w + gap) inches for cbar+space
    right_frac = 1.0 - (cbar_w + gap) / fig_w

    gs = fig.add_gridspec(
        1, n_years + 1,
        width_ratios=[panel_w] * n_years + [cbar_w],
        wspace=0.06,
        left=0.07,
        right=right_frac,
        top=0.80,
        bottom=0.14,
    )

    axes    = [fig.add_subplot(gs[0, i]) for i in range(n_years)]
    cbar_ax = fig.add_subplot(gs[0, n_years])

    for ax in axes[1:]:
        ax.sharey(axes[0])

    fig.suptitle(
        f"{island_label} — Heat Pump (HP) Activity  "
        f"(% of annual peak hourly consumption)",
        fontsize=11, fontweight="bold", y=0.95,
    )

    last_im = None
    for idx, year in enumerate(years):
        matrix, peak_gwh, avg_act, h_active, total_gwh = \
            build_activity_matrix(hp, island, year)
        im = draw_activity_heatmap(
            axes[idx], matrix, year, peak_gwh, avg_act, h_active, total_gwh,
            show_ylab=(idx == 0),
        )
        if im is not None:
            last_im = im
        if idx > 0:
            plt.setp(axes[idx].get_yticklabels(), visible=False)

    # ── Colorbar anchored to its own dedicated axis ───────────────────────────
    if last_im is not None:
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("Heat pump activity (%)", fontsize=9, labelpad=8)
        cbar.set_ticks([0, 20, 40, 60, 80, 100])
        cbar.ax.tick_params(labelsize=8)
        for t in [20, 40, 60, 80]:
            cbar.ax.axhline(t, color="white", linewidth=0.5, alpha=0.6)
    else:
        cbar_ax.set_visible(False)

    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)      # no bbox_inches="tight"
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)
    hp  = extract_hp_consumption(data)

    if hp.empty:
        log.error("No HP data. Verify tech='%s', commodity='%s' in commodity_balance.",
                  HP_TECH, HP_COMMODITY)
        return

    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island_hp(hp, island, years, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island figures written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--out",     default=OUTPUT_DIR)
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+", default=YEARS)
    parser.add_argument("--tech",    default=HP_TECH)
    args = parser.parse_args()

    HP_TECH = args.tech
    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
 
##############################original for capacities, generation and end use cases####################
import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.patches import Patch

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/capacity_overview_minload"
YEARS      = ["2020", "2030", "2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

# ── Generation technologies (panels 1 & 2) ───────────────────────────────────
EL_PRODUCTION_TECHS = [
    "PV_B", "PV_N",
    "WindOnshore_B", "WindOnshore_N", "WindOffshore_N",
    "Wave_N",
    "Hydro_B", "Hydro_N",
    "BG_B", "BG_N",
    "NG_plant",
    "DG",
    "Geothermal_B",
]

# ── Capacity correction factors (model scaling artefacts) ────────────────────
# Capacities are UNDERREPORTED in the GDX by these percentages.
# True value = reported / (1 - underreport_fraction)
#   Hydro_B, Hydro_N  : underreported by 55 % → scale = 1 / 0.45 ≈ 2.222
#   Geothermal_B      : underreported by 78 % → scale = 1 / 0.22 ≈ 4.545
CAPACITY_SCALE = {
    "Hydro_B":      1 / (1 - 0.55),   # ÷ 0.45
    "Hydro_N":      1 / (1 - 0.55),   # ÷ 0.45
    "Geothermal_B": 1 / (1 - 0.78),   # ÷ 0.22
}

# ── End-use / converter technologies (panel 3) ───────────────────────────────
END_USE_TECHS = [
    "Battery",
    "DW_Electric_converter",
    "HP",
    "cook_el",
    "Industry_EL",
    "RO",
    "AEL",
    "Ammonia_synthesis",
    "DAC",
    "LDV_el",
    "HDV_el",
    "MDV_el",
    "Bus_el",
    "Two_wheel_el",
    "Aviation_el",
    "Ship_BEV",
    "Demand",
]

# ── Technology display labels ─────────────────────────────────────────────────
TECH_LABELS = {
    # Generation
    "PV_B":              "Solar PV (existing)",
    "PV_N":              "Solar PV (new)",
    "WindOnshore_B":     "Wind onshore (existing)",
    "WindOnshore_N":     "Wind onshore (new)",
    "WindOffshore_N":    "Wind offshore",
    "Wave_N":            "Wave",
    "Hydro_B":           "Hydro (existing)",
    "Hydro_N":           "Hydro (new)",
    "BG_B":              "Biogas (existing)",
    "BG_N":              "Biogas (new)",
    "NG_plant":          "Natural gas",
    "DG":                "Diesel generator",
    "Geothermal_B":      "Geothermal",
    # End-use
    "Battery":               "Battery storage",
    "DW_Electric_converter": "Electric hot water",
    "HP":                    "Heat pump",
    "cook_el":               "Electric cooking",
    "Industry_EL":           "Industrial electr.",
    "RO":                    "Desalination (RO)",
    "AEL":                   "Electrolyser (AEL)",
    "Ammonia_synthesis":     "Ammonia synthesis",
    "DAC":                   "Direct air capture",
    "LDV_el":                "Light-duty EV",
    "HDV_el":                "Heavy-duty EV",
    "MDV_el":                "Medium-duty EV",
    "Bus_el":                "Electric bus",
    "Two_wheel_el":          "E-motorbike/scooter",
    "Aviation_el":           "Electric aviation",
    "Ship_BEV":              "Electric shipping",
    "Demand":                "Residential demand",
}

# ── Colour palettes ───────────────────────────────────────────────────────────
GEN_COLORS = {
    "PV_B":           "#F0C234",
    "PV_N":           "#E6A817",
    "WindOnshore_B":  "#56B4E9",
    "WindOnshore_N":  "#0072B2",
    "WindOffshore_N": "#004D80",
    "Wave_N":         "#009E73",
    "Hydro_B":        "#44AA99",
    "Hydro_N":        "#117733",
    "BG_B":           "#CC79A7",
    "BG_N":           "#882255",
    "NG_plant":       "#E69F00",
    "DG":             "#D55E00",
    "Geothermal_B":   "#8B4513",
}

ENDUSE_COLORS = {
    "Battery":               "#1A85FF",
    "DW_Electric_converter": "#D41159",
    "HP":                    "#FFC20A",
    "cook_el":               "#994F00",
    "Industry_EL":           "#006CD1",
    "RO":                    "#26C6DA",
    "AEL":                   "#40B0A6",
    "Ammonia_synthesis":     "#E1BE6A",
    "DAC":                   "#A0522D",
    "LDV_el":                "#5D3A9B",
    "HDV_el":                "#E66100",
    "MDV_el":                "#009E73",
    "Bus_el":                "#0072B2",
    "Two_wheel_el":          "#CC79A7",
    "Aviation_el":           "#882255",
    "Ship_BEV":              "#D55E00",
    "Demand":                "#888888",
}

FALLBACK_COLOR = "#AAAAAA"

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   9,
    "ytick.labelsize":   8,
    "legend.fontsize":   7.5,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "xtick.direction":   "out",
    "ytick.direction":   "out",
    "xtick.major.size":  3,
    "ytick.major.size":  3,
    "grid.linewidth":    0.4,
    "grid.alpha":        0.35,
    "grid.color":        "#888888",
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.bbox":      "tight",
    "savefig.facecolor": "white",
})

BAR_WIDTH = 0.6


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


# ─────────────────────────────────────────────────────────────────────────────
# PANEL 1 — INSTALLED CAPACITY
# ─────────────────────────────────────────────────────────────────────────────

def extract_capacity(data: dict) -> pd.DataFrame:
    """
    converter_caps | capType == 'total' | commodity == 'Elec'
    Returns DataFrame: island, year (str), tech, capacity_mw (float)

    Unit conversion : GW → MW  (multiply by 1000)
    Capacity corrections (model scaling artefacts):
        Hydro_B, Hydro_N  : multiply by (1 - 0.55) = 0.45
        Geothermal_B      : multiply by (1 - 0.78) = 0.22
    """
    cc = data["converter_caps"]
    mask = (
        (cc["capType"]   == "total") &
        (cc["commodity"] == "Elec") &
        (cc["techs"].isin(EL_PRODUCTION_TECHS))
    )
    df = cc.loc[mask].copy()
    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "techs":         "tech",
        "Value":         "capacity_mw",
    }, inplace=True)
    df["year"] = df["year"].astype(str)

    # Step 1: GW → MW
    df["capacity_mw"] = df["capacity_mw"] * 1000.0

    # Step 2: apply per-tech correction factors
    for tech, scale in CAPACITY_SCALE.items():
        mask_tech = df["tech"] == tech
        df.loc[mask_tech, "capacity_mw"] *= scale
        n = mask_tech.sum()
        if n > 0:
            log.info(
                "Capacity correction applied: %s × %.2f  (%d rows)",
                tech, scale, n,
            )

    return df[["island", "year", "tech", "capacity_mw"]]


def capacity_pivot(cap: pd.DataFrame, island: str) -> pd.DataFrame:
    sub = cap[cap["island"] == island]
    if sub.empty:
        return pd.DataFrame(index=YEARS)
    pivot = (
        sub.pivot_table(index="year", columns="tech",
                        values="capacity_mw", aggfunc="sum")
           .reindex(YEARS, fill_value=0.0)
           .fillna(0.0)
    )
    ordered = [t for t in EL_PRODUCTION_TECHS if t in pivot.columns
               and (pivot[t] > 0).any()]
    return pivot[ordered] if ordered else pd.DataFrame(index=YEARS)


# ─────────────────────────────────────────────────────────────────────────────
# PANEL 2 — ELECTRICITY GENERATION
# ─────────────────────────────────────────────────────────────────────────────

def extract_generation(data: dict) -> pd.DataFrame:
    """
    commodity_balance_annual | commodity == 'Elec' | balanceType == 'net' | Value > 0
    Returns DataFrame: island, year (str), tech, generation_gwh (float)
    """
    cb = data["commodity_balance_annual"]
    mask = (
        (cb["commodity"]   == "Elec") &
        (cb["balanceType"] == "net") &
        (cb["techs"].isin(EL_PRODUCTION_TECHS)) &
        (cb["Value"]       >  0)
    )
    df = cb.loc[mask].copy()
    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "techs":         "tech",
        "Value":         "generation_gwh",
    }, inplace=True)
    df["year"] = df["year"].astype(str)
    return df[["island", "year", "tech", "generation_gwh"]]


def generation_pivot(gen: pd.DataFrame, island: str) -> pd.DataFrame:
    sub = gen[gen["island"] == island]
    if sub.empty:
        return pd.DataFrame(index=YEARS)
    pivot = (
        sub.pivot_table(index="year", columns="tech",
                        values="generation_gwh", aggfunc="sum")
           .reindex(YEARS, fill_value=0.0)
           .fillna(0.0)
    )
    ordered = [t for t in EL_PRODUCTION_TECHS if t in pivot.columns
               and (pivot[t] > 0).any()]
    return pivot[ordered] if ordered else pd.DataFrame(index=YEARS)


# ─────────────────────────────────────────────────────────────────────────────
# PANEL 3 — ELECTRICITY END-USE
# ─────────────────────────────────────────────────────────────────────────────

def extract_enduse(data: dict) -> pd.DataFrame:
    """
    commodity_balance_annual | commodity == 'Elec' | balanceType == 'net' | Value < 0
    Filters to END_USE_TECHS.
    Returns DataFrame: island, year (str), tech, enduse_gwh (float, positive)
    """
    cb = data["commodity_balance_annual"]
    mask = (
        (cb["commodity"]   == "Elec") &
        (cb["balanceType"] == "net") &
        (cb["techs"].isin(END_USE_TECHS)) &
        (cb["Value"]       <  0)
    )
    df = cb.loc[mask].copy()
    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "techs":         "tech",
        "Value":         "enduse_gwh",
    }, inplace=True)
    df["enduse_gwh"] = df["enduse_gwh"].abs()
    df["year"] = df["year"].astype(str)
    return df[["island", "year", "tech", "enduse_gwh"]]


def enduse_pivot(eu: pd.DataFrame, island: str) -> pd.DataFrame:
    sub = eu[eu["island"] == island]
    if sub.empty:
        return pd.DataFrame(index=YEARS)
    pivot = (
        sub.pivot_table(index="year", columns="tech",
                        values="enduse_gwh", aggfunc="sum")
           .reindex(YEARS, fill_value=0.0)
           .fillna(0.0)
    )
    ordered = [t for t in END_USE_TECHS if t in pivot.columns
               and (pivot[t] > 0).any()]
    return pivot[ordered] if ordered else pd.DataFrame(index=YEARS)


# ─────────────────────────────────────────────────────────────────────────────
# STACKED BAR HELPER
# ─────────────────────────────────────────────────────────────────────────────

def draw_stacked_bar(ax, pivot: pd.DataFrame, color_map: dict,
                     ylabel: str, title: str, unit_scale: float = 1.0):
    if pivot.empty or pivot.shape[1] == 0:
        ax.text(0.5, 0.5, "No data", ha="center", va="center",
                transform=ax.transAxes, fontsize=9, color="#888888")
        ax.set_title(title, fontweight="bold", pad=6)
        return

    x      = np.arange(len(YEARS))
    bottom = np.zeros(len(YEARS))

    for tech in pivot.columns:
        values = pivot[tech].values * unit_scale
        color  = color_map.get(tech, FALLBACK_COLOR)
        label  = TECH_LABELS.get(tech, tech)
        ax.bar(x, values, BAR_WIDTH,
               bottom=bottom, label=label, color=color,
               edgecolor="white", linewidth=0.3, zorder=3)
        bottom += values

    for i, total in enumerate(bottom):
        if total > 0:
            ax.text(x[i], total * 1.01, f"{total:,.0f}",
                    ha="center", va="bottom", fontsize=7, color="#333333")

    ax.set_xticks(x)
    ax.set_xticklabels(YEARS, fontsize=9)
    ax.set_xlabel("Year", labelpad=4)
    ax.set_ylabel(ylabel, labelpad=4)
    ax.set_title(title, fontweight="bold", pad=6)
    ax.set_xlim(-0.5, len(YEARS) - 0.5)
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v:,.0f}")
    )


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island(cap_df: pd.DataFrame, gen_df: pd.DataFrame,
                eu_df: pd.DataFrame, island: str,
                out_dir: Path, fmt: str, dpi: int):
    island_label = island.replace("_model", "")

    cap_piv = capacity_pivot(cap_df, island)
    gen_piv = generation_pivot(gen_df, island)
    eu_piv  = enduse_pivot(eu_df, island)

    fig, axes = plt.subplots(1, 3, figsize=(18, 6))
    fig.suptitle(
        f"{island_label} — Installed Capacity, Electricity Generation "
        f"& End-Use  (2020 – 2050)",
        fontsize=12, fontweight="bold", y=1.02,
    )

    draw_stacked_bar(axes[0], cap_piv, GEN_COLORS,
                     "Installed capacity (MW)", "Installed Capacity")
    draw_stacked_bar(axes[1], gen_piv, GEN_COLORS,
                     "Generation (GWh)", "Electricity Generation")
    draw_stacked_bar(axes[2], eu_piv,  ENDUSE_COLORS,
                     "Electricity consumed (GWh)", "Electricity End-Use")

    # ── Generation legend (panels 1 & 2) ─────────────────────────────────────
    gen_techs_present = list(dict.fromkeys(
        list(cap_piv.columns) + list(gen_piv.columns)
    ))
    gen_handles = [
        Patch(facecolor=GEN_COLORS.get(t, FALLBACK_COLOR),
              edgecolor="white", linewidth=0.3,
              label=TECH_LABELS.get(t, t))
        for t in EL_PRODUCTION_TECHS if t in gen_techs_present
    ]
    if gen_handles:
        leg1 = axes[0].legend(
            handles=gen_handles,
            loc="upper center", bbox_to_anchor=(1.0, -0.18),
            ncol=min(len(gen_handles), 4), frameon=False,
            fontsize=7.5, title="Generation technologies", title_fontsize=8,
        )
        axes[0].add_artist(leg1)

    # ── End-use legend (panel 3) ──────────────────────────────────────────────
    eu_handles = [
        Patch(facecolor=ENDUSE_COLORS.get(t, FALLBACK_COLOR),
              edgecolor="white", linewidth=0.3,
              label=TECH_LABELS.get(t, t))
        for t in END_USE_TECHS if t in list(eu_piv.columns)
    ]
    if eu_handles:
        axes[2].legend(
            handles=eu_handles,
            loc="upper center", bbox_to_anchor=(0.5, -0.18),
            ncol=min(len(eu_handles), 4), frameon=False,
            fontsize=7.5, title="End-use sectors", title_fontsize=8,
        )

    plt.tight_layout(rect=[0, 0.12, 1, 1])
    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    global YEARS
    YEARS = years

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)

    log.info("Extracting installed capacities ...")
    cap_df = extract_capacity(data)

    log.info("Extracting electricity generation ...")
    gen_df = extract_generation(data)

    log.info("Extracting electricity end-use ...")
    eu_df  = extract_enduse(data)

    for island in islands:
        log.info("Plotting %s ...", island)
        plot_island(cap_df, gen_df, eu_df, island, out_path, fmt, dpi)

    log.info("Done. All figures saved to: %s", out_path)
    print(f"\nAll {len(islands)} island figures written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH)
    parser.add_argument("--out",     default=OUTPUT_DIR)
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",     default=200, type=int)
    parser.add_argument("--islands", nargs="+", default=ISLANDS)
    parser.add_argument("--years",   nargs="+",
                        default=["2020", "2030", "2040", "2050"])
    args = parser.parse_args()

    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
    ######################################################all commodity demand #############################
# DHW_el: Domestic hot water_elec 
# Dummy_EL: maritime demand_efuel
# Elec: Residential demand
# Methanol: maritime demand_efuel1
# T_Aviation_el: Aviation demand_elec
# T_Bus_el: Bus_elec
# T_HDV_el: HDV_elec
# T_Industry_EH: industry_heat_HP_ST
# T_LDV_el: LDV_el
# T_MDV_el: MDV_el
# T_Two_wheel_el: Two_wheel_el
# T_cook_el: cooking_heat_elec
# T_ship_el: Maritime_demand_elec
# eKerosene: aviation_efuel
# DHW_LPG: Hot_water_LPG
# Heat_cooking_cooking_Heat_Biomass
# Heat_industry: industry_heat_diesel
# T_Aviation_th: Aviation_demand_fossil
# T_Bus_th: Bus_fossil
# T_HDV_th: HDV_fossil
# T_LDV_th: LDV_fossil
# T_MDV_th: MDV_fossil
# T_Marine_f_th: Maritime_demand_fossil1
# T_Marine_th: Maritime_demand_fossil2
# T_Two_wheel_th: Two_wheel_fossil
# T_HDV_BF: HDV_biofuel
# T_LDV_BF: LDV_biofuel
# T_MDV_BF: MDV_biofuel
"""
Pacific Island Countries (PICs) — Final Energy Demand by Commodity
===================================================================
Reads 'commodity_balance_annual' from the GAMS GDX file and extracts
all final demand flows (techs == "Demand", balanceType == "net", Value < 0)
for 28 energy commodities across 14 islands and 4 target years.

Produces two sets of outputs:
  1. One figure per island  — 4 year-bars stacked by commodity (GWh)
  2. One combined overview figure — all 14 islands, 4 year-bars each

Output
------
  figures/final_demand/
      CI_model.png  ...  VU_model.png   ← per-island figures
      all_islands_overview.png          ← combined overview

Usage
-----
    python pic_final_demand.py
    python pic_final_demand.py --gdx path/to/results.gdx
                               --out figures/final_demand
                               --fmt png --dpi 200
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.patches import Patch

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/final_demand"
YEARS      = ["2020", "2030", "2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

# ── Commodity display labels (GDX name → legend label) ───────────────────────
COMMODITY_LABELS = {
    # ── Electric end-uses ────────────────────────────────────────────────────
    "Elec":           "Residential demand (elec)",
    "DHW_el":         "Domestic hot water (elec)",
    "T_cook_el":      "Cooking heat (elec)",
    "T_Industry_EH":  "Industry heat (elec/solar thermal)",
    "T_LDV_el":       "LDV (elec)",
    "T_MDV_el":       "MDV (elec)",
    "T_HDV_el":       "HDV (elec)",
    "T_Bus_el":       "Bus (elec)",
    "T_Two_wheel_el": "Two and three wheeler (elec)",
    "T_Aviation_el":  "Aviation (elec)",
    "T_ship_el":      "Maritime (elec)",
    # ── E-fuels ──────────────────────────────────────────────────────────────
    "eKerosene":      "Aviation (e-fuel)",
    "Dummy_EL":       "Maritime (e-fuel)",
    "Methanol":       "Maritime (e-fuel)",
    # ── Fossil end-uses ───────────────────────────────────────────────────────
    "Heat_industry":  "Industry heat (diesel boiler)",
    "T_LDV_th":       "LDV (fossil)",
    "T_MDV_th":       "MDV (fossil)",
    "T_HDV_th":       "HDV (fossil)",
    "T_Bus_th":       "Bus (fossil)",
    "T_Two_wheel_th": "Two and three wheeler (fossil)",
    "T_Aviation_th":  "Aviation (fossil)",
    "T_Marine_th":    "Maritime (fossil)",
    "T_Marine_f_th":  "Maritime (fossil)",
    # ── Biofuels ─────────────────────────────────────────────────────────────
    "T_LDV_BF":       "LDV (biofuel)",
    "T_MDV_BF":       "MDV (biofuel)",
    "T_HDV_BF":       "HDV (biofuel)",
    # ── Other heat / LPG ─────────────────────────────────────────────────────
    "Heat_cooking":   "Cooking heat (biomass)",
    "DHW_LPG":        "Domestic hot water (LPG)",
}

# Ordered list — electric first, e-fuels, fossil, biofuel, other
COMMODITY_ORDER = list(COMMODITY_LABELS.keys())

# ── Colour palette — grouped by energy carrier ───────────────────────────────
COMMODITY_COLORS = {
    # Electric (blues / teals)
    "Elec":           "#0072B2",
    "DHW_el":         "#56B4E9",
    "T_cook_el":      "#009E73",
    "T_Industry_EH":  "#00CED1",
    "T_LDV_el":       "#1A85FF",
    "T_MDV_el":       "#40B0A6",
    "T_HDV_el":       "#44AA99",
    "T_Bus_el":       "#117733",
    "T_Two_wheel_el": "#26C6DA",
    "T_Aviation_el":  "#0D47A1",
    "T_ship_el":      "#006CD1",
    # E-fuels (purples)
    "eKerosene":      "#882255",
    "Dummy_EL":       "#CC79A7",
    "Methanol":       "#AA4499",
    # Fossil (oranges / reds / browns)
    "Heat_industry":  "#D55E00",
    "T_LDV_th":       "#E69F00",
    "T_MDV_th":       "#F0C234",
    "T_HDV_th":       "#E66100",
    "T_Bus_th":       "#A65200",
    "T_Two_wheel_th": "#DDCC77",
    "T_Aviation_th":  "#994F00",
    "T_Marine_th":    "#8B4513",
    "T_Marine_f_th":  "#A0522D",
    # Biofuels (greens)
    "T_LDV_BF":       "#4DAF4A",
    "T_MDV_BF":       "#33A02C",
    "T_HDV_BF":       "#1B7837",
    # Other heat / LPG (greys / neutrals)
    "Heat_cooking":   "#BBBBBB",
    "DHW_LPG":        "#888888",
}

BAR_WIDTH  = 0.6
YEAR_INTS  = [int(y) for y in YEARS]

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   9,
    "ytick.labelsize":   8,
    "legend.fontsize":   7.5,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "xtick.direction":   "out",
    "ytick.direction":   "out",
    "xtick.major.size":  3,
    "ytick.major.size":  3,
    "grid.linewidth":    0.4,
    "grid.alpha":        0.35,
    "grid.color":        "#888888",
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.bbox":      "tight",
    "savefig.facecolor": "white",
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_demand(data: dict) -> pd.DataFrame:
    """
    commodity_balance_annual
        | techs       == "Demand"
        | balanceType == "net"
        | Value       <  0          (consumption from the system)
        | commodity   in COMMODITY_ORDER

    Returns DataFrame: island, year (str), commodity, demand_gwh (float, positive)
    Values in the GDX are negative (demand withdraws from the balance);
    we take absolute value for plotting.
    """
    cb = data["commodity_balance_annual"]

    mask = (
        (cb["techs"]       == "Demand") &
        (cb["balanceType"] == "net") &
        (cb["commodity"].isin(COMMODITY_ORDER)) &
        (cb["Value"]       <  0)
    )
    df = cb.loc[mask].copy()

    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "commodity":     "commodity",
        "Value":         "demand_gwh",
    }, inplace=True)

    df["demand_gwh"] = df["demand_gwh"].abs()
    df["year"]       = df["year"].astype(str)

    # Aggregate (should be one row per island-year-commodity, but safe)
    df = (
        df.groupby(["island", "year", "commodity"], as_index=False)["demand_gwh"]
        .sum()
    )

    log.info("Demand rows extracted: %d", len(df))
    log.info("Commodities found: %s", sorted(df["commodity"].unique()))
    return df[["island", "year", "commodity", "demand_gwh"]]


def demand_pivot(df: pd.DataFrame, island: str) -> pd.DataFrame:
    """
    Returns (years × commodities) pivot of demand in GWh for one island.
    Rows = YEARS, columns = commodities present for this island.
    Ordered by COMMODITY_ORDER; only non-zero commodities kept.
    """
    sub = df[df["island"] == island]
    if sub.empty:
        return pd.DataFrame(index=YEARS)

    pivot = (
        sub.pivot_table(index="year", columns="commodity",
                        values="demand_gwh", aggfunc="sum")
           .reindex(YEARS, fill_value=0.0)
           .fillna(0.0)
    )
    ordered = [c for c in COMMODITY_ORDER
               if c in pivot.columns and (pivot[c] > 0).any()]
    return pivot[ordered] if ordered else pd.DataFrame(index=YEARS)


# ─────────────────────────────────────────────────────────────────────────────
# LEGEND BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_legend_handles(commodities_present: list) -> list:
    """Return Patch handles for the commodities actually present,
    reversed so the legend reads bottom-to-top instead of top-to-bottom."""
    handles = [
        Patch(facecolor=COMMODITY_COLORS.get(c, "#AAAAAA"),
              edgecolor="white", linewidth=0.3,
              label=COMMODITY_LABELS.get(c, c))
        for c in COMMODITY_ORDER
        if c in commodities_present
    ]
    return handles[::-1]


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island_demand(pivot: pd.DataFrame, island: str,
                       out_dir: Path, fmt: str, dpi: int):
    """Stacked bar chart for one island — 4 year bars, stacked by commodity."""
    island_label = island.replace("_model", "")

    fig, ax = plt.subplots(figsize=(7, 5))
    fig.suptitle(
        f"{island_label} — Final Energy Demand by Commodity (2020–2050)",
        fontsize=11, fontweight="bold", y=1.02,
    )

    if pivot.empty or pivot.shape[1] == 0:
        ax.text(0.5, 0.5, "No demand data", ha="center", va="center",
                transform=ax.transAxes, fontsize=10, color="#888888")
    else:
        x      = np.arange(len(YEARS))
        bottom = np.zeros(len(YEARS))

        for commodity in pivot.columns:
            values = pivot[commodity].values
            color  = COMMODITY_COLORS.get(commodity, "#AAAAAA")
            ax.bar(x, values, BAR_WIDTH,
                   bottom=bottom,
                   color=color,
                   edgecolor="white", linewidth=0.3,
                   zorder=3)
            bottom += values

        # Total value labels on top
        for i, total in enumerate(bottom):
            if total > 0:
                ax.text(x[i], total * 1.01, f"{total:,.0f}",
                        ha="center", va="bottom", fontsize=7.5, color="#333333")

        ax.set_xticks(x)
        ax.set_xticklabels(YEARS, fontsize=9)
        ax.set_xlabel("Year", labelpad=4)
        ax.set_ylabel("Final energy demand (GWh)", labelpad=4)
        ax.set_xlim(-0.5, len(YEARS) - 0.5)
        ax.set_ylim(bottom=0)
        ax.yaxis.grid(True, zorder=0)
        ax.set_axisbelow(True)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda v, _: f"{v:,.0f}")
        )

        # Legend — on the right side of the plot
        handles = build_legend_handles(list(pivot.columns))
        ax.legend(
            handles=handles,
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            frameon=False,
            fontsize=7.2,
            title="Energy carrier / end-use",
            title_fontsize=7.5,
        )

    plt.tight_layout()
    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# COMBINED OVERVIEW FIGURE  (all 14 islands in one chart)
# ─────────────────────────────────────────────────────────────────────────────

def plot_overview(df: pd.DataFrame, islands: list,
                  out_dir: Path, fmt: str, dpi: int):
    """
    One figure with all islands on the x-axis.
    Each island has 4 grouped year-bars, stacked by commodity.
    """
    n_islands   = len(islands)
    n_years     = len(YEARS)
    group_width = n_years * BAR_WIDTH + 0.4    # total width per island group
    x_groups    = np.arange(n_islands) * group_width
    year_offsets = np.arange(n_years) * BAR_WIDTH - (n_years - 1) * BAR_WIDTH / 2

    # Collect all commodities present across all islands
    all_commodities = [c for c in COMMODITY_ORDER
                       if c in df["commodity"].values]

    fig, ax = plt.subplots(figsize=(max(18, n_islands * 1.8), 6))
    fig.suptitle(
        "Final Energy Demand by Commodity — All Pacific Islands (2020–2050)",
        fontsize=12, fontweight="bold", y=1.02,
    )

    year_colors_fallback = ["#0072B2", "#E69F00", "#009E73", "#D55E00"]

    for g_idx, island in enumerate(islands):
        pivot = demand_pivot(df, island)
        if pivot.empty or pivot.shape[1] == 0:
            continue

        for y_idx, year in enumerate(YEARS):
            bar_x  = x_groups[g_idx] + year_offsets[y_idx]
            bottom = 0.0
            for commodity in all_commodities:
                if commodity not in pivot.columns:
                    continue
                val   = float(pivot.loc[year, commodity]) if year in pivot.index else 0.0
                color = COMMODITY_COLORS.get(commodity, "#AAAAAA")
                ax.bar(bar_x, val, BAR_WIDTH * 0.92,
                       bottom=bottom, color=color,
                       edgecolor="white", linewidth=0.2, zorder=3)
                bottom += val

    # X-axis: island labels centred on each group
    island_labels = [i.replace("_model", "") for i in islands]
    ax.set_xticks(x_groups)
    ax.set_xticklabels(island_labels, rotation=35, ha="right", fontsize=8)
    ax.set_xlabel("Island", labelpad=6)
    ax.set_ylabel("Final energy demand (GWh)", labelpad=4)
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v:,.0f}")
    )

    # Year sub-labels below each group
    for g_idx in range(n_islands):
        for y_idx, year in enumerate(YEARS):
            ax.text(
                x_groups[g_idx] + year_offsets[y_idx],
                -ax.get_ylim()[1] * 0.04,
                year[-2:],    # e.g. "20", "30"
                ha="center", va="top", fontsize=6.5, color="#555555",
            )

    # Commodity legend on the right
    handles = build_legend_handles(all_commodities)
    ax.legend(
        handles=handles,
        loc="center left",
        bbox_to_anchor=(1.01, 0.5),
        frameon=False,
        fontsize=7.2,
        title="Energy carrier / end-use",
        title_fontsize=7.5,
        ncol=1,
    )

    plt.tight_layout()
    save_path = out_dir / f"all_islands_overview.{fmt}"
    fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    log.info("Overview saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    global YEARS
    YEARS = years

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # ── Load GDX ──────────────────────────────────────────────────────────────
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)

    # ── Extract demand ────────────────────────────────────────────────────────
    log.info("Extracting final demand ...")
    df = extract_demand(data)

    # ── Per-island figures ────────────────────────────────────────────────────
    for island in islands:
        log.info("Plotting %s ...", island)
        pivot = demand_pivot(df, island)
        plot_island_demand(pivot, island, out_path, fmt, dpi)

    # ── Combined overview ─────────────────────────────────────────────────────
    log.info("Plotting combined overview ...")
    plot_overview(df, islands, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\n{len(islands)} island figures + overview written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH,
                        help="Path to GAMS GDX file")
    parser.add_argument("--out",     default=OUTPUT_DIR,
                        help="Output directory for figures")
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"],
                        help="Output file format (default: png)")
    parser.add_argument("--dpi",     default=200, type=int,
                        help="Save resolution in DPI (default: 200)")
    parser.add_argument("--islands", nargs="+", default=ISLANDS,
                        help="Subset of islands to plot (default: all 14)")
    parser.add_argument("--years",   nargs="+", default=YEARS,
                        help="Target years (default: 2020 2030 2040 2050)")
    args = parser.parse_args()

    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)
    
    ####################heat generation #####################################################
    """
Pacific Island Countries (PICs) — Heat Generation by Technology
================================================================
Reads 'commodity_balance_annual' from the GAMS GDX file and extracts
positive (production) flows for specific tech–commodity pairs representing
all forms of heat supply across 14 islands and 4 target years.

Filter applied
--------------
    balanceType == "net"
    Value       >  0          (positive = output / production)
    (tech, commodity) in the table below:

    Tech                  Commodity        Heat type
    ────────────────────  ───────────────  ──────────────────────────
    DW_Electric_converter DHW_el           Electric domestic hot water
    DW_LPG_converter      DHW_LPG          LPG domestic hot water
    Industry              Heat_industry    Industrial heat (diesel)
    cook_b                Heat_cooking     Biomass cooking
    cook_el               T_cook_el        Electric cooking
    cook_LPG              T_cook_LPG       LPG cooking
    HP                    Heat             Heat-pump space/water heat
    Industry_EL           T_Industry_EH    Industrial heat (elec/HP/ST)
    ST_N                  Heat             Solar thermal heat

All commodity types are treated as "heat" in the plot; the legend
uses technology names to distinguish the source.

Output
------
  figures/heat_generation/
      CI_model.png  ...  VU_model.png   ← per-island figures
      all_islands_overview.png          ← combined overview

Usage
-----
    python pic_heat_generation.py
    python pic_heat_generation.py --gdx path/to/results.gdx
                                  --out figures/heat_generation
                                  --fmt png --dpi 200
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.patches import Patch

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/heat_generation"
YEARS      = ["2020","2030", "2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

# ── Tech–commodity pairs to extract (exact matching, no cross-contamination) ──
# Each tuple is (tech, commodity_produced).
HEAT_PAIRS = [
    ("DW_Electric_converter", "DHW_el"),
    ("DW_LPG_converter",      "DHW_LPG"),
    ("Industry",              "Heat_industry"),
    ("cook_b",                "Heat_cooking"),
    ("cook_el",               "T_cook_el"),
    ("cook_LPG",              "T_cook_LPG"),
    ("HP",                    "Heat"),
    ("Industry_EL",           "T_Industry_EH"),
    ("ST_N",                  "Heat"),
]

HEAT_PAIRS_SET = set(HEAT_PAIRS)

# Ordered list of unique techs (defines stack order bottom → top)
HEAT_TECHS_ORDER = list(dict.fromkeys(t for t, _ in HEAT_PAIRS))

# ── Technology display labels ─────────────────────────────────────────────────
TECH_LABELS = {
    "DW_Electric_converter": "Electric domestic hot water",
    "DW_LPG_converter":      "LPG domestic hot water",
    "Industry":              "Industry heat (diesel)",
    "cook_b":                "Biomass cooking",
    "cook_el":               "Electric cooking",
    "cook_LPG":              "LPG cooking",
    "HP":                    "Heat pump",
    "Industry_EL":           "Industry direct electric heat",
    "ST_N":                  "Solar thermal",
}

# ── Colour palette — colorblind-safe (Wong 2011 + extensions) ─────────────────
TECH_COLORS = {
    "DW_Electric_converter": "#0072B2",   # blue
    "DW_LPG_converter":      "#E69F00",   # amber
    "Industry":              "#D55E00",   # vermillion
    "cook_b":                "#117733",   # dark green
    "cook_el":               "#56B4E9",   # sky blue
    "cook_LPG":              "#F0E442",   # yellow
    "HP":                    "#009E73",   # teal
    "Industry_EL":           "#CC79A7",   # pink
    "ST_N":                  "#E6A817",   # dark amber / solar
}

BAR_WIDTH = 0.6

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE
# ─────────────────────────────────────────────────────────────────────────────
rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.titlesize":    10,
    "axes.labelsize":    9,
    "xtick.labelsize":   9,
    "ytick.labelsize":   8,
    "legend.fontsize":   7.5,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "xtick.direction":   "out",
    "ytick.direction":   "out",
    "xtick.major.size":  3,
    "ytick.major.size":  3,
    "grid.linewidth":    0.4,
    "grid.alpha":        0.35,
    "grid.color":        "#888888",
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.bbox":      "tight",
    "savefig.facecolor": "white",
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_heat(data: dict) -> pd.DataFrame:
    """
    commodity_balance_annual
        | balanceType == "net"
        | Value       >  0              (positive = production output)
        | (tech, commodity) in HEAT_PAIRS_SET

    Returns DataFrame: island, year (str), tech, heat_gwh (float, positive)
    All commodity types are aggregated to a single "heat" column per tech.
    """
    cb = data["commodity_balance_annual"]

    # Pre-filter to candidate techs and commodities (fast)
    all_techs = [t for t, _ in HEAT_PAIRS]
    all_comms = [c for _, c in HEAT_PAIRS]

    pre = cb.loc[
        (cb["balanceType"] == "positive") &
        (cb["Value"]       >  0) &
        (cb["techs"].isin(all_techs)) &
        (cb["commodity"].isin(all_comms))
    ].copy()

    if pre.empty:
        log.warning("No heat production rows found — check tech/commodity names.")
        return pd.DataFrame(columns=["island", "year", "tech", "heat_gwh"])

    # Exact pair filtering (removes e.g. HP producing DHW_el, which is not in the table)
    pre["_pair"] = list(zip(pre["techs"], pre["commodity"]))
    df = pre[pre["_pair"].isin(HEAT_PAIRS_SET)].copy()
    df.drop(columns=["_pair"], inplace=True)

    df.rename(columns={
        "accNodesModel": "island",
        "accYears":      "year",
        "techs":         "tech",
        "Value":         "heat_gwh",
    }, inplace=True)
    df["year"] = df["year"].astype(str)

    # Sum over commodities — all output from a tech counts as "heat"
    df = (
        df.groupby(["island", "year", "tech"], as_index=False)["heat_gwh"]
        .sum()
    )

    log.info("Heat rows extracted: %d", len(df))
    log.info("Techs found        : %s", sorted(df["tech"].unique()))
    return df[["island", "year", "tech", "heat_gwh"]]


def heat_pivot(df: pd.DataFrame, island: str) -> pd.DataFrame:
    """
    Returns (years × techs) pivot of heat production in GWh for one island.
    Rows = YEARS, columns = techs active for this island (non-zero only).
    """
    sub = df[df["island"] == island]
    if sub.empty:
        return pd.DataFrame(index=YEARS)

    pivot = (
        sub.pivot_table(index="year", columns="tech",
                        values="heat_gwh", aggfunc="sum")
           .reindex(YEARS, fill_value=0.0)
           .fillna(0.0)
    )
    ordered = [t for t in HEAT_TECHS_ORDER
               if t in pivot.columns and (pivot[t] > 0).any()]
    return pivot[ordered] if ordered else pd.DataFrame(index=YEARS)


# ─────────────────────────────────────────────────────────────────────────────
# LEGEND BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_legend_handles(techs_present: list) -> list:
    return [
        Patch(facecolor=TECH_COLORS.get(t, "#AAAAAA"),
              edgecolor="white", linewidth=0.3,
              label=TECH_LABELS.get(t, t))
        for t in HEAT_TECHS_ORDER
        if t in techs_present
    ]


# ─────────────────────────────────────────────────────────────────────────────
# STACKED BAR HELPER
# ─────────────────────────────────────────────────────────────────────────────

def draw_stacked_bar(ax, pivot: pd.DataFrame, ylabel: str, title: str):
    """Draw a stacked heat bar chart for one island into ax."""
    if pivot.empty or pivot.shape[1] == 0:
        ax.text(0.5, 0.5, "No heat data", ha="center", va="center",
                transform=ax.transAxes, fontsize=9, color="#888888")
        ax.set_title(title, fontweight="bold", pad=6)
        return

    x      = np.arange(len(YEARS))
    bottom = np.zeros(len(YEARS))

    for tech in pivot.columns:
        values = pivot[tech].values
        color  = TECH_COLORS.get(tech, "#AAAAAA")
        label  = TECH_LABELS.get(tech, tech)
        ax.bar(x, values, BAR_WIDTH,
               bottom=bottom, label=label, color=color,
               edgecolor="white", linewidth=0.3, zorder=3)
        bottom += values

    # Total label on top of each bar
    for i, total in enumerate(bottom):
        if total > 0:
            ax.text(x[i], total * 1.01, f"{total:,.1f}",
                    ha="center", va="bottom", fontsize=7.5, color="#333333")

    ax.set_xticks(x)
    ax.set_xticklabels(YEARS, fontsize=9)
    ax.set_xlabel("Year", labelpad=4)
    ax.set_ylabel(ylabel, labelpad=4)
    ax.set_title(title, fontweight="bold", pad=6)
    ax.set_xlim(-0.5, len(YEARS) - 0.5)
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v:,.1f}")
    )


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island_heat(pivot: pd.DataFrame, island: str,
                     out_dir: Path, fmt: str, dpi: int):
    """One figure per island — 4 year bars stacked by heat technology."""
    island_label = island.replace("_model", "")

    fig, ax = plt.subplots(figsize=(7, 5))
    fig.suptitle(
        f"{island_label} — Heat Generation by Technology (2020–2050)",
        fontsize=11, fontweight="bold", y=1.02,
    )

    draw_stacked_bar(ax, pivot,
                     ylabel="Heat generation (GWh)",
                     title="")

    handles = build_legend_handles(list(pivot.columns) if not pivot.empty else [])
    if handles:
        ax.legend(
            handles=handles,
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            frameon=False,
            fontsize=7.2,
            title="Heat technology",
            title_fontsize=7.5,
        )

    plt.tight_layout()
    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# COMBINED OVERVIEW FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_overview(df: pd.DataFrame, islands: list,
                  out_dir: Path, fmt: str, dpi: int):
    """All 14 islands on one chart — 4 year-bars each, stacked by tech."""
    n_islands   = len(islands)
    n_years     = len(YEARS)
    group_width = n_years * BAR_WIDTH + 0.4
    x_groups    = np.arange(n_islands) * group_width
    year_offsets = (np.arange(n_years) * BAR_WIDTH
                    - (n_years - 1) * BAR_WIDTH / 2)

    all_techs_present = [t for t in HEAT_TECHS_ORDER
                         if t in df["tech"].values]

    fig, ax = plt.subplots(figsize=(max(18, n_islands * 2.0), 6))
    fig.suptitle(
        "Heat Generation by Technology — All Pacific Islands (2020–2050)",
        fontsize=12, fontweight="bold", y=1.02,
    )

    for g_idx, island in enumerate(islands):
        pivot = heat_pivot(df, island)
        if pivot.empty or pivot.shape[1] == 0:
            continue

        for y_idx, year in enumerate(YEARS):
            bar_x  = x_groups[g_idx] + year_offsets[y_idx]
            bottom = 0.0
            for tech in all_techs_present:
                if tech not in pivot.columns:
                    continue
                val   = float(pivot.loc[year, tech]) if year in pivot.index else 0.0
                color = TECH_COLORS.get(tech, "#AAAAAA")
                ax.bar(bar_x, val, BAR_WIDTH * 0.92,
                       bottom=bottom, color=color,
                       edgecolor="white", linewidth=0.2, zorder=3)
                bottom += val

    # X-axis island labels
    island_labels = [i.replace("_model", "") for i in islands]
    ax.set_xticks(x_groups)
    ax.set_xticklabels(island_labels, rotation=35, ha="right", fontsize=8)
    ax.set_xlabel("Island", labelpad=6)
    ax.set_ylabel("Heat generation (GWh)", labelpad=4)
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{v:,.1f}")
    )

    # Year sub-labels below each group
    for g_idx in range(n_islands):
        for y_idx, year in enumerate(YEARS):
            ax.text(
                x_groups[g_idx] + year_offsets[y_idx],
                -ax.get_ylim()[1] * 0.04,
                year[-2:],
                ha="center", va="top", fontsize=6.5, color="#555555",
            )

    # Legend on the right
    handles = build_legend_handles(all_techs_present)
    ax.legend(
        handles=handles,
        loc="center left",
        bbox_to_anchor=(1.01, 0.5),
        frameon=False,
        fontsize=7.2,
        title="Heat technology",
        title_fontsize=7.5,
    )

    plt.tight_layout()
    save_path = out_dir / f"all_islands_overview.{fmt}"
    fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    log.info("Overview saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path: str, out_dir: str, fmt: str, dpi: int,
         islands: list, years: list, preloaded_data=None):

    global YEARS
    YEARS = years

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)

    log.info("Extracting heat generation ...")
    df = extract_heat(data)

    for island in islands:
        log.info("Plotting %s ...", island)
        pivot = heat_pivot(df, island)
        plot_island_heat(pivot, island, out_path, fmt, dpi)

    log.info("Plotting combined overview ...")
    plot_overview(df, islands, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\n{len(islands)} island figures + overview written to: {out_path}/")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",     default=GDX_PATH,
                        help="Path to GAMS GDX file")
    parser.add_argument("--out",     default=OUTPUT_DIR,
                        help="Output directory for figures")
    parser.add_argument("--fmt",     default="png",
                        choices=["png", "pdf", "svg", "tiff"],
                        help="Output file format (default: png)")
    parser.add_argument("--dpi",     default=200, type=int,
                        help="Save resolution in DPI (default: 200)")
    parser.add_argument("--islands", nargs="+", default=ISLANDS,
                        help="Subset of islands to plot (default: all 14)")
    parser.add_argument("--years",   nargs="+", default=YEARS,
                        help="Target years (default: 2020 2030 2040 2050)")
    args = parser.parse_args()

    main(args.gdx, args.out, args.fmt, args.dpi, args.islands, args.years)

############################################################################

#     ################################Sankey diagram pre-processing for sankeymatic#################################
# """
# SankeyMatic Prompt Generator — Excel version
# Reads Input/Output/Value columns from an .xlsx file and outputs SankeyMatic-ready text.

# Usage:
#     python sankey_from_excel.py                          # uses EXCEL_PATH below
#     python sankey_from_excel.py path/to/file.xlsx        # override path
#     python sankey_from_excel.py path/to/file.xlsx Sheet2 # override path + sheet name
# """

import sys
import pandas as pd
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\sankey_2050.xlsx"
SHEET_NAME  = 0          # 0 = first sheet; change to "SheetName" if needed
OUTPUT_FILE = "sankey_output.txt"

# Fixed loss lines appended after the data rows (value is always *)
LOSS_TECHS = [
    "AEL",
    "FTL",
    "Ammonia_synthesis",
    "Methanol_synthesis",
    "Power",
    "cook_el",
    "Industry_EL",
    "DW_Electric_converter",
    "BG_N",
]
# ─────────────────────────────────────────────────────────────────────────────

def parse_value(raw) -> str:
    """Return a rounded integer string or '*' for empty/wildcard values."""
    if pd.isna(raw) or str(raw).strip() in ("", "*"):
        return "*"
    s = str(raw).strip().replace(",", ".")
    try:
        return str(round(float(s)))
    except ValueError:
        return "*"

def table_to_sankey(df: pd.DataFrame) -> str:
    df.columns = [str(c).strip() for c in df.columns]
    cols = [c for c in df.columns if c][:3]
    if len(cols) < 2:
        raise ValueError(f"Expected at least 2 columns, found: {df.columns.tolist()}")
    col_in  = cols[0]
    col_out = cols[1]
    col_val = cols[2] if len(cols) > 2 else None

    lines = []
    for _, row in df.iterrows():
        inp = str(row[col_in]).strip()  if pd.notna(row[col_in])  else ""
        out = str(row[col_out]).strip() if pd.notna(row[col_out]) else ""
        val = parse_value(row[col_val]) if col_val else "*"
        if not inp:
            continue
        if val == "*":
            lines.append(f"{inp} [*] {out}")
        else:
            lines.append(f"{inp} [{val}] {out}")

    # Append fixed loss lines
    lines.append("")  # blank separator
    for tech in LOSS_TECHS:
        lines.append(f"{tech} [*] Loss")

    return "\n".join(lines)

def main():
    path  = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    sheet = sys.argv[2] if len(sys.argv) > 2 else SHEET_NAME
    print(f"Reading: {path}  |  Sheet: {sheet}")
    try:
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    except FileNotFoundError:
        print(f"\n❌  File not found: {path}")
        print("    Check the path or pass it as a command-line argument.")
        sys.exit(1)

    result = table_to_sankey(df)
    print("\n── SankeyMatic output ──────────────────────────────────────\n")
    print(result)

    out_path = Path(OUTPUT_FILE)
    out_path.write_text(result, encoding="utf-8")
    print(f"\n✓ Saved to: {out_path.resolve()}")

if __name__ == "__main__":
    main()
#  ###################################Sankey excel creater##########################
# """
# extract_sankey.py
# -----------------
# Extracts commodity-balance flows for year 2050 (all islands summed as "Global")
# from a GDX file and writes a 3-column Excel sheet: Input | Output | Value (GWh).

# Usage:
#     python extract_sankey.py --gdx path/to/results.gdx --out sankey_2050.xlsx
# """

# """
# extract_sankey.py
# -----------------
# Extracts commodity-balance flows for year 2050 (all islands summed as "Global")
# from a GDX file and writes a 3-column Excel sheet: Input | Output | Value (GWh).

# Usage:
#     python extract_sankey.py --gdx path/to/results.gdx --out sankey_2050.xlsx
# """

# """
# extract_sankey.py
# -----------------
# Extracts commodity-balance flows for year 2050 (all islands summed as "Global")
# from a GDX file and writes a 3-column Excel sheet: Input | Output | Value (GWh).

# Usage:
#     python extract_sankey.py --gdx path/to/results.gdx --out sankey_2050.xlsx
# """

import argparse
import logging
from pathlib import Path

import gdxpds
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ── CLI defaults (edit these or override via --gdx / --out) ──────────────────
GDX_PATH = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S1_minload.gdx"
OUTPUT_PATH = "sankey_2050.xlsx"
YEAR = "2050"
BALANCE_TYPE = "net"

# Explicit list of the 14 island-level nodes.
# Only these are summed to form "Global" — this prevents double-counting
# if the GDX also contains pre-aggregated regional/total nodes.
ISLANDS = [
    "CI_model", "FJ_model", "FSM_model", "KB_model", "MI_model",
    "NU_model", "NE_model", "PU_model", "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model", "VU_model",
]


def load_cb(gdx_path: str) -> pd.DataFrame:
    """Load commodity_balance_annual from GDX, filter to year + net + global sum.

    IMPORTANT: We filter to ONLY the 14 named island nodes before summing.
    REMix GDX files often contain additional aggregate/regional nodes (e.g. a
    Pacific-total node) alongside the per-island nodes.  Summing all
    accNodesModel values without this filter causes every flow to appear twice
    — once from the per-island rows and once from the pre-aggregated total row.
    """
    log.info("Loading GDX …")
    # data = gdxpds.to_dataframes(gdx_path)
    cb = data["commodity_balance_annual"]
    log.info("commodity_balance_annual shape: %s", cb.shape)
    log.info("Columns: %s", cb.columns.tolist())

    # ── Diagnostic: show every unique node so the user can spot extras ────────
    all_nodes = sorted(cb["accNodesModel"].unique())
    log.info("All accNodesModel values in GDX (%d): %s", len(all_nodes), all_nodes)

    # ── Filter to year, balance type AND only the 14 island nodes ────────────
    cb = cb[
        (cb["accYears"] == YEAR) &
        (cb["balanceType"] == BALANCE_TYPE) &
        (cb["accNodesModel"].isin(ISLANDS))
    ].copy()

    island_count = cb["accNodesModel"].nunique()
    log.info("Islands found after filter: %d / %d", island_count, len(ISLANDS))
    missing = set(ISLANDS) - set(cb["accNodesModel"].unique())
    if missing:
        log.warning("These islands had NO data for year %s: %s", YEAR, missing)

    # ── Sum across the 14 islands → single "Global" value per tech+commodity ──
    cb = (
        cb.groupby(["techs", "commodity"], as_index=False)["Value"]
        .sum()
    )
    log.info("After global aggregation: %d unique tech-commodity pairs", len(cb))
    return cb


def get_val(cb: pd.DataFrame, techs, commodity: str, sign: str = "pos") -> float:
    """Return summed absolute value for given tech(s), commodity, and sign."""
    if isinstance(techs, str):
        techs = [techs]
    mask = cb["techs"].isin(techs) & (cb["commodity"] == commodity)
    vals = cb.loc[mask, "Value"]
    if sign == "pos":
        result = vals[vals > 0].sum()
    else:
        result = vals[vals < 0].sum()
    return abs(float(result))


def build_rows(cb: pd.DataFrame) -> list[dict]:
    rows = []

    def add(inp, out, val):
        if val and val > 0:
            rows.append({"Input": inp, "Output": out, "Value (GWh)": round(val, 4)})

    # ── 1. Electricity generation: tech → Power ───────────────────────────────
    el_gen_techs = [
        "BG_N", "PV_N", "WindOnshore_N", "Hydro_N", "Wave_N", "DG",
        "NG_plant", "BG_B", "PV_B", "WindOnshore_B", "Hydro_B",
        "Geothermal_B", "WindOffshore_N",
    ]
    for tech in el_gen_techs:
        val = get_val(cb, tech, "Elec", "pos")
        add(tech, "Power", val)

    # ── 2. Biomass input to BG_N, BG_B ───────────────────────────────────────
    for tech in ["BG_N", "BG_B"]:
        val = get_val(cb, tech, "Biomass", "neg")
        add("Biomass", tech, val)

    # ── 3. Power consumed by converter technologies ───────────────────────────
    power_consumers = ["AEL", "DAC", "RO", "Industry_EL", "HP",
                        "DW_Electric_converter", "Ammonia_synthesis"]
    for tech in power_consumers:
        val = get_val(cb, tech, "Elec", "neg")
        add("Power", tech, val)

    # ── 4. AEL outputs: Hydrogen and Pure_water (via RO) ─────────────────────
    h2_val = get_val(cb, "AEL", "Hydrogen", "pos")
    add("AEL", "Hydrogen", h2_val)

    pw_val = get_val(cb, "RO", "Pure_water", "pos")
    add("RO", "AEL", pw_val)          # Pure_water from RO feeds AEL

    # ── 5. Hydrogen & CO2 inputs to FTL, Methanol_synthesis, Ammonia_synthesis ─
    for tech in ["FTL", "Methanol_synthesis", "Ammonia_synthesis"]:
        h2 = get_val(cb, tech, "Hydrogen", "neg")
        add("Hydrogen", tech, h2)

    for tech in ["FTL", "Methanol_synthesis"]:
        co2 = get_val(cb, tech, "co", "neg")
        add("DAC", tech, co2)

    # ── 6. Heat production: ST_N → Heat, HP → Heat ───────────────────────────
    for tech in ["ST_N", "HP"]:
        val = get_val(cb, tech, "Heat", "pos")
        add(tech, "Heat", val)

    # ── 7. Industry_EL → T_Industry_EH (written as Industry_EL → Heat) ────────
    val = get_val(cb, "Industry_EL", "T_Industry_EH", "pos")
    add("Industry_EL", "Heat", val)

    # ── 8. DW_Electric_converter → DHW_el (written as → Heat) ────────────────
    val = get_val(cb, "DW_Electric_converter", "DHW_el", "pos")
    add("DW_Electric_converter", "Heat", val)

    # ── 9. DW_LPG_converter → DHW_LPG (written as → Heat) ───────────────────
    val = get_val(cb, "DW_LPG_converter", "DHW_LPG", "pos")
    add("DW_LPG_converter", "Heat", val)

    # ── 10. DW_LPG_converter → LPG (negative value = fuel input) ─────────────
    val = get_val(cb, "DW_LPG_converter", "LPG", "neg")
    add("DW_LPG_converter", "LPG", val)

    # ── 11. Heat → DAC (DAC consumes heat; negative in DAC row) ──────────────
    val = get_val(cb, "DAC", "Heat", "neg")
    add("Heat", "DAC", val)

    # ── 12. Demand flows ──────────────────────────────────────────────────────
    # Power → Residential demand  (Demand tech, Elec commodity)
    val = get_val(cb, "Demand", "Elec", "neg")
    add("Power", "Residential demand", val)

    # Power → Transportation demand  (sum of electric transport sub-commodities)
    transport_el_commodities = [
        "T_MDV_el", "T_Bus_el", "T_Two_wheel_el",
        "T_HDV_el", "T_LDV_el", "T_ship_el", "T_Aviation_el",
    ]
    transport_val = 0.0
    for comm in transport_el_commodities:
        transport_val += get_val(cb, "Demand", comm, "neg")
    add("Power", "Transportation demand", transport_val)

    # Heat → Heating demand  (T_cook_el + T_cook_LPG + DHW_LPG + T_Industry_EH)
    heat_demand_commodities = ["T_cook_el", "T_cook_LPG", "DHW_LPG", "T_Industry_EH", "DHW_el"]
    heat_demand_val = 0.0
    for comm in heat_demand_commodities:
        heat_demand_val += get_val(cb, "Demand", comm, "neg")
    add("Heat", "Heating demand", heat_demand_val)

    # FTL → Transportation demand  (eKerosene)
    val = get_val(cb, "Demand", "eKerosene", "neg")
    add("FTL", "Transportation demand", val)

    # Ammonia_synthesis → Transportation demand
    # Dummy_Ammonia - Dummy_EL: +ve value; written as Ammonia_synthesis - Transportation demand
    dummy_ammonia = get_val(cb, "Dummy_Ammonia", "Dummy_EL", "pos")
    add("Ammonia_synthesis", "Transportation demand", dummy_ammonia)

    # Methanol_synthesis → Transportation demand
    # Two sources ADDED together:
    #   (a) Demand tech, Methanol commodity: -ve value (demand-derived)
    #   (b) Dummy_Methanol - Dummy_EL: +ve value
    demand_methanol = get_val(cb, "Demand", "Methanol", "neg")
    dummy_methanol  = get_val(cb, "Dummy_Methanol", "Dummy_EL", "pos")
    add("Methanol_synthesis", "Transportation demand", demand_methanol + dummy_methanol)

    # ── 13. cook_el: Power consumed, then delivered as Heat ──────────────────
    # cook_el - Elec: -ve value → Power - cook_el
    cook_el_power = get_val(cb, "cook_el", "Elec", "neg")
    add("Power", "cook_el", cook_el_power)
    # cook_el - Heat: same value as above → cook_el - Heat
    add("cook_el", "Heat", cook_el_power)

    return rows


def write_excel(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sankey 2050"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2F5496")
    cell_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Input", "Output", "Value (GWh)"]
    col_widths = [30, 30, 15]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill("solid", start_color="EBF0FA")

    for r_idx, row in enumerate(rows, start=2):
        fill = alt_fill if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(["Input", "Output", "Value (GWh)"], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[key])
            cell.font = cell_font
            cell.border = border
            cell.alignment = center if c_idx == 3 else left
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(output_path)
    log.info("Excel saved → %s  (%d rows)", output_path, len(rows))


def main(gdx_path: str, output_path: str) -> None:
    cb = load_cb(gdx_path)
    rows = build_rows(cb)

    log.info("Total rows to write: %d", len(rows))
    for r in rows:
        log.debug("  %s → %s : %.2f", r["Input"], r["Output"], r["Value (GWh)"])

    write_excel(rows, output_path)
    print(f"\nDone! {len(rows)} rows written to: {output_path}")
    print("\nExtracted flows:")
    print(f"{'Input':<30} {'Output':<30} {'Value (GWh)':>12}")
    print("-" * 74)
    for r in rows:
        print(f"{r['Input']:<30} {r['Output']:<30} {r['Value (GWh)']:>12.4f}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract Sankey flows from GDX to Excel")
    parser.add_argument("--gdx", default=GDX_PATH, help="Path to GDX file")
    parser.add_argument("--out", default=OUTPUT_PATH, help="Output Excel path")
    args = parser.parse_args()
    main(args.gdx, args.out)
####################AEL capacities##################################################
"""
Pacific Island Countries (PICs) — AEL Installed Capacity (2040 & 2050)
=======================================================================
Reads 'converter_caps' from the GAMS GDX file and plots the total
installed capacity of the AEL electrolyser (Hydrogen commodity) for
all 14 islands across 2040 and 2050.

    Source : converter_caps | techs == 'AEL' | commodity == 'Hydrogen' | capType == 'total'
    Unit   : GW → MW (× 1000)

Output
------
  figures/S_23/ael_capacity_minload/
      ael_capacity_overview.png
"""

"""
Pacific Island Countries (PICs) — AEL Installed Capacity (2040 & 2050)
=======================================================================
One figure per island — simple grouped bar chart showing AEL total
installed capacity for 2040 and 2050.

    Source : converter_caps | techs == 'AEL' | commodity == 'Hydrogen' | capType == 'total'
    Unit   : GW → MW (× 1000)
"""

import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.patches import Patch

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/ael_capacity_minload"
YEARS      = ["2040", "2050"]
ISLANDS    = [
    "CI_model", "FJ_model", "FSM_model", "KB_model",  "MI_model",
    "NU_model", "NE_model", "PU_model",  "PNG_model", "SA_model",
    "SI_model", "TA_model", "TU_model",  "VU_model",
]

AEL_TECH      = "AEL"
AEL_COMMODITY = "Hydrogen"

# Colorbrewer-safe, print-friendly
YEAR_COLORS = {
    "2040": "#2196F3",   # medium blue
    "2050": "#FF5722",   # deep orange
}

BAR_WIDTH = 0.45

rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         10,
    "axes.titlesize":    11,
    "axes.labelsize":    10,
    "xtick.labelsize":   10,
    "ytick.labelsize":   9,
    "legend.fontsize":   9,
    "axes.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "xtick.direction":   "out",
    "ytick.direction":   "out",
    "xtick.major.size":  3,
    "ytick.major.size":  3,
    "grid.linewidth":    0.4,
    "grid.alpha":        0.35,
    "grid.color":        "#888888",
    "figure.dpi":        120,
    "savefig.dpi":       200,
    "savefig.bbox":      "tight",
    "savefig.facecolor": "white",
})


# ─────────────────────────────────────────────────────────────────────────────
# DATA
# ─────────────────────────────────────────────────────────────────────────────

def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    return gdxpds.to_dataframes(path)


def extract_ael_capacity(data: dict) -> pd.DataFrame:
    cc   = data["converter_caps"]
    mask = (
        (cc["techs"]     == AEL_TECH) &
        (cc["commodity"] == AEL_COMMODITY) &
        (cc["capType"]   == "total")
    )
    df = cc.loc[mask].copy()

    if df.empty:
        log.warning("No rows found — check tech/commodity/capType names.")
        return pd.DataFrame(columns=["island", "year", "capacity_mw"])

    df.rename(columns={"accNodesModel": "island",
                        "accYears":      "year",
                        "Value":         "capacity_mw"}, inplace=True)
    df["year"]        = df["year"].astype(str)
    df["capacity_mw"] = df["capacity_mw"] * 1000.0   # GW → MW
    df = df[df["year"].isin(YEARS)]
    log.info("Rows extracted: %d", len(df))
    return df[["island", "year", "capacity_mw"]]


# ─────────────────────────────────────────────────────────────────────────────
# PER-ISLAND FIGURE
# ─────────────────────────────────────────────────────────────────────────────

def plot_island(df: pd.DataFrame, island: str, out_dir: Path, fmt: str, dpi: int):
    island_label = island.replace("_model", "")

    sub = df[df["island"] == island]
    # Build value dict {year: MW}, default 0 if not deployed
    values = {y: 0.0 for y in YEARS}
    for _, row in sub.iterrows():
        values[row["year"]] = row["capacity_mw"]

    x       = np.arange(len(YEARS))
    heights = [values[y] for y in YEARS]
    colors  = [YEAR_COLORS[y] for y in YEARS]
    y_max   = max(heights) if max(heights) > 0 else 1.0

    fig, ax = plt.subplots(figsize=(5, 4.5))

    bars = ax.bar(x, heights, BAR_WIDTH, color=colors,
                  edgecolor="white", linewidth=0.4, zorder=3)

    # Value labels above bars
    for bar, val in zip(bars, heights):
        if val > 0:
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + y_max * 0.02,
                f"{val:,.0f}",
                ha="center", va="bottom", fontsize=9, color="#333333",
            )

    ax.set_xticks(x)
    ax.set_xticklabels(YEARS, fontsize=10)
    ax.set_xlabel("Year", labelpad=4)
    ax.set_ylabel("Installed capacity (MW)", labelpad=4)
    ax.set_title(f"{island_label} — AEL Total Installed Capacity",
                 fontweight="bold", pad=8)
    ax.set_xlim(-0.6, len(YEARS) - 0.4)
    ax.set_ylim(0, y_max * 1.18)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))

    handles = [Patch(facecolor=YEAR_COLORS[y], edgecolor="white",
                     linewidth=0.3, label=y) for y in YEARS]
    ax.legend(handles=handles, frameon=False, fontsize=9,
              title="Year", title_fontsize=9)

    plt.tight_layout()
    save_path = out_dir / f"{island}.{fmt}"
    fig.savefig(save_path, dpi=dpi)
    plt.close(fig)
    log.info("  Saved -> %s", save_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(gdx_path, out_dir, fmt, dpi, preloaded_data=None):
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)
    df   = extract_ael_capacity(data)

    if df.empty:
        log.error("No data extracted.")
        return

    for island in ISLANDS:
        log.info("Plotting %s ...", island)
        plot_island(df, island, out_path, fmt, dpi)

    log.info("Done. Figures saved to: %s", out_path)
    print(f"\n14 island figures written to: {out_path}/")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--gdx", default=GDX_PATH)
    parser.add_argument("--out", default=OUTPUT_DIR)
    parser.add_argument("--fmt", default="png",
                        choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi", default=200, type=int)
    args = parser.parse_args()
    main(args.gdx, args.out, args.fmt, args.dpi)
    
    #############################Trade flow practice figures###############################

########################################
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.path import Path
import matplotlib.colors as mc
import matplotlib.patheffects as pe
import math as _math

# ─────────────────────────────────────────────────────────────────────────────
# DATA
# ─────────────────────────────────────────────────────────────────────────────
COUNTRIES = [
    "Solomon Islands", "Vanuatu", "Papua New Guinea",
    "Kiribati", "Fiji", "Fed. States Micronesia",
    "Tonga", "Niue", "Samoa",
    "Tuvalu", "Nauru", "Cook Islands",
    "Marshall Islands", "Palau",
]
N   = len(COUNTRIES)
IDX = {c: i for i, c in enumerate(COUNTRIES)}

FLOWS_RAW = [
    ("Solomon Islands",        "Vanuatu",                  103.889),
    ("Kiribati",               "Solomon Islands",           21.0983),
    ("Kiribati",               "Vanuatu",                   17.5141),
    ("Fiji",                   "Kiribati",                  16.5211),
    ("Fiji",                   "Vanuatu",                    8.51074),
    ("Papua New Guinea",       "Solomon Islands",            1.38939),
    ("Fed. States Micronesia", "Papua New Guinea",           0.54899),
    ("Niue",                   "Tonga",                      0.00360621),
    ("Cook Islands",           "Samoa",                      0.0027766),
    ("Samoa",                  "Tonga",                      0.00262711),
    ("Cook Islands",           "Niue",                       0.00212216),
    ("Marshall Islands",       "Tuvalu",                     0.000899626),
    ("Cook Islands",           "Tonga",                      0.000840416),
    ("Kiribati",               "Marshall Islands",           0.000302717),
    ("Fed. States Micronesia", "Marshall Islands",           0.000226764),
    ("Papua New Guinea",       "Palau",                      0.000262955),
    ("Fiji",                   "Nauru",                      0.000383604),
    ("Fed. States Micronesia", "Palau",                      0.000632451),
    ("Fiji",                   "Samoa",                      0.00155268),
    ("Kiribati",               "Tuvalu",                     0.00165708),
    ("Fiji",                   "Tuvalu",                     0.00166598),
    ("Fiji",                   "Tonga",                      0.00757608),
]

# Filter out flows below threshold
FLOWS_RAW = [(src, dst, vol) for src, dst, vol in FLOWS_RAW if vol >= 0.0010]

FLOW_DIR = {}
for src, dst, _ in FLOWS_RAW:
    i, j = IDX[src], IDX[dst]
    FLOW_DIR[(i, j)] = True
    FLOW_DIR[(j, i)] = False

VOL = np.zeros((N, N))
for src, dst, vol in FLOWS_RAW:
    i, j = IDX[src], IDX[dst]
    VOL[i][j] = vol
    VOL[j][i] = vol

log_vals        = [np.log10(v) for _, _, v in FLOWS_RAW]
lv_min, lv_max  = min(log_vals), max(log_vals)

# ── compute total flow per country (matches badge value) ──────────────────────
country_total = {}
for src, dst, vol in FLOWS_RAW:
    country_total[src] = country_total.get(src, 0) + vol
    country_total[dst] = country_total.get(dst, 0) + vol

# ── tight round scale ceiling — based on TOTAL flow ───────────────────────────
def nice_cap(val):
    if val <= 0:
        return 1
    mag  = 10 ** _math.floor(_math.log10(val))
    step = mag / 10
    return _math.ceil(val / step) * step

SCALE_CAP = {c: nice_cap(country_total.get(c, 1)) for c in COUNTRIES}

def fmt_gwh(v):
    if v == 0:      return "0"
    if v >= 10_000: return f"{v/1000:.1f}k"
    if v >= 1_000:  return f"{v:.0f}"
    if v >= 100:    return f"{v:.0f}"
    if v >= 10:     return f"{v:.0f}"
    if v >= 1:      return f"{v:.1f}"
    if v >= 0.1:    return f"{v:.2f}"
    return f"{v:.3f}"

# ─────────────────────────────────────────────────────────────────────────────
# PALETTE
# ─────────────────────────────────────────────────────────────────────────────
PALETTE = [
    "#E05A2B",  # Solomon Islands
    "#C0392B",  # Vanuatu
    "#8B1A1A",  # Papua New Guinea
    "#D4820A",  # Kiribati
    "#2471A3",  # Fiji
    "#117864",  # Fed. States Micronesia
    "#6C3483",  # Tonga
    "#1A5276",  # Niue
    "#1E8449",  # Samoa
    "#2E86C1",  # Tuvalu
    "#A04000",  # Nauru
    "#5D6D7E",  # Cook Islands
    "#839192",  # Marshall Islands
    "#AAB7B8",  # Palau
]

# ─────────────────────────────────────────────────────────────────────────────
# GEOMETRY
# ─────────────────────────────────────────────────────────────────────────────
PAD   = 0.055
R_OUT = 1.00
R_IN  = 0.88

arc_span = (2 * np.pi - N * PAD) / N

def compute_equal_arcs():
    starts, ends = [], []
    a = np.pi / 2
    for _ in range(N):
        starts.append(a)
        ends.append(a + arc_span)
        a += arc_span + PAD
    return np.array(starts), np.array(ends)

starts, ends = compute_equal_arcs()

LOG_OFF = 1.0
def log_scaled(vol):
    return np.log(vol + LOG_OFF) / np.log(10)

def compute_sub_arcs():
    sub = {}
    for i in range(N):
        t0, t1 = starts[i], ends[i]
        span   = t1 - t0
        nbrs   = [(j, VOL[i][j]) for j in range(N)
                  if j != i and VOL[i][j] > 0]
        if not nbrs:
            continue
        total_vol = sum(v for _, v in nbrs)
        cursor    = t0
        for j, v in nbrs:
            frac = v / total_vol
            sub[(i, j)] = (cursor, cursor + frac * span)
            cursor += frac * span
    return sub

subs = compute_sub_arcs()

def ribbon_path(a0, a1, b0, b1, r=R_IN, na=60, nb=40):
    th  = np.linspace(a0, a1, na)
    src = np.c_[r*np.cos(th), r*np.sin(th)]
    th  = np.linspace(b0, b1, na)
    tgt = np.c_[r*np.cos(th), r*np.sin(th)]
    def quad(p, q, n):
        u = np.linspace(0, 1, n)[:, None]
        return (1-u)**2*p + 2*u*(1-u)*np.zeros(2) + u**2*q
    e1  = quad(src[-1], tgt[0], nb)
    e2  = quad(tgt[-1], src[0], nb)
    pts = np.vstack([src, e1[1:], tgt, e2[1:]])
    codes = [Path.MOVETO] + [Path.LINETO]*(len(pts)-2) + [Path.CLOSEPOLY]
    return Path(pts, codes)

def arc_band_path(t0, t1, ro=R_OUT, ri=R_IN, n=300):
    th = np.linspace(t0, t1, n)
    pts = np.vstack([
        np.c_[np.cos(th)*ro, np.sin(th)*ro],
        np.c_[np.cos(th[::-1])*ri, np.sin(th[::-1])*ri],
    ])
    codes = [Path.MOVETO] + [Path.LINETO]*(len(pts)-2) + [Path.CLOSEPOLY]
    return Path(pts, codes)

def darker(hex_c, f=0.35):
    return tuple(np.array(mc.to_rgb(hex_c)) * (1 - f))

def bezier_tip(src_exit, dst_entry, t, r=R_IN):
    P0 = np.array([r*np.cos(src_exit),  r*np.sin(src_exit)])
    P2 = np.array([r*np.cos(dst_entry), r*np.sin(dst_entry)])
    C  = np.zeros(2)
    pt   = (1-t)**2*P0 + 2*t*(1-t)*C + t**2*P2
    tang = 2*(1-t)*(C - P0) + 2*t*(P2 - C)
    return pt, tang

# ─────────────────────────────────────────────────────────────────────────────
# TICK SCALE
# ─────────────────────────────────────────────────────────────────────────────
N_TICKS  = 5
TICK_R0  = R_OUT + 0.010
TICK_R1  = R_OUT + 0.032
TICK_R1E = R_OUT + 0.048

def draw_ticks(ax, t0, t1, cap, color):
    base_rgb = np.array(mc.to_rgb(color))
    tick_angs = np.linspace(t0, t1, N_TICKS)
    for k, ang in enumerate(tick_angs):
        is_edge = (k == 0 or k == N_TICKS - 1)
        r1  = TICK_R1E if is_edge else TICK_R1
        lw  = 1.5 if is_edge else 0.7
        blend = 1.0 - 0.5*(1.0 - abs(2*k/(N_TICKS-1) - 1))
        tc  = tuple(1.0 - (1.0 - base_rgb)*blend)
        ax.plot([np.cos(ang)*TICK_R0, np.cos(ang)*r1],
                [np.sin(ang)*TICK_R0, np.sin(ang)*r1],
                color=tc, lw=lw, solid_capstyle="round", zorder=5)

    R_TK_LBL = R_OUT + 0.075
    # White halo for tick labels — crisp on white background
    shadow = [pe.withStroke(linewidth=1.8, foreground="#ffffff")]
    for ang, val_str in [(t0, "0"), (t1, fmt_gwh(cap))]:
        mid_deg = np.degrees(ang) % 360
        rot = (mid_deg + 90) % 360
        if 90 < rot < 270: rot = (rot+180) % 360
        ax.text(np.cos(ang)*R_TK_LBL, np.sin(ang)*R_TK_LBL,
                val_str,
                color=color, fontsize=6.5, ha="center", va="center",
                rotation=rot, rotation_mode="anchor",
                fontweight="bold", fontfamily="monospace",
                path_effects=shadow, alpha=0.90, zorder=7)

# ─────────────────────────────────────────────────────────────────────────────
# LABEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
SWAP_ROWS = {
    "Solomon Islands",
    "Marshall Islands",
    "Cook Islands",
    "Papua New Guinea",
    "Fed. States Micronesia",
}

LABEL_OVERRIDE = {
    "Solomon Islands":       ("Solomon",     "Islands"),
    "Papua New Guinea":      ("Papua",       "New Guinea"),
    "Marshall Islands":      ("Marshall",    "Islands"),
    "Cook Islands":          ("Cook",        "Islands"),
    "Fed. States Micronesia":("Fed. States", "Micronesia"),
}

def draw_label(ax, name, mid_angle, color, idx):
    R_LBL_IN  = R_OUT + 0.105
    R_LBL_OUT = R_OUT + 0.168

    mid_deg = np.degrees(mid_angle) % 360
    rot = (mid_deg + 90) % 360
    if 90 < rot < 270: rot = (rot+180) % 360
    # White halo — clean on white paper, still separates from coloured arcs
    shadow = [pe.withStroke(linewidth=2.5, foreground="#ffffff")]

    if name in LABEL_OVERRIDE:
        l1, l2 = LABEL_OVERRIDE[name]
        if name in SWAP_ROWS:
            l1, l2 = l2, l1
        for txt, r in [(l1, R_LBL_IN), (l2, R_LBL_OUT)]:
            ax.text(np.cos(mid_angle)*r, np.sin(mid_angle)*r, txt,
                    color=color, fontsize=8.8, ha="center", va="center",
                    rotation=rot, rotation_mode="anchor",
                    fontweight="700", fontfamily="sans-serif",
                    path_effects=shadow, zorder=8)
    else:
        words = name.split()
        if len(words) > 1:
            best, bd = 1, 999
            for k in range(1, len(words)):
                d = abs(len(" ".join(words[:k])) - len(" ".join(words[k:])))
                if d < bd: bd, best = d, k
            l1 = " ".join(words[:best])
            l2 = " ".join(words[best:])
            for txt, r in [(l1, R_LBL_IN), (l2, R_LBL_OUT)]:
                ax.text(np.cos(mid_angle)*r, np.sin(mid_angle)*r, txt,
                        color=color, fontsize=8.8, ha="center", va="center",
                        rotation=rot, rotation_mode="anchor",
                        fontweight="700", fontfamily="sans-serif",
                        path_effects=shadow, zorder=8)
        else:
            ax.text(np.cos(mid_angle)*R_LBL_IN, np.sin(mid_angle)*R_LBL_IN,
                    name, color=color, fontsize=9.2, ha="center", va="center",
                    rotation=rot, rotation_mode="anchor",
                    fontweight="700", fontfamily="sans-serif",
                    path_effects=shadow, zorder=8)

# ─────────────────────────────────────────────────────────────────────────────
# FIGURE — white background
# ─────────────────────────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(13, 13), facecolor="white")
ax.set_facecolor("white")
ax.set_aspect("equal")
ax.set_xlim(-2.10, 2.10)
ax.set_ylim(-2.10, 2.10)
ax.axis("off")

# depth rings
for r in [0.28, 0.52, 0.76]:
    ax.add_patch(plt.Circle((0,0), r, fill=False,
                              color="#00000015", lw=0.5, zorder=0))

# ── RIBBONS ───────────────────────────────────────────────────────────────────
for src_n, dst_n, vol in FLOWS_RAW:
    i, j = IDX[src_n], IDX[dst_n]
    si = subs.get((i, j))
    sj = subs.get((j, i))
    if si is None or sj is None: continue
    path  = ribbon_path(*si, *sj)
    lv    = np.log10(vol)
    alpha = 0.12 + 0.55*((lv - lv_min)/(lv_max - lv_min))
    c     = PALETTE[i]
    ax.add_patch(mpatches.PathPatch(
        path, facecolor=c, edgecolor=darker(c, 0.15),
        linewidth=0.3, alpha=alpha, zorder=1))

# ── ARROWS ────────────────────────────────────────────────────────────────────
ARROW_T = 0.78
for src_n, dst_n, vol in FLOWS_RAW:
    i, j = IDX[src_n], IDX[dst_n]
    si = subs.get((i, j))
    sj = subs.get((j, i))
    if si is None or sj is None: continue

    flow_i_to_j = FLOW_DIR.get((i, j), True)
    src_exit, dst_entry = (si[1], sj[0]) if flow_i_to_j else (sj[1], si[0])
    arrow_color = PALETTE[i] if flow_i_to_j else PALETTE[j]

    tip, tang = bezier_tip(src_exit, dst_entry, ARROW_T)
    tnorm = np.linalg.norm(tang)
    if tnorm < 1e-9: continue
    tang /= tnorm

    lv    = np.log10(vol)
    scale = 0.25 + 0.75*((lv - lv_min)/(lv_max - lv_min))
    tail  = tip - tang*(0.028 + 0.048*scale)

    ax.annotate("", xy=tip, xytext=tail,
        arrowprops=dict(arrowstyle="-|>", color=arrow_color,
                        lw=0.5 + 1.1*scale,
                        mutation_scale=5 + 11*scale),
        zorder=4)

# ── ARC BANDS + TICKS + LABELS ────────────────────────────────────────────────
for i in range(N):
    t0, t1 = starts[i], ends[i]
    c    = PALETTE[i]
    name = COUNTRIES[i]
    cap  = SCALE_CAP[name]

    ax.add_patch(mpatches.PathPatch(
        arc_band_path(t0, t1), facecolor=c, edgecolor="none",
        alpha=0.93, zorder=2))

    th = np.linspace(t0, t1, 300)
    ax.plot(np.cos(th)*R_IN,  np.sin(th)*R_IN,
            color="white", lw=0.5, alpha=0.18, zorder=3)
    ax.plot(np.cos(th)*R_OUT, np.sin(th)*R_OUT,
            color="white", lw=0.3, alpha=0.10, zorder=3)

    draw_ticks(ax, t0, t1, cap, c)

    mid_angle = (t0 + t1) / 2
    total_gwh = country_total.get(name, 0)
    badge_str = fmt_gwh(total_gwh) + " GWh"
    mid_deg   = np.degrees(mid_angle) % 360
    rot = (mid_deg + 90) % 360
    if 90 < rot < 270: rot = (rot+180) % 360
    # Badge text: white on coloured arc band — dark halo for maximum contrast
    # (text always sits on the coloured band, not on white paper)
    ax.text(np.cos(mid_angle)*(R_IN + 0.040),
            np.sin(mid_angle)*(R_IN + 0.040),
            badge_str, color="white", fontsize=7.0,
            ha="center", va="center",
            rotation=rot, rotation_mode="anchor",
            fontfamily="monospace", fontweight="bold", alpha=1.0, zorder=5,
            path_effects=[
                pe.withStroke(linewidth=2.5, foreground=darker(c, 0.55)),
            ])

    draw_label(ax, name, mid_angle, c, i)

# ── CENTRE ────────────────────────────────────────────────────────────────────
ax.text(0,  0.08, "PACIFIC",
        color="#111111", fontsize=19, ha="center",
        fontweight="900", fontfamily="sans-serif", alpha=0.92, zorder=5)
ax.text(0, -0.08, "AMMONIA\nTRADE 2050",
        color="#444444", fontsize=11, ha="center",
        fontweight="400", fontfamily="sans-serif",
        linespacing=1.6, zorder=5)

# ── LEGEND ────────────────────────────────────────────────────────────────────
lx, ly = -2.02, -1.68
ax.text(lx, ly+0.22, "Flow volume tiers",
        color="#222222", fontsize=8.5, fontweight="700",
        fontfamily="sans-serif")
tiers = [
    (">1,000 GWh  (major)",  "#E05A2B", 0.65),
    ("10–1,000 GWh  (mid)",  "#D4820A", 0.45),
    ("<10 GWh  (minor)",     "#2471A3", 0.25),
]
for label, col, alpha in tiers:
    ax.add_patch(mpatches.FancyBboxPatch(
        (lx, ly), 0.13, 0.08,
        boxstyle="round,pad=0.01",
        facecolor=col, alpha=alpha, edgecolor="none"))
    ax.text(lx+0.17, ly+0.040, label,
            color="#222222", fontsize=7.8, va="center",
            fontfamily="sans-serif")
    ly -= 0.115

ax.annotate("", xy=(lx+0.14, ly-0.025), xytext=(lx, ly-0.025),
            arrowprops=dict(arrowstyle="-|>", color="#333333",
                            lw=1.1, mutation_scale=8))
ax.text(lx+0.17, ly-0.025, "Ammonia flow direction",
        color="#333333", fontsize=7.8, va="center",
        fontfamily="sans-serif")

# footnote
ax.text(2.02, -1.88,
        "Equal arc size per nation\n"
        "Ribbon width ∝ trade volume (linear)\n"
        "Tick scale: 0 → total flow per nation",
        color="#555555", fontsize=6.5, ha="right", va="bottom",
        fontfamily="sans-serif", linespacing=1.8)

plt.tight_layout(pad=0)
out = "pacific_ammonia_white.png"
plt.savefig(out, dpi=300, bbox_inches="tight", facecolor="white")
print(f"Saved: {out}")
########################################################################Efuel connectivity#################
import csv
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.path import Path
import matplotlib.colors as mc
import matplotlib.patheffects as pe
import math as _math

# ─────────────────────────────────────────────────────────────────────────────
# STATIC DATA (identical to original figure — countries, palette, geometry)
# ─────────────────────────────────────────────────────────────────────────────
COUNTRIES = [
    "Solomon Islands", "Vanuatu", "Papua New Guinea",
    "Kiribati", "Fiji", "Fed. States Micronesia",
    "Tonga", "Niue", "Samoa",
    "Tuvalu", "Nauru", "Cook Islands",
    "Marshall Islands", "Palau",
]
N   = len(COUNTRIES)
IDX = {c: i for i, c in enumerate(COUNTRIES)}

CODE_TO_NAME = {
    "SI": "Solomon Islands",
    "VU": "Vanuatu",
    "PNG": "Papua New Guinea",
    "KB": "Kiribati",
    "FJ": "Fiji",
    "FSM": "Fed. States Micronesia",
    "TA": "Tonga",
    "NU": "Niue",
    "SA": "Samoa",
    "TU": "Tuvalu",
    "NE": "Nauru",
    "CI": "Cook Islands",
    "MI": "Marshall Islands",
    "PU": "Palau",
}

COMMODITY_LABEL = {
    "Ammonia": "AMMONIA",
    "eKerosene": "eKEROSENE",
    "Methanol": "METHANOL",
}

PALETTE = [
    "#E05A2B",  # Solomon Islands
    "#C0392B",  # Vanuatu
    "#8B1A1A",  # Papua New Guinea
    "#D4820A",  # Kiribati
    "#2471A3",  # Fiji
    "#117864",  # Fed. States Micronesia
    "#6C3483",  # Tonga
    "#1A5276",  # Niue
    "#1E8449",  # Samoa
    "#2E86C1",  # Tuvalu
    "#A04000",  # Nauru
    "#5D6D7E",  # Cook Islands
    "#839192",  # Marshall Islands
    "#AAB7B8",  # Palau
]

PAD   = 0.055
R_OUT = 1.00
R_IN  = 0.88
arc_span = (2 * np.pi - N * PAD) / N

def compute_equal_arcs():
    starts, ends = [], []
    a = np.pi / 2
    for _ in range(N):
        starts.append(a)
        ends.append(a + arc_span)
        a += arc_span + PAD
    return np.array(starts), np.array(ends)

starts, ends = compute_equal_arcs()

LOG_OFF = 1.0

def fmt_gwh(v):
    if v == 0:      return "0"
    if v >= 10_000: return f"{v/1000:.1f}k"
    if v >= 1_000:  return f"{v:.0f}"
    if v >= 100:    return f"{v:.0f}"
    if v >= 10:     return f"{v:.0f}"
    if v >= 1:      return f"{v:.1f}"
    if v >= 0.1:    return f"{v:.2f}"
    return f"{v:.3f}"

def nice_cap(val):
    if val <= 0:
        return 1
    mag  = 10 ** _math.floor(_math.log10(val))
    step = mag / 10
    return _math.ceil(val / step) * step

def ribbon_path(a0, a1, b0, b1, r=R_IN, na=60, nb=40):
    th  = np.linspace(a0, a1, na)
    src = np.c_[r*np.cos(th), r*np.sin(th)]
    th  = np.linspace(b0, b1, na)
    tgt = np.c_[r*np.cos(th), r*np.sin(th)]
    def quad(p, q, n):
        u = np.linspace(0, 1, n)[:, None]
        return (1-u)**2*p + 2*u*(1-u)*np.zeros(2) + u**2*q
    e1  = quad(src[-1], tgt[0], nb)
    e2  = quad(tgt[-1], src[0], nb)
    pts = np.vstack([src, e1[1:], tgt, e2[1:]])
    codes = [Path.MOVETO] + [Path.LINETO]*(len(pts)-2) + [Path.CLOSEPOLY]
    return Path(pts, codes)

def arc_band_path(t0, t1, ro=R_OUT, ri=R_IN, n=300):
    th = np.linspace(t0, t1, n)
    pts = np.vstack([
        np.c_[np.cos(th)*ro, np.sin(th)*ro],
        np.c_[np.cos(th[::-1])*ri, np.sin(th[::-1])*ri],
    ])
    codes = [Path.MOVETO] + [Path.LINETO]*(len(pts)-2) + [Path.CLOSEPOLY]
    return Path(pts, codes)

def darker(hex_c, f=0.35):
    return tuple(np.array(mc.to_rgb(hex_c)) * (1 - f))

def bezier_tip(src_exit, dst_entry, t, r=R_IN):
    P0 = np.array([r*np.cos(src_exit),  r*np.sin(src_exit)])
    P2 = np.array([r*np.cos(dst_entry), r*np.sin(dst_entry)])
    C  = np.zeros(2)
    pt   = (1-t)**2*P0 + 2*t*(1-t)*C + t**2*P2
    tang = 2*(1-t)*(C - P0) + 2*t*(P2 - C)
    return pt, tang

N_TICKS  = 5
TICK_R0  = R_OUT + 0.010
TICK_R1  = R_OUT + 0.032
TICK_R1E = R_OUT + 0.048

def draw_ticks(ax, t0, t1, cap, color):
    base_rgb = np.array(mc.to_rgb(color))
    tick_angs = np.linspace(t0, t1, N_TICKS)
    for k, ang in enumerate(tick_angs):
        is_edge = (k == 0 or k == N_TICKS - 1)
        r1  = TICK_R1E if is_edge else TICK_R1
        lw  = 1.5 if is_edge else 0.7
        blend = 1.0 - 0.5*(1.0 - abs(2*k/(N_TICKS-1) - 1))
        tc  = tuple(1.0 - (1.0 - base_rgb)*blend)
        ax.plot([np.cos(ang)*TICK_R0, np.cos(ang)*r1],
                [np.sin(ang)*TICK_R0, np.sin(ang)*r1],
                color=tc, lw=lw, solid_capstyle="round", zorder=5)

    R_TK_LBL = R_OUT + 0.075
    shadow = [pe.withStroke(linewidth=1.8, foreground="#ffffff")]
    for ang, val_str in [(t0, "0"), (t1, fmt_gwh(cap))]:
        mid_deg = np.degrees(ang) % 360
        rot = (mid_deg + 90) % 360
        if 90 < rot < 270: rot = (rot+180) % 360
        ax.text(np.cos(ang)*R_TK_LBL, np.sin(ang)*R_TK_LBL,
                val_str,
                color=color, fontsize=6.5, ha="center", va="center",
                rotation=rot, rotation_mode="anchor",
                fontweight="bold", fontfamily="monospace",
                path_effects=shadow, alpha=0.90, zorder=7)

SWAP_ROWS = {
    "Solomon Islands",
    "Marshall Islands",
    "Cook Islands",
    "Papua New Guinea",
    "Fed. States Micronesia",
}

LABEL_OVERRIDE = {
    "Solomon Islands":       ("Solomon",     "Islands"),
    "Papua New Guinea":      ("Papua",       "New Guinea"),
    "Marshall Islands":      ("Marshall",    "Islands"),
    "Cook Islands":          ("Cook",        "Islands"),
    "Fed. States Micronesia":("Fed. States", "Micronesia"),
}

def draw_label(ax, name, mid_angle, color, idx):
    R_LBL_IN  = R_OUT + 0.105
    R_LBL_OUT = R_OUT + 0.168

    mid_deg = np.degrees(mid_angle) % 360
    rot = (mid_deg + 90) % 360
    if 90 < rot < 270: rot = (rot+180) % 360
    shadow = [pe.withStroke(linewidth=2.5, foreground="#ffffff")]

    if name in LABEL_OVERRIDE:
        l1, l2 = LABEL_OVERRIDE[name]
        if name in SWAP_ROWS:
            l1, l2 = l2, l1
        for txt, r in [(l1, R_LBL_IN), (l2, R_LBL_OUT)]:
            ax.text(np.cos(mid_angle)*r, np.sin(mid_angle)*r, txt,
                    color=color, fontsize=8.8, ha="center", va="center",
                    rotation=rot, rotation_mode="anchor",
                    fontweight="700", fontfamily="sans-serif",
                    path_effects=shadow, zorder=8)
    else:
        words = name.split()
        if len(words) > 1:
            best, bd = 1, 999
            for k in range(1, len(words)):
                d = abs(len(" ".join(words[:k])) - len(" ".join(words[k:])))
                if d < bd: bd, best = d, k
            l1 = " ".join(words[:best])
            l2 = " ".join(words[best:])
            for txt, r in [(l1, R_LBL_IN), (l2, R_LBL_OUT)]:
                ax.text(np.cos(mid_angle)*r, np.sin(mid_angle)*r, txt,
                        color=color, fontsize=8.8, ha="center", va="center",
                        rotation=rot, rotation_mode="anchor",
                        fontweight="700", fontfamily="sans-serif",
                        path_effects=shadow, zorder=8)
        else:
            ax.text(np.cos(mid_angle)*R_LBL_IN, np.sin(mid_angle)*R_LBL_IN,
                    name, color=color, fontsize=9.2, ha="center", va="center",
                    rotation=rot, rotation_mode="anchor",
                    fontweight="700", fontfamily="sans-serif",
                    path_effects=shadow, zorder=8)

# ─────────────────────────────────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────────────────────────────────
groups = {}  # (commodity, year) -> list of (src_name, dst_name, signed_value)
with open("flows_data.csv") as f:
    reader = csv.reader(f)
    for row in reader:
        src_code, dst_code, link, year, port, commodity, netlabel, val = row
        src_name = CODE_TO_NAME[src_code.replace("_model", "")]
        dst_name = CODE_TO_NAME[dst_code.replace("_model", "")]
        val = float(val)
        key = (commodity, year)
        groups.setdefault(key, []).append((src_name, dst_name, val))

# ─────────────────────────────────────────────────────────────────────────────
# FIGURE GENERATOR (identical configuration to original script)
# ─────────────────────────────────────────────────────────────────────────────
def generate_figure(flows_signed, commodity, year, out_path):
    # Apply sign convention: +ve = start->end (src->dst), -ve = end->start (dst->src)
    flows_raw_full = []
    for src, dst, v in flows_signed:
        if v >= 0:
            flows_raw_full.append((src, dst, abs(v)))
        else:
            flows_raw_full.append((dst, src, abs(v)))

    # Same threshold filter as original script
    FLOWS_RAW = [(s, d, vol) for s, d, vol in flows_raw_full if vol >= 0.0010]

    FLOW_DIR = {}
    for src, dst, _ in FLOWS_RAW:
        i, j = IDX[src], IDX[dst]
        FLOW_DIR[(i, j)] = True
        FLOW_DIR[(j, i)] = False

    VOL = np.zeros((N, N))
    for src, dst, vol in FLOWS_RAW:
        i, j = IDX[src], IDX[dst]
        VOL[i][j] = vol
        VOL[j][i] = vol

    log_vals = [np.log10(v) for _, _, v in FLOWS_RAW]
    lv_min, lv_max = min(log_vals), max(log_vals)

    country_total = {}
    for src, dst, vol in FLOWS_RAW:
        country_total[src] = country_total.get(src, 0) + vol
        country_total[dst] = country_total.get(dst, 0) + vol

    SCALE_CAP = {c: nice_cap(country_total.get(c, 1)) for c in COUNTRIES}

    def compute_sub_arcs():
        sub = {}
        for i in range(N):
            t0, t1 = starts[i], ends[i]
            span   = t1 - t0
            nbrs   = [(j, VOL[i][j]) for j in range(N)
                      if j != i and VOL[i][j] > 0]
            if not nbrs:
                continue
            total_vol = sum(v for _, v in nbrs)
            cursor    = t0
            for j, v in nbrs:
                frac = v / total_vol
                sub[(i, j)] = (cursor, cursor + frac * span)
                cursor += frac * span
        return sub

    subs = compute_sub_arcs()

    fig, ax = plt.subplots(figsize=(13, 13), facecolor="white")
    ax.set_facecolor("white")
    ax.set_aspect("equal")
    ax.set_xlim(-2.10, 2.10)
    ax.set_ylim(-2.10, 2.10)
    ax.axis("off")

    for r in [0.28, 0.52, 0.76]:
        ax.add_patch(plt.Circle((0,0), r, fill=False,
                                  color="#00000015", lw=0.5, zorder=0))

    # ── RIBBONS ──
    for src_n, dst_n, vol in FLOWS_RAW:
        i, j = IDX[src_n], IDX[dst_n]
        si = subs.get((i, j))
        sj = subs.get((j, i))
        if si is None or sj is None: continue
        path  = ribbon_path(*si, *sj)
        lv    = np.log10(vol)
        alpha = 0.12 + 0.55*((lv - lv_min)/(lv_max - lv_min)) if lv_max > lv_min else 0.4
        c     = PALETTE[i]
        ax.add_patch(mpatches.PathPatch(
            path, facecolor=c, edgecolor=darker(c, 0.15),
            linewidth=0.3, alpha=alpha, zorder=1))

    # ── ARROWS ──
    ARROW_T = 0.78
    for src_n, dst_n, vol in FLOWS_RAW:
        i, j = IDX[src_n], IDX[dst_n]
        si = subs.get((i, j))
        sj = subs.get((j, i))
        if si is None or sj is None: continue

        flow_i_to_j = FLOW_DIR.get((i, j), True)
        src_exit, dst_entry = (si[1], sj[0]) if flow_i_to_j else (sj[1], si[0])
        arrow_color = PALETTE[i] if flow_i_to_j else PALETTE[j]

        tip, tang = bezier_tip(src_exit, dst_entry, ARROW_T)
        tnorm = np.linalg.norm(tang)
        if tnorm < 1e-9: continue
        tang /= tnorm

        lv    = np.log10(vol)
        scale = 0.25 + 0.75*((lv - lv_min)/(lv_max - lv_min)) if lv_max > lv_min else 0.6
        tail  = tip - tang*(0.028 + 0.048*scale)

        ax.annotate("", xy=tip, xytext=tail,
            arrowprops=dict(arrowstyle="-|>", color=arrow_color,
                            lw=0.5 + 1.1*scale,
                            mutation_scale=5 + 11*scale),
            zorder=4)

    # ── ARC BANDS + TICKS + LABELS ──
    for i in range(N):
        t0, t1 = starts[i], ends[i]
        c    = PALETTE[i]
        name = COUNTRIES[i]
        cap  = SCALE_CAP[name]

        ax.add_patch(mpatches.PathPatch(
            arc_band_path(t0, t1), facecolor=c, edgecolor="none",
            alpha=0.93, zorder=2))

        th = np.linspace(t0, t1, 300)
        ax.plot(np.cos(th)*R_IN,  np.sin(th)*R_IN,
                color="white", lw=0.5, alpha=0.18, zorder=3)
        ax.plot(np.cos(th)*R_OUT, np.sin(th)*R_OUT,
                color="white", lw=0.3, alpha=0.10, zorder=3)

        draw_ticks(ax, t0, t1, cap, c)

        mid_angle = (t0 + t1) / 2
        total_gwh = country_total.get(name, 0)
        badge_str = fmt_gwh(total_gwh) + " GWh"
        mid_deg   = np.degrees(mid_angle) % 360
        rot = (mid_deg + 90) % 360
        if 90 < rot < 270: rot = (rot+180) % 360
        ax.text(np.cos(mid_angle)*(R_IN + 0.040),
                np.sin(mid_angle)*(R_IN + 0.040),
                badge_str, color="white", fontsize=7.0,
                ha="center", va="center",
                rotation=rot, rotation_mode="anchor",
                fontfamily="monospace", fontweight="bold", alpha=1.0, zorder=5,
                path_effects=[
                    pe.withStroke(linewidth=2.5, foreground=darker(c, 0.55)),
                ])

        draw_label(ax, name, mid_angle, c, i)

    # ── CENTRE (commodity/year reflect this figure) ──
    ax.text(0,  0.08, "PACIFIC",
            color="#111111", fontsize=19, ha="center",
            fontweight="900", fontfamily="sans-serif", alpha=0.92, zorder=5)
    ax.text(0, -0.08, f"{COMMODITY_LABEL[commodity]}\nTRADE {year}",
            color="#444444", fontsize=11, ha="center",
            fontweight="400", fontfamily="sans-serif",
            linespacing=1.6, zorder=5)

    # ── LEGEND ──
    lx, ly = -2.02, -1.68
    ax.text(lx, ly+0.22, "Flow volume tiers",
            color="#222222", fontsize=8.5, fontweight="700",
            fontfamily="sans-serif")
    tiers = [
        (">1,000 GWh  (major)",  "#E05A2B", 0.65),
        ("10–1,000 GWh  (mid)",  "#D4820A", 0.45),
        ("<10 GWh  (minor)",     "#2471A3", 0.25),
    ]
    for label, col, alpha in tiers:
        ax.add_patch(mpatches.FancyBboxPatch(
            (lx, ly), 0.13, 0.08,
            boxstyle="round,pad=0.01",
            facecolor=col, alpha=alpha, edgecolor="none"))
        ax.text(lx+0.17, ly+0.040, label,
                color="#222222", fontsize=7.8, va="center",
                fontfamily="sans-serif")
        ly -= 0.115

    ax.annotate("", xy=(lx+0.14, ly-0.025), xytext=(lx, ly-0.025),
                arrowprops=dict(arrowstyle="-|>", color="#333333",
                                lw=1.1, mutation_scale=8))
    ax.text(lx+0.17, ly-0.025, "Ammonia flow direction",
            color="#333333", fontsize=7.8, va="center",
            fontfamily="sans-serif")

    ax.text(2.02, -1.88,
            "Equal arc size per nation\n"
            "Ribbon width \u221d trade volume (linear)\n"
            "Tick scale: 0 \u2192 total flow per nation",
            color="#555555", fontsize=6.5, ha="right", va="bottom",
            fontfamily="sans-serif", linespacing=1.8)

    plt.tight_layout(pad=0)
    plt.savefig(out_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out_path}  ({len(FLOWS_RAW)} flows plotted)")

# ─────────────────────────────────────────────────────────────────────────────
# GENERATE ALL 6 FIGURES
# ─────────────────────────────────────────────────────────────────────────────
for commodity in ["Ammonia", "eKerosene", "Methanol"]:
    for year in ["2040", "2050"]:
        flows_signed = groups.get((commodity, year), [])
        out_path = f"pacific_{commodity.lower()}_{year}.png"
        generate_figure(flows_signed, commodity, year, out_path)
# #     ############################system costs deduction####################################
import argparse
import logging
from pathlib import Path

import gdxpds
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib import rcParams

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s")
log = logging.getLogger(__name__)

GDX_PATH   = r"C:\Local\remix-pic\REMix-Pacific_Island_Countries\Process\results\IP_2050_Final_S23_minload.gdx"
OUTPUT_DIR = "figures/S_23/system_cost"
YEARS      = ["2020", "2030", "2040", "2050"]

COST_COMPONENTS = ["Invest", "OMFix", "OMVar", "FuelCost"]

COMPONENT_COLORS = {
    "Invest":   "#1B3A5C",
    "OMFix":    "#2E6F95",
    "OMVar":    "#3D9DAA",
    "FuelCost": "#C06B3A",
}

COMPONENT_LABELS = {
    "Invest":   "Investment (CAPEX)",
    "OMFix":    "Fixed O&M",
    "OMVar":    "Variable O&M",
    "FuelCost": "Fuel Cost",
}

rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         10,
    "axes.titlesize":    11,
    "axes.labelsize":    10,
    "xtick.labelsize":   10,
    "ytick.labelsize":   10,
    "legend.fontsize":   9,
    "axes.linewidth":    0.8,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "xtick.direction":   "out",
    "ytick.direction":   "out",
    "xtick.major.size":  4,
    "ytick.major.size":  4,
    "grid.linewidth":    0.4,
    "grid.alpha":        0.4,
    "grid.color":        "#AAAAAA",
    "figure.dpi":        120,
    "savefig.dpi":       300,
    "savefig.bbox":      "tight",
    "savefig.facecolor": "white",
})


def load_gdx(path: str) -> dict:
    log.info("Loading GDX: %s", path)
    data = gdxpds.to_dataframes(path)
    log.info("Available symbols: %s", sorted(data.keys()))
    return data


def extract_cost_components(data: dict, years: list) -> pd.DataFrame:
    ia = data["indicator_accounting"]
    log.info("indicator_accounting columns: %s", list(ia.columns))

    all_indicators = COST_COMPONENTS + ["SystemCost"]

    mask = (
        (ia["accNodesModel"].str.lower() == "global") &
        (ia["accYears"].astype(str).isin(years)) &
        (ia["indicator"].isin(all_indicators))
    )
    sub = ia.loc[mask, ["accYears", "indicator", "Value"]].copy()
    sub["accYears"] = sub["accYears"].astype(str)

    log.info("Rows extracted:\n%s", sub.to_string(index=False))

    wide = sub.pivot_table(
        index="accYears", columns="indicator", values="Value", aggfunc="sum"
    ).reset_index()
    wide.rename(columns={"accYears": "year"}, inplace=True)
    wide = wide.sort_values("year").reset_index(drop=True)

    for comp in COST_COMPONENTS:
        if comp not in wide.columns:
            log.warning("Component '%s' not found — filling with 0.", comp)
            wide[comp] = 0.0

    wide["components_sum"] = wide[COST_COMPONENTS].sum(axis=1)
    if "SystemCost" in wide.columns:
        wide["diff_pct"] = (
            (wide["components_sum"] - wide["SystemCost"]) / wide["SystemCost"] * 100
        ).round(2)
        log.info("Component sum vs SystemCost:\n%s",
                  wide[["year", "components_sum", "SystemCost", "diff_pct"]].to_string(index=False))

    return wide


def plot_stacked_cost(df: pd.DataFrame, out_dir: Path, fmt: str, dpi: int):
    if df.empty:
        log.error("No cost data to plot.")
        return

    years  = df["year"].tolist()
    x      = np.arange(len(years))
    bar_w  = 0.52

    # Wider figure to accommodate the external legend on the right
    fig, ax = plt.subplots(figsize=(8.5, 5.5))

    bottoms = np.zeros(len(years))

    for comp in COST_COMPONENTS:
        values = df[comp].fillna(0).values
        bars = ax.bar(
            x, values,
            bottom=bottoms,
            width=bar_w,
            color=COMPONENT_COLORS[comp],
            edgecolor="white",
            linewidth=0.5,
            label=COMPONENT_LABELS[comp],
            zorder=3,
        )
        for bar, val, bot in zip(bars, values, bottoms):
            if val > df["components_sum"].max() * 0.04:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bot + val / 2,
                    f"{val:,.1f}",
                    ha="center", va="center",
                    fontsize=7.5, color="white", fontweight="600",
                )
        bottoms = bottoms + values

    # Total label on top of each bar
    for i, total in enumerate(df["components_sum"].values):
        ax.text(
            x[i],
            total + df["components_sum"].max() * 0.012,
            f"{total:,.1f}",
            ha="center", va="bottom",
            fontsize=9, color="#222222", fontweight="600",
        )

    ax.set_xticks(x)
    ax.set_xticklabels(years, fontsize=10)
    ax.set_xlabel("Year", labelpad=6)
    ax.set_ylabel("Total System Cost (M$)", labelpad=6)
    ax.set_title(
        "Total System Cost by Component — Pacific Island Countries\n"
        "(All Islands, Cumulative)",
        fontsize=11, fontweight="bold", pad=12,
    )

    ax.set_xlim(-0.5, len(years) - 0.5)
    ax.set_ylim(0, df["components_sum"].max() * 1.16)
    ax.yaxis.grid(True, zorder=0)
    ax.set_axisbelow(True)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.spines["left"].set_linewidth(0.8)
    ax.spines["bottom"].set_linewidth(0.8)

    # ── Legend placed OUTSIDE the axes to the right — never covers bars ───────
    ax.legend(
        loc="upper left",
        bbox_to_anchor=(1.02, 1.0),
        borderaxespad=0,
        frameon=True,
        framealpha=0.95,
        edgecolor="#cccccc",
        fancybox=False,
    )

    # Shrink axes to leave room for legend; bbox_inches="tight" captures it all
    fig.subplots_adjust(right=0.78)

    out_dir.mkdir(parents=True, exist_ok=True)
    save_path = out_dir / f"system_cost_stacked.{fmt}"
    fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    log.info("Saved -> %s", save_path)
    print(f"\nFigure saved to: {save_path}")


def main(gdx_path: str, out_dir: str, fmt: str, dpi: int, years: list, preloaded_data=None):
    # if preloaded_data is not None:
    #     data = preloaded_data
    # else:
    #     data = load_data(gdx_path)
    df   = extract_cost_components(data, years)

    if df.empty:
        log.error(
            "No rows matched. Check accNodesModel='global', "
            "indicators %s, years %s exist in indicator_accounting.",
            COST_COMPONENTS, years,
        )
        return

    plot_stacked_cost(df, Path(out_dir), fmt, dpi)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Stacked bar chart of system cost components for Pacific islands.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--gdx",   default=GDX_PATH)
    parser.add_argument("--out",   default=OUTPUT_DIR)
    parser.add_argument("--fmt",   default="png", choices=["png", "pdf", "svg", "tiff"])
    parser.add_argument("--dpi",   default=300, type=int)
    parser.add_argument("--years", nargs="+", default=YEARS)
    args = parser.parse_args()

    main(args.gdx, args.out, args.fmt, args.dpi, args.years)
    ############################NZ import####################################################
# import numpy as np
# import matplotlib
# matplotlib.use("Agg")
# import matplotlib.pyplot as plt
# import matplotlib.patches as mpatches
# from matplotlib.path import Path
# import matplotlib.colors as mc
# import matplotlib.patheffects as pe
# import math as _math

# # ─────────────────────────────────────────────────────────────────────────────
# # DATA
# # ─────────────────────────────────────────────────────────────────────────────
# COUNTRIES = [
#     "Solomon Islands", "Vanuatu", "Papua New Guinea",
#     "Kiribati", "Fiji", "Fed. States Micronesia",
#     "Tonga", "Niue", "Samoa",
#     "Tuvalu", "Nauru", "Cook Islands",
#     "Marshall Islands", "Palau",
# ]
# N = len(COUNTRIES)

# NZ_FLOWS = {
#     "Solomon Islands":        0,
#     "Vanuatu":                450.0,
#     "Papua New Guinea":       800.0,
#     "Kiribati":               300.0,
#     "Fiji":                   600.0,
#     "Fed. States Micronesia": 0.0,
#     "Tonga":                  350.0,
#     "Niue":                   120.0,
#     "Samoa":                  0.0,
#     "Tuvalu":                 100.0,
#     "Nauru":                  0.0,
#     "Cook Islands":           200.0,
#     "Marshall Islands":       0.0,
#     "Palau":                  0.0,
# }

# NZ_TOTAL  = sum(NZ_FLOWS.values())
# _nonzero  = [v for v in NZ_FLOWS.values() if v > 0]
# NZ_lv_min = np.log10(min(_nonzero)) if _nonzero else 0.0
# NZ_lv_max = np.log10(max(_nonzero)) if _nonzero else 1.0

# def nice_cap(val):
#     if val <= 0: return 0
#     mag  = 10 ** _math.floor(_math.log10(val))
#     step = mag / 10
#     return _math.ceil(val / step) * step

# def fmt_gwh(v):
#     if v == 0:      return "0"
#     if v >= 10_000: return f"{v/1000:.1f}k"
#     if v >= 1_000:  return f"{v:.0f}"
#     if v >= 100:    return f"{v:.0f}"
#     if v >= 10:     return f"{v:.0f}"
#     if v >= 1:      return f"{v:.1f}"
#     if v >= 0.1:    return f"{v:.2f}"
#     return f"{v:.3f}"

# # ─────────────────────────────────────────────────────────────────────────────
# # PALETTE  (identical to pacific_ammonia_white.py)
# # ─────────────────────────────────────────────────────────────────────────────
# PALETTE = [
#     "#E05A2B",  # Solomon Islands
#     "#C0392B",  # Vanuatu
#     "#8B1A1A",  # Papua New Guinea
#     "#D4820A",  # Kiribati
#     "#2471A3",  # Fiji
#     "#117864",  # Fed. States Micronesia
#     "#6C3483",  # Tonga
#     "#1A5276",  # Niue
#     "#1E8449",  # Samoa
#     "#2E86C1",  # Tuvalu
#     "#A04000",  # Nauru
#     "#5D6D7E",  # Cook Islands
#     "#839192",  # Marshall Islands
#     "#AAB7B8",  # Palau
# ]
# NZ_COLOR = "#00843D"

# def darker(hex_c, f=0.35):
#     return tuple(np.array(mc.to_rgb(hex_c)) * (1 - f))

# # ─────────────────────────────────────────────────────────────────────────────
# # GEOMETRY — equal arcs for all nations
# # ─────────────────────────────────────────────────────────────────────────────
# PAD          = 0.060
# R_OUT        = 1.00
# R_IN         = 0.88
# NZ_RING_R    = 0.285
# arc_span_val = (2 * np.pi - N * PAD) / N

# def compute_equal_arcs():
#     starts, ends = [], []
#     a = np.pi / 2
#     for _ in range(N):
#         starts.append(a)
#         ends.append(a + arc_span_val)
#         a += arc_span_val + PAD
#     return np.array(starts), np.array(ends)

# starts, ends = compute_equal_arcs()

# # ─────────────────────────────────────────────────────────────────────────────
# # NZ ring sub-arcs — proportional to NZ_FLOWS
# # ─────────────────────────────────────────────────────────────────────────────
# sorted_by_mid = sorted(range(N), key=lambda i: (starts[i] + ends[i]) / 2)

# nz_ring_sub = {}
# cursor = 0.0
# for i in sorted_by_mid:
#     name = COUNTRIES[i]
#     frac = NZ_FLOWS[name] / NZ_TOTAL
#     span = frac * 2 * np.pi
#     nz_ring_sub[name] = (cursor, cursor + span)
#     cursor += span

# # ─────────────────────────────────────────────────────────────────────────────
# # PATH HELPERS
# # ─────────────────────────────────────────────────────────────────────────────
# def arc_band_path(t0, t1, ro, ri, n=300):
#     th  = np.linspace(t0, t1, n)
#     pts = np.vstack([
#         np.c_[np.cos(th)*ro, np.sin(th)*ro],
#         np.c_[np.cos(th[::-1])*ri, np.sin(th[::-1])*ri],
#     ])
#     codes = [Path.MOVETO] + [Path.LINETO]*(len(pts)-2) + [Path.CLOSEPOLY]
#     return Path(pts, codes)

# def nz_ribbon_path(src_a0, src_a1, dst_a0, dst_a1,
#                     r_src=None, r_dst=None, n_arc=80, n_bez=80):
#     if r_src is None: r_src = NZ_RING_R
#     if r_dst is None: r_dst = R_IN

#     th_src  = np.linspace(src_a0, src_a1, n_arc)
#     src_pts = np.c_[r_src * np.cos(th_src), r_src * np.sin(th_src)]

#     th_dst  = np.linspace(dst_a0, dst_a1, n_arc)
#     dst_pts = np.c_[r_dst * np.cos(th_dst), r_dst * np.sin(th_dst)]

#     R_ctrl = (r_src + r_dst) * 0.38

#     def bez_edge(p0, p1, n=n_bez):
#         a0_ = np.arctan2(p0[1], p0[0])
#         a1_ = np.arctan2(p1[1], p1[0])
#         diff = a1_ - a0_
#         if diff >  np.pi: diff -= 2 * np.pi
#         if diff < -np.pi: diff += 2 * np.pi
#         ang_c = a0_ + diff / 2
#         ctrl  = np.array([R_ctrl * np.cos(ang_c), R_ctrl * np.sin(ang_c)])
#         u     = np.linspace(0, 1, n)[:, None]
#         return (1-u)**2 * p0 + 2*u*(1-u)*ctrl + u**2 * p1

#     e_left  = bez_edge(src_pts[0],  dst_pts[0])
#     e_right = bez_edge(src_pts[-1], dst_pts[-1])

#     all_pts = np.vstack([src_pts, e_right[1:], dst_pts[::-1], e_left[::-1][1:]])
#     codes   = [Path.MOVETO] + [Path.LINETO]*(len(all_pts)-2) + [Path.CLOSEPOLY]
#     return Path(all_pts, codes)

# # ─────────────────────────────────────────────────────────────────────────────
# # TICK SCALE  (matching white-bg version)
# # ─────────────────────────────────────────────────────────────────────────────
# N_TICKS  = 5
# TICK_R0  = R_OUT + 0.010
# TICK_R1  = R_OUT + 0.032
# TICK_R1E = R_OUT + 0.048

# def draw_ticks(ax, t0, t1, cap, color):
#     base_rgb  = np.array(mc.to_rgb(color))
#     tick_angs = np.linspace(t0, t1, N_TICKS)
#     for k, ang in enumerate(tick_angs):
#         is_edge = (k == 0 or k == N_TICKS - 1)
#         r1    = TICK_R1E if is_edge else TICK_R1
#         lw    = 1.5 if is_edge else 0.7
#         blend = 1.0 - 0.5*(1.0 - abs(2*k/(N_TICKS-1) - 1))
#         tc    = tuple(1.0 - (1.0 - base_rgb)*blend)
#         ax.plot([np.cos(ang)*TICK_R0, np.cos(ang)*r1],
#                 [np.sin(ang)*TICK_R0, np.sin(ang)*r1],
#                 color=tc, lw=lw, solid_capstyle="round", zorder=5)

#     R_TK_LBL = R_OUT + 0.075
#     shadow    = [pe.withStroke(linewidth=1.8, foreground="#ffffff")]
#     for ang, val_str in [(t0, "0"), (t1, fmt_gwh(cap))]:
#         mid_deg = np.degrees(ang) % 360
#         rot     = (mid_deg + 90) % 360
#         if 90 < rot < 270: rot = (rot+180) % 360
#         ax.text(np.cos(ang)*R_TK_LBL, np.sin(ang)*R_TK_LBL, val_str,
#                 color=color, fontsize=6.5, ha="center", va="center",
#                 rotation=rot, rotation_mode="anchor",
#                 fontweight="bold", fontfamily="monospace",
#                 path_effects=shadow, alpha=0.90, zorder=7)

# # ─────────────────────────────────────────────────────────────────────────────
# # LABELS  (matching white-bg version: white halo, fontsize 8.8/9.2, weight 700)
# # ─────────────────────────────────────────────────────────────────────────────
# SWAP_ROWS = {
#     "Solomon Islands", "Marshall Islands", "Cook Islands",
#     "Papua New Guinea", "Fed. States Micronesia",
# }
# LABEL_OVERRIDE = {
#     "Solomon Islands":        ("Solomon",     "Islands"),
#     "Papua New Guinea":       ("Papua",       "New Guinea"),
#     "Marshall Islands":       ("Marshall",    "Islands"),
#     "Cook Islands":           ("Cook",        "Islands"),
#     "Fed. States Micronesia": ("Fed. States", "Micronesia"),
# }

# def draw_label(ax, name, mid_angle, color):
#     R_LBL_IN  = R_OUT + 0.105
#     R_LBL_OUT = R_OUT + 0.168
#     mid_deg = np.degrees(mid_angle) % 360
#     rot     = (mid_deg + 90) % 360
#     if 90 < rot < 270: rot = (rot+180) % 360
#     shadow = [pe.withStroke(linewidth=2.5, foreground="#ffffff")]

#     def place(txt, r, fs=8.8):
#         ax.text(np.cos(mid_angle)*r, np.sin(mid_angle)*r, txt,
#                 color=color, fontsize=fs, ha="center", va="center",
#                 rotation=rot, rotation_mode="anchor",
#                 fontweight="700", fontfamily="sans-serif",
#                 path_effects=shadow, zorder=8)

#     if name in LABEL_OVERRIDE:
#         l1, l2 = LABEL_OVERRIDE[name]
#         if name in SWAP_ROWS: l1, l2 = l2, l1
#         place(l1, R_LBL_IN)
#         place(l2, R_LBL_OUT)
#     else:
#         words = name.split()
#         if len(words) > 1:
#             best, bd = 1, 999
#             for k in range(1, len(words)):
#                 d = abs(len(" ".join(words[:k])) - len(" ".join(words[k:])))
#                 if d < bd: bd, best = d, k
#             place(" ".join(words[:best]), R_LBL_IN)
#             place(" ".join(words[best:]), R_LBL_OUT)
#         else:
#             place(name, R_LBL_IN, fs=9.2)

# # ─────────────────────────────────────────────────────────────────────────────
# # NZ CENTRAL NODE
# # ─────────────────────────────────────────────────────────────────────────────
# def draw_nz_node(ax):
#     R = NZ_RING_R

#     # Atmospheric glow
#     for r_off, lw, alpha in [(0.038, 14.0, 0.018), (0.022, 8.0, 0.040),
#                                (0.010, 4.0, 0.090), (0.003, 2.0, 0.22)]:
#         ax.add_patch(plt.Circle((0, 0), R+r_off, fill=False,
#                                  color=NZ_COLOR, lw=lw, alpha=alpha, zorder=19))

#     # Light disc (white bg — use very light green instead of near-black)
#     ax.add_patch(plt.Circle((0, 0), R*0.97,
#                               facecolor="#f0f7f2", edgecolor="none",
#                               alpha=1.0, zorder=18))

#     # Colour slices on ring edge
#     for i in sorted_by_mid:
#         name = COUNTRIES[i]
#         a0, a1 = nz_ring_sub[name]
#         th = np.linspace(a0, a1, 80)
#         r0_, r1_ = R * 0.90, R * 1.00
#         pts = np.vstack([
#             np.c_[np.cos(th)*r0_, np.sin(th)*r0_],
#             np.c_[np.cos(th[::-1])*r1_, np.sin(th[::-1])*r1_],
#         ])
#         codes = [Path.MOVETO] + [Path.LINETO]*(len(pts)-2) + [Path.CLOSEPOLY]
#         ax.add_patch(mpatches.PathPatch(
#             Path(pts, codes),
#             facecolor=PALETTE[i], edgecolor="none",
#             alpha=0.80, zorder=20))

#     # Main ring border
#     ax.add_patch(plt.Circle((0, 0), R, fill=False,
#                               color=NZ_COLOR, lw=1.4, alpha=0.92, zorder=21))

#     # 24 tick marks
#     for k in range(24):
#         ang     = 2*np.pi*k/24
#         is_card = (k % 6 == 0)
#         is_mid  = (k % 2 == 0)
#         r0_ = R
#         r1_ = R + (0.034 if is_card else 0.018 if is_mid else 0.010)
#         lw  = 1.2 if is_card else 0.7 if is_mid else 0.4
#         al  = 0.80 if is_card else 0.55 if is_mid else 0.28
#         ax.plot([np.cos(ang)*r0_, np.cos(ang)*r1_],
#                 [np.sin(ang)*r0_, np.sin(ang)*r1_],
#                 color=NZ_COLOR, lw=lw, alpha=al,
#                 solid_capstyle="round", zorder=21)

#     # Labels — dark text on light disc, white halo matching island labels
#     shadow = [pe.withStroke(linewidth=2.5, foreground="#ffffff")]

#     ax.text(0,  R * 0.32, "NEW",
#             color=NZ_COLOR, fontsize=14, ha="center", va="center",
#             fontweight="900", fontfamily="sans-serif",
#             path_effects=shadow, zorder=24, alpha=1.0)

#     hw = R * 0.55
#     ax.plot([-hw, hw], [0, 0],
#             color=NZ_COLOR, lw=0.7, alpha=0.40, zorder=23)

#     ax.text(0, -R * 0.32, "ZEALAND",
#             color=NZ_COLOR, fontsize=9.5, ha="center", va="center",
#             fontweight="900", fontfamily="sans-serif",
#             path_effects=shadow, zorder=24, alpha=1.0)

# # ─────────────────────────────────────────────────────────────────────────────
# # NZ → ISLAND RIBBONS
# # ─────────────────────────────────────────────────────────────────────────────
# def draw_nz_ribbons(ax):
#     for i in sorted_by_mid:
#         name = COUNTRIES[i]
#         vol  = NZ_FLOWS.get(name, 0)
#         if vol <= 0: continue

#         src_a0, src_a1 = nz_ring_sub[name]
#         dst_a0 = starts[i]
#         dst_a1 = ends[i]

#         path = nz_ribbon_path(src_a0, src_a1, dst_a0, dst_a1,
#                                r_src=NZ_RING_R, r_dst=R_IN)

#         if vol <= 0: continue
#         lv    = np.log10(vol)
#         t     = (lv - NZ_lv_min) / max(NZ_lv_max - NZ_lv_min, 1e-9)
#         alpha = float(np.clip(0.28 + 0.45 * t, 0.0, 1.0))

#         isl_rgb  = np.array(mc.to_rgb(PALETTE[i]))
#         nz_rgb   = np.array(mc.to_rgb(NZ_COLOR))
#         fill_col = tuple(isl_rgb * 0.40 + nz_rgb * 0.60)
#         edge_col = tuple(isl_rgb * 0.20 + nz_rgb * 0.80)

#         ax.add_patch(mpatches.PathPatch(
#             path,
#             facecolor=fill_col,
#             edgecolor=edge_col,
#             linewidth=0.20,
#             alpha=alpha,
#             zorder=3,
#         ))



# # ─────────────────────────────────────────────────────────────────────────────
# # FIGURE — white background (matching pacific_ammonia_white.py)
# # ─────────────────────────────────────────────────────────────────────────────
# fig, ax = plt.subplots(figsize=(16, 16), facecolor="white")
# ax.set_facecolor("white")
# ax.set_aspect("equal")
# ax.set_xlim(-1.95, 1.95)
# ax.set_ylim(-1.95, 1.95)
# ax.axis("off")

# # Depth rings — very faint on white
# for r in [0.10, 0.18, 0.26]:
#     ax.add_patch(plt.Circle((0, 0), r, fill=False,
#                               color="#00000015", lw=0.5, zorder=0))

# # 1. Ribbons behind arcs
# draw_nz_ribbons(ax)

# # 2. Island arc bands + ticks + labels
# for i in range(N):
#     t0, t1 = starts[i], ends[i]
#     c      = PALETTE[i]
#     name   = COUNTRIES[i]
#     cap    = nice_cap(NZ_FLOWS.get(name, 1))

#     ax.add_patch(mpatches.PathPatch(
#         arc_band_path(t0, t1, R_OUT, R_IN),
#         facecolor=c, edgecolor="none",
#         alpha=0.93, zorder=7))

#     th = np.linspace(t0, t1, 300)
#     ax.plot(np.cos(th)*R_IN,  np.sin(th)*R_IN,
#             color="white", lw=0.5, alpha=0.18, zorder=8)
#     ax.plot(np.cos(th)*R_OUT, np.sin(th)*R_OUT,
#             color="white", lw=0.3, alpha=0.10, zorder=8)

#     draw_ticks(ax, t0, t1, cap, c)

#     # Badge: show "0 GWh" for no-trade islands; white text on coloured band
#     mid_angle = (t0 + t1) / 2
#     vol_val   = NZ_FLOWS.get(name, 0)
#     badge_str = fmt_gwh(vol_val) + " GWh"
#     mid_deg   = np.degrees(mid_angle) % 360
#     rot       = (mid_deg + 90) % 360
#     if 90 < rot < 270: rot = (rot+180) % 360
#     ax.text(np.cos(mid_angle)*(R_IN + 0.040),
#             np.sin(mid_angle)*(R_IN + 0.040),
#             badge_str, color="white", fontsize=7.0,
#             ha="center", va="center",
#             rotation=rot, rotation_mode="anchor",
#             fontfamily="monospace", fontweight="bold", alpha=1.0, zorder=9,
#             path_effects=[pe.withStroke(linewidth=2.5, foreground=darker(c, 0.55))])

#     draw_label(ax, name, mid_angle, c)

# # 3. NZ node on top
# draw_nz_node(ax)

# # ─────────────────────────────────────────────────────────────────────────────
# # CENTRE TITLE  (matching pacific_ammonia_white.py style)
# # ─────────────────────────────────────────────────────────────────────────────
# # (NZ node occupies centre — title placed just below the node)
# ax.text(0, -(NZ_RING_R + 0.08), "",
#         color="#111111", fontsize=11, ha="center",
#         fontweight="900", fontfamily="sans-serif", alpha=0.92, zorder=5)
# ax.text(0, -(NZ_RING_R + 0.16), "",
#         color="#444444", fontsize=8, ha="center",
#         fontweight="400", fontfamily="sans-serif",
#         linespacing=1.6, zorder=5)

# # ─────────────────────────────────────────────────────────────────────────────
# # LEGEND  (matching pacific_ammonia_white.py style)
# # ─────────────────────────────────────────────────────────────────────────────
# lx, ly = -1.90, -1.62

# ax.text(lx, ly + 0.20, "NZ Ammonia Flows",
#         color="#222222", fontsize=8.5, fontweight="700",
#         fontfamily="sans-serif")

# # Gradient opacity bar
# bar_w, bar_h = 0.38, 0.07
# bar_n = 60
# for bi in range(bar_n):
#     t_    = bi / bar_n
#     frac_ = 0.28 + 0.45 * t_
#     isl_rgb = np.array(mc.to_rgb("#D4820A"))
#     nz_rgb  = np.array(mc.to_rgb(NZ_COLOR))
#     col_    = tuple(isl_rgb * 0.40 + nz_rgb * 0.60)
#     ax.add_patch(mpatches.Rectangle(
#         (lx + bi * bar_w / bar_n, ly),
#         bar_w / bar_n, bar_h,
#         facecolor=col_, alpha=frac_, edgecolor="none"))

# _bar_vals = [v for v in NZ_FLOWS.values() if v > 0]
# _bar_min  = min(_bar_vals) if _bar_vals else 0
# _bar_max  = max(_bar_vals) if _bar_vals else 0
# ax.text(lx,           ly - 0.022, f"{fmt_gwh(_bar_min)} GWh",
#         color="#333333", fontsize=6.5, va="top", fontfamily="monospace")
# ax.text(lx + bar_w,   ly - 0.022, f"{fmt_gwh(_bar_max)} GWh",
#         color="#333333", fontsize=6.5, va="top", ha="right", fontfamily="monospace")
# ax.text(lx + bar_w/2, ly + bar_h + 0.010, "ribbon opacity ∝ flow volume",
#         color="#333333", fontsize=6.5, va="bottom", ha="center",
#         fontfamily="sans-serif")

# ly -= 0.145
# ax.annotate("", xy=(lx+0.14, ly), xytext=(lx, ly),
#             arrowprops=dict(arrowstyle="-|>", color="#333333",
#                             lw=1.1, mutation_scale=8))
# ax.text(lx+0.17, ly, "NZ → island direction",
#         color="#333333", fontsize=7.8, va="center",
#         fontfamily="sans-serif")

# # Footnote
# ax.text(1.90, -1.84,
#         "Equal arc per nation  ·  Ribbon fills full arc  ·  NZ ring slice ∝ flow volume\n"
#         "NZ flows: placeholder values — update NZ_FLOWS dict",
#         color="#555555", fontsize=6.5, ha="right", va="bottom",
#         fontfamily="sans-serif", linespacing=1.8)

# plt.tight_layout(pad=0)
# out = "pacific_ammonia_nz_white.png"
# plt.savefig(out, dpi=600, bbox_inches="tight", facecolor="white")
# print(f"Saved: {out}")
#######################tornedo for S1 vs S2################
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# ─────────────────────────────────────────────────────────────────────────────
# RAW DATA (S1 and S2), format: model,year,tech,carrier,kind,value
# ─────────────────────────────────────────────────────────────────────────────
S1_RAW = """
CI_model,2050,BG_N,Elec,total,0.0043282
FJ_model,2050,BG_N,Elec,total,0.0773755
FSM_model,2050,BG_N,Elec,total,0.00833579
KB_model,2050,BG_N,Elec,total,0.00296663
MI_model,2050,BG_N,Elec,total,0.00606418
NE_model,2050,BG_N,Elec,total,0.000269841
NU_model,2050,BG_N,Elec,total,0.00275552
PNG_model,2050,BG_N,Elec,total,0.393256
PU_model,2050,BG_N,Elec,total,0.000900001
SA_model,2050,BG_N,Elec,total,0.00981532
SI_model,2050,BG_N,Elec,total,0.0140425
TA_model,2050,BG_N,Elec,total,0.00743916
TU_model,2050,BG_N,Elec,total,0.000957496
VU_model,2050,BG_N,Elec,total,0.0137659
PNG_model,2050,Geothermal_B,Elec,total,0.011
FJ_model,2050,Hydro_B,Elec,total,0.0625
FSM_model,2050,Hydro_B,Elec,total,0.000225
PNG_model,2050,Hydro_B,Elec,total,0.115
SA_model,2050,Hydro_B,Elec,total,0.0063
SI_model,2050,Hydro_B,Elec,total,0.00018
VU_model,2050,Hydro_B,Elec,total,0.00054
SI_model,2050,Hydro_N,Elec,total,0.00675
CI_model,2050,PV_N,Elec,total,0.306028
FJ_model,2050,PV_N,Elec,total,4.3852
FSM_model,2050,PV_N,Elec,total,0.966284
KB_model,2050,PV_N,Elec,total,0.503328
MI_model,2050,PV_N,Elec,total,0.19933
NE_model,2050,PV_N,Elec,total,0.0191824
NU_model,2050,PV_N,Elec,total,0.15
PNG_model,2050,PV_N,Elec,total,46.6182
PU_model,2050,PV_N,Elec,total,0.645286
SA_model,2050,PV_N,Elec,total,1.03854
SI_model,2050,PV_N,Elec,total,2.11228
TA_model,2050,PV_N,Elec,total,0.90092
TU_model,2050,PV_N,Elec,total,0.0385623
VU_model,2050,PV_N,Elec,total,1.55543
FJ_model,2050,Wave_N,Elec,total,1.31617e-06
FSM_model,2050,Wave_N,Elec,total,2.56603e-07
MI_model,2050,Wave_N,Elec,total,2.711e-07
PNG_model,2050,Wave_N,Elec,total,2.10817e-07
SA_model,2050,Wave_N,Elec,total,0.0363967
SI_model,2050,Wave_N,Elec,total,1.9874e-07
TA_model,2050,Wave_N,Elec,total,0.0237774
TU_model,2050,Wave_N,Elec,total,0.000993571
VU_model,2050,Wave_N,Elec,total,2.17938e-07
FJ_model,2050,WindOffshore_N,Elec,total,8.60466e-07
FSM_model,2050,WindOffshore_N,Elec,total,7.89072e-08
MI_model,2050,WindOffshore_N,Elec,total,1.62223e-07
PNG_model,2050,WindOffshore_N,Elec,total,1.14246
PU_model,2050,WindOffshore_N,Elec,total,8.28967e-08
SA_model,2050,WindOffshore_N,Elec,total,1.12508e-07
SI_model,2050,WindOffshore_N,Elec,total,0.00223501
TA_model,2050,WindOffshore_N,Elec,total,2.49156e-07
VU_model,2050,WindOffshore_N,Elec,total,1.38047e-07
CI_model,2050,WindOnshore_N,Elec,total,0.0153509
FJ_model,2050,WindOnshore_N,Elec,total,0.222045
FSM_model,2050,WindOnshore_N,Elec,total,0.0506737
KB_model,2050,WindOnshore_N,Elec,total,0.0331596
MI_model,2050,WindOnshore_N,Elec,total,0.0145821
NE_model,2050,WindOnshore_N,Elec,total,0.00102808
NU_model,2050,WindOnshore_N,Elec,total,0.00672199
PNG_model,2050,WindOnshore_N,Elec,total,8.82325e-08
PU_model,2050,WindOnshore_N,Elec,total,0.0522325
SA_model,2050,WindOnshore_N,Elec,total,0.0257704
SI_model,2050,WindOnshore_N,Elec,total,0.150343
TA_model,2050,WindOnshore_N,Elec,total,0.0494258
TU_model,2050,WindOnshore_N,Elec,total,0.00125648
VU_model,2050,WindOnshore_N,Elec,total,0.0375388
CI_model,2050,AEL,Hydrogen,total,0.0269299
FJ_model,2050,AEL,Hydrogen,total,0.369212
FSM_model,2050,AEL,Hydrogen,total,0.080836
KB_model,2050,AEL,Hydrogen,total,0.0503114
MI_model,2050,AEL,Hydrogen,total,0.00804784
NE_model,2050,AEL,Hydrogen,total,0.00151937
NU_model,2050,AEL,Hydrogen,total,0.0130142
PNG_model,2050,AEL,Hydrogen,total,3.72724
PU_model,2050,AEL,Hydrogen,total,0.0577704
SA_model,2050,AEL,Hydrogen,total,0.0925911
SI_model,2050,AEL,Hydrogen,total,0.20116
TA_model,2050,AEL,Hydrogen,total,0.0906472
TU_model,2050,AEL,Hydrogen,total,0.00293863
VU_model,2050,AEL,Hydrogen,total,0.173052
""".strip()

S2_RAW = """
CI_model,2050,BG_N,Elec,total,0.00546013
CI_model,2050,PV_N,Elec,total,0.239174
CI_model,2050,WindOnshore_N,Elec,total,0.0139545
FJ_model,2050,BG_N,Elec,total,0.0817962
FJ_model,2050,Hydro_B,Elec,total,0.0625
FJ_model,2050,PV_N,Elec,total,2.38576
FJ_model,2050,Wave_N,Elec,total,0.0597574
FJ_model,2050,WindOffshore_N,Elec,total,1.86152e-06
FJ_model,2050,WindOnshore_N,Elec,total,0.184344
FSM_model,2050,BG_N,Elec,total,0.00704326
FSM_model,2050,Hydro_B,Elec,total,0.000225
FSM_model,2050,PV_N,Elec,total,1.20573
FSM_model,2050,Wave_N,Elec,total,2.56994e-07
FSM_model,2050,WindOffshore_N,Elec,total,9.04927e-08
FSM_model,2050,WindOnshore_N,Elec,total,0.0283609
KB_model,2050,BG_N,Elec,total,0.00284609
KB_model,2050,PV_N,Elec,total,2.01764
KB_model,2050,WindOnshore_N,Elec,total,0.0133368
MI_model,2050,BG_N,Elec,total,0.00601156
MI_model,2050,PV_N,Elec,total,0.265617
MI_model,2050,Wave_N,Elec,total,2.46299e-07
MI_model,2050,WindOffshore_N,Elec,total,1.6821e-07
MI_model,2050,WindOnshore_N,Elec,total,0.0182311
NE_model,2050,BG_N,Elec,total,0.000533125
NE_model,2050,PV_N,Elec,total,0.00758763
NE_model,2050,WindOnshore_N,Elec,total,0.00102343
NU_model,2050,BG_N,Elec,total,0.00213536
NU_model,2050,PV_N,Elec,total,0.15
NU_model,2050,WindOnshore_N,Elec,total,0.00689564
PNG_model,2050,BG_N,Elec,total,0.452
PNG_model,2050,Geothermal_B,Elec,total,0.011
PNG_model,2050,Hydro_B,Elec,total,0.115
PNG_model,2050,PV_N,Elec,total,29.899
PNG_model,2050,Wave_N,Elec,total,2.24169e-07
PNG_model,2050,WindOffshore_N,Elec,total,1.26316
PNG_model,2050,WindOnshore_N,Elec,total,1.03228e-07
PU_model,2050,BG_N,Elec,total,0.0009
PU_model,2050,PV_N,Elec,total,0.780567
PU_model,2050,WindOffshore_N,Elec,total,9.47039e-08
PU_model,2050,WindOnshore_N,Elec,total,0.0353757
SA_model,2050,BG_N,Elec,total,0.00981538
SA_model,2050,Hydro_B,Elec,total,0.0063
SA_model,2050,PV_N,Elec,total,0.86464
SA_model,2050,Wave_N,Elec,total,0.0469004
SA_model,2050,WindOffshore_N,Elec,total,1.26035e-07
SA_model,2050,WindOnshore_N,Elec,total,0.0215807
SI_model,2050,BG_N,Elec,total,0.0132497
SI_model,2050,Hydro_B,Elec,total,0.00018
SI_model,2050,Hydro_N,Elec,total,0.00675
SI_model,2050,PV_N,Elec,total,1.58719
SI_model,2050,Wave_N,Elec,total,2.34532e-07
SI_model,2050,WindOffshore_N,Elec,total,0.00223506
SI_model,2050,WindOnshore_N,Elec,total,0.136366
TA_model,2050,BG_N,Elec,total,0.00657477
TA_model,2050,PV_N,Elec,total,1.14671
TA_model,2050,Wave_N,Elec,total,0.012722
TA_model,2050,WindOffshore_N,Elec,total,2.71264e-07
TA_model,2050,WindOnshore_N,Elec,total,0.0555545
TU_model,2050,BG_N,Elec,total,0.00100646
TU_model,2050,PV_N,Elec,total,0.0210809
TU_model,2050,Wave_N,Elec,total,0.00104206
TU_model,2050,WindOnshore_N,Elec,total,0.00125607
VU_model,2050,BG_N,Elec,total,0.00889104
VU_model,2050,Hydro_B,Elec,total,0.00054
VU_model,2050,PV_N,Elec,total,14.292
VU_model,2050,Wave_N,Elec,total,2.22458e-07
VU_model,2050,WindOffshore_N,Elec,total,1.49828e-07
VU_model,2050,WindOnshore_N,Elec,total,0.0430727
CI_model,2050,AEL,Hydrogen,total,0.0181044
FJ_model,2050,AEL,Hydrogen,total,0.122464
FSM_model,2050,AEL,Hydrogen,total,0.111432
KB_model,2050,AEL,Hydrogen,total,0.344721
MI_model,2050,AEL,Hydrogen,total,0.0211043
NE_model,2050,AEL,Hydrogen,total,6.66688e-05
NU_model,2050,AEL,Hydrogen,total,0.0116271
PNG_model,2050,AEL,Hydrogen,total,1.67074
PU_model,2050,AEL,Hydrogen,total,0.0718617
SA_model,2050,AEL,Hydrogen,total,0.0746599
SI_model,2050,AEL,Hydrogen,total,0.11784
TA_model,2050,AEL,Hydrogen,total,0.12255
TU_model,2050,AEL,Hydrogen,total,0.00055243
VU_model,2050,AEL,Hydrogen,total,2.82125
""".strip()

CODE_TO_NAME = {
    "SI": "Solomon Islands", "VU": "Vanuatu", "PNG": "Papua New Guinea",
    "KB": "Kiribati", "FJ": "Fiji", "FSM": "Fed. States Micronesia",
    "TA": "Tonga", "NU": "Niue", "SA": "Samoa", "TU": "Tuvalu",
    "NE": "Nauru", "CI": "Cook Islands", "MI": "Marshall Islands", "PU": "Palau",
}

# Technology groups — combine onshore + offshore wind, and the two hydro types
TECH_GROUP = {
    "PV_N":            "Solar PV",
    "WindOnshore_N":   "Wind",
    "WindOffshore_N":  "Wind",
    "BG_N":            "Biomass",
    "Hydro_B":         "Hydro",
    "Hydro_N":         "Hydro",
    "Wave_N":          "Wave",
    "Geothermal_B":    "Geothermal",
    "AEL":             "AEL (Electrolyser)",
}

TECH_ORDER = ["Solar PV", "Wind", "Biomass", "AEL (Electrolyser)"]

TECH_COLORS = {
    "Solar PV":            "#F2A93C",
    "Wind":                "#3B7DD8",
    "Biomass":             "#3F9142",
    "Hydro":               "#17BEBB",
    "Wave":                "#6A4C93",
    "Geothermal":          "#D7263D",
    "AEL (Electrolyser)":  "#E83F94",
}

def parse(raw):
    cap = {}  # country -> group -> value (summed)
    for line in raw.splitlines():
        parts = line.split(",")
        code = parts[0].replace("_model", "")
        tech = parts[2]
        val = float(parts[5])
        country = CODE_TO_NAME[code]
        group = TECH_GROUP.get(tech)
        if group is None:
            continue
        cap.setdefault(country, {})
        cap[country][group] = cap[country].get(group, 0.0) + val
    return cap

cap_s1 = parse(S1_RAW)
cap_s2 = parse(S2_RAW)

COUNTRIES = list(CODE_TO_NAME.values())

# Ignore near-zero base capacities (numerical noise), % change would be meaningless
QUIET_THRESHOLD = 0.001  # GW

data = {}  # country -> {tech: pct_diff}
for c in COUNTRIES:
    data[c] = {}
    for t in TECH_ORDER:
        v1 = cap_s1.get(c, {}).get(t, 0.0)
        v2 = cap_s2.get(c, {}).get(t, 0.0)
        if v1 < QUIET_THRESHOLD:
            continue
        pct = (v2 - v1) / v1 * 100.0
        data[c][t] = pct

# Drop countries with no valid comparisons at all
COUNTRIES = [c for c in COUNTRIES if data[c]]

# Sort alphabetically — reverse order so A→Z reads top-to-bottom in the chart
COUNTRIES_SORTED = sorted(COUNTRIES, reverse=True)

# ─────────────────────────────────────────────────────────────────────────────
# PLOT — symlog x-axis so both small (~1%) and huge (~800%) swings read clearly
# ─────────────────────────────────────────────────────────────────────────────
plt.rcParams["font.family"]     = "DejaVu Sans"
plt.rcParams["axes.edgecolor"]  = "#333333"
plt.rcParams["text.color"]      = "#1a1a1a"
plt.rcParams["axes.labelcolor"] = "#1a1a1a"
plt.rcParams["xtick.color"]     = "#333333"
plt.rcParams["ytick.color"]     = "#1a1a1a"

bar_h = 0.19
gap   = 0.03
LINTHRESH = 10  # linear region within +/-10%, log-compressed beyond that

fig, ax = plt.subplots(figsize=(14.5, 11.5), facecolor="white")
ax.set_facecolor("white")
ax.set_xscale("symlog", linthresh=LINTHRESH, linscale=1.2, base=10)

y_positions = {}
for gi, c in enumerate(COUNTRIES_SORTED):
    techs_present = [t for t in TECH_ORDER if t in data[c]]
    n_here = len(techs_present)
    group_h = n_here * (bar_h + gap)
    base_y = sum(
        len([t for t in TECH_ORDER if t in data[cc]]) * (bar_h + gap) + 0.42
        for cc in COUNTRIES_SORTED[:gi]
    )
    y_positions[c] = (base_y, group_h)

AXIS_LIMIT = 1000

# ── Alternating row bands, one per country, for a clearer visual separation ──
for gi, c in enumerate(COUNTRIES_SORTED):
    base_y, group_h = y_positions[c]
    if gi % 2 == 0:
        ax.axhspan(base_y - 0.18, base_y + group_h - gap + 0.18,
                   color="#F4F5F7", zorder=0, lw=0)

# ── Bars ──
for gi, c in enumerate(COUNTRIES_SORTED):
    base_y, group_h = y_positions[c]
    techs_present = [t for t in TECH_ORDER if t in data[c]]
    for ti, t in enumerate(techs_present):
        y = base_y + ti * (bar_h + gap)
        val = data[c][t]
        color = TECH_COLORS[t]
        ax.barh(y, val, height=bar_h, color=color, edgecolor="white",
                 linewidth=0.5, zorder=3, align="edge")

for gi, c in enumerate(COUNTRIES_SORTED):
    base_y, group_h = y_positions[c]
    mid_y = base_y + group_h/2 - gap/2
    ax.text(-AXIS_LIMIT*1.4, mid_y, c, fontsize=14, fontweight="bold",
             ha="right", va="center", color="#1a1a1a", clip_on=False)
    if gi > 0:
        ax.axhline(base_y - 0.21, color="#dddddd", lw=0.8, zorder=1)

top_y = y_positions[COUNTRIES_SORTED[-1]][0] + y_positions[COUNTRIES_SORTED[-1]][1] + 0.15

ax.axvline(0, color="#111111", lw=1.8, zorder=5)

TICKS = [-1000, -100, -10, 0, 10, 100, 1000]
for gx in TICKS:
    if gx == 0:
        continue
    ax.axvline(gx, color="#eeeeee", lw=0.8, zorder=0)

ax.set_xlim(-AXIS_LIMIT, AXIS_LIMIT)
ax.set_ylim(-0.6, top_y)
ax.set_yticks([])
ax.set_xlabel("Capacity difference — S2 vs S1  (%, symmetric log scale)",
               fontsize=14.5, labelpad=14, fontweight="bold")

ax.set_xticks(TICKS)
ax.set_xticklabels([f"{'+' if x>0 else ''}{x}%" for x in TICKS], fontsize=12.5)

for spine in ["top", "right", "left"]:
    ax.spines[spine].set_visible(False)
ax.spines["bottom"].set_color("#333333")
ax.tick_params(axis="y", length=0)
ax.tick_params(axis="x", length=0)

legend_handles = [plt.Rectangle((0,0), 1, 1, color=TECH_COLORS[t]) for t in TECH_ORDER]
ax.legend(legend_handles, TECH_ORDER, loc="upper left",
           bbox_to_anchor=(1.02, 1.0), ncol=1, frameon=True,
           fontsize=12.5, handlelength=1.5, handleheight=1.5,
           labelspacing=0.7, borderpad=1.0,
           facecolor="white", edgecolor="#cccccc", framealpha=0.95)

fig.text(0.99, 0.01,
          "Islands sorted alphabetically (A\u2192Z, top to bottom)\n"
          "Near-zero base capacities (<0.001 GW) excluded to avoid noisy % values",
          fontsize=9, ha="right", va="bottom", color="#777777", style="italic")

plt.subplots_adjust(top=0.97, bottom=0.10, left=0.22, right=0.84)
out = "capacity_diff_tornado_real_v2.png"
plt.savefig(out, dpi=600, facecolor="white", bbox_inches="tight")
print(f"Saved: {out}")

####################Final global figures for publication###########################
#####################system costs#########################
"""
Combined system cost figure for 14 Pacific Island Countries
=============================================================
Same data as before — restyled to match the shared look of the
generation / capacity / heat figures: same engine (build_x + the
scenario_figure renderer), same group separators, scenario-label
strip, legend placement, and total-label formatting.
"""
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

plt.rcParams["font.family"] = "DejaVu Sans"

# ── Scenario layout (shared with the other combined figures) ─────────────────
BARS = [("IES", 2020), ("IES", 2030), ("IES", 2040), ("IES", 2050),
        ("PES", 2040), ("PES", 2050),
        ("IES-NZ", 2040), ("IES-NZ", 2050),
        ("PES-NZ", 2040), ("PES-NZ", 2050)]
GROUP_GAP = 1.8

# ══════════════════════════════════════════════════════════════════════════════
# DATA — system cost components per scenario / year (values unchanged)
# ══════════════════════════════════════════════════════════════════════════════
RAW_COST = {
    ("IES", 2020):    {"FuelCost": 3209.2,  "OMFix": 215.618},
    ("IES", 2030):    {"FuelCost": 3839.82, "Invest": 389.952, "OMFix": 205.049},
    ("IES", 2040):    {"FuelCost": 2988.42, "Invest": 1004.45, "OMFix": 382.877, "OMVar": 9.66609},
    ("IES", 2050):    {"FuelCost": 13.3338, "Invest": 3158.62, "OMFix": 1055.49, "OMVar": 48.4734},

    ("PES", 2040):    {"FuelCost": 2990.53, "Invest": 1000.68, "OMFix": 381.255, "OMVar": 9.79345},
    ("PES", 2050):    {"FuelCost": 25.894,  "Invest": 3098.23, "OMFix": 1043.19, "OMVar": 54.9363},

    ("IES-NZ", 2040): {"FuelCost": 3006.81, "Invest": 990.256, "OMFix": 378.473, "OMVar": 9.64604},
    ("IES-NZ", 2050): {"FuelCost": 796.492, "Invest": 2494.73, "OMFix": 833.859, "OMVar": 35.1693},

    ("PES-NZ", 2040): {"FuelCost": 2995.83, "Invest": 996.656, "OMFix": 380.074, "OMVar": 9.66601},
    ("PES-NZ", 2050): {"FuelCost": 858.925, "Invest": 2438.83, "OMFix": 813.961, "OMVar": 34.8906},
}

# 2020 / 2030 only exist for IES in the source data (kept as-is; other
# scenarios simply have no bar for those years, same as the source script)

COMP_ORDER = ["Invest", "FuelCost", "OMFix", "OMVar"]     # stacking order, bottom -> top
COMP_LABEL = {
    "Invest":   "Investment (CAPEX)",
    "FuelCost": "Fuel cost",
    "OMFix":    "O&M fixed",
    "OMVar":    "O&M variable",
}
COMP_COLORS = {
    "Invest":   "#2E7D32",  # green
    "FuelCost": "#5AA9E6",  # light blue
    "OMFix":    "#F4A020",  # amber
    "OMVar":    "#8E5CC0",  # purple
}

def merge(raw, scale=1/1000.0):
    """scale converts M$ (source units) -> B$ (display units)"""
    out = {c: 0.0 for c in COMP_ORDER}
    for c, v in raw.items():
        out[c] = v * scale
    return out

COST = {k: merge(v) for k, v in RAW_COST.items()}

# ══════════════════════════════════════════════════════════════════════════════
# ENGINE  (identical structure to the generation / capacity / heat figures)
# ══════════════════════════════════════════════════════════════════════════════
def build_x(bars, gap=GROUP_GAP):
    x, cur, groups, prev = [], 0.0, {}, None
    for sc, yr in bars:
        if prev is not None and sc != prev:
            cur += gap
        x.append(cur); groups.setdefault(sc, []).append(cur)
        cur += 1.0; prev = sc
    return np.array(x), groups


def scenario_figure(DATA, bars, order, colors, label_map, ylabel, title, outfile,
                     broken=True, brk=None, ymax=None,
                     tick_fmt="{:.0f}", total_fmt="{:,.1f}", y_tick_step=None,
                     legend_title="Technology"):
    x, groups = build_x(bars)
    mat = np.array([[DATA[(sc, yr)].get(t, 0.0) for t in order] for sc, yr in bars])
    totals = mat.sum(axis=1)

    # half-gap offset used for every separator (interior AND closing one),
    # so all separators sit an equal visual distance from their neighbouring bars
    sep_offset = (1.0 + GROUP_GAP) / 2.0

    if broken:
        fig, (axt, axb) = plt.subplots(
            2, 1, sharex=True, figsize=(11, 6.6),
            gridspec_kw={"height_ratios": [2.4, 1.0]})
        axes = (axt, axb)
    else:
        fig, axb = plt.subplots(figsize=(11, 6.6)); axt, axes = None, (axb,)

    for ax in axes:
        bottom = np.zeros(len(bars))
        for ti, t in enumerate(order):
            ax.bar(x, mat[:, ti], bottom=bottom, width=0.82,
                   color=colors[t], label=label_map.get(t, t), zorder=3, edgecolor="none")
            bottom += mat[:, ti]

    ref = totals.max()
    if broken:
        axt.set_ylim(brk, (ymax or ref * 1.06))
        axb.set_ylim(0, brk)
        axt.spines["bottom"].set_visible(False); axb.spines["top"].set_visible(False)
        for ax in axes: ax.spines["right"].set_visible(False)
        axt.spines["top"].set_visible(False); axt.tick_params(bottom=False)
        d = 0.008
        kw = dict(transform=axt.transAxes, color="#666666", clip_on=False, linewidth=1)
        axt.plot((-d, +d), (-d, +d), **kw); axt.plot((1 - d, 1 + d), (-d, +d), **kw)
        kw.update(transform=axb.transAxes)
        axb.plot((-d, +d), (1 - d * 3, 1 + d * 3), **kw)
        axb.plot((1 - d, 1 + d), (1 - d * 3, 1 + d * 3), **kw)
        top_ax = axt
        # avoid the break-point value being drawn twice (once on each axis)
        axt.yaxis.set_major_locator(mticker.MaxNLocator(prune="lower"))
    else:
        axb.set_ylim(0, ymax or ref * 1.10)
        axb.spines["top"].set_visible(False); axb.spines["right"].set_visible(False)
        top_ax = axb

    for ax in axes:
        ax.yaxis.grid(True, color="#E5E5E5", linewidth=0.9, zorder=0)
        ax.set_axisbelow(True); ax.tick_params(axis="x", length=0)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: tick_fmt.format(v)))
        if y_tick_step:
            ax.yaxis.set_major_locator(mticker.MultipleLocator(y_tick_step))

    for xi, t in zip(x, totals):
        ax = axt if (broken and t > brk) else axb
        ax.text(xi, t + ref * 0.012, total_fmt.format(t),
                ha="center", va="bottom", fontsize=8.5, color="#333333")

    axb.set_xlim(x.min() - 0.7, x.max() + sep_offset)
    axb.set_xticks(x); axb.set_xticklabels([str(yr) for _, yr in bars], fontsize=9.5)

    fig.supylabel(ylabel, x=0.03, fontsize=12)
    top_ax.set_title(title, fontsize=13.5, fontweight="bold", pad=12,
                     loc="left", color="#1a1a1a")
    h, l = axes[0].get_legend_handles_labels()
    top_ax.legend(h[::-1], l[::-1], loc="upper left", bbox_to_anchor=(1.01, 1.0),
                  frameon=False, fontsize=10, title=legend_title, title_fontsize=11)

    plt.subplots_adjust(left=0.10, right=0.80, top=0.91, bottom=0.17, hspace=0.06)

    # ── Scenario labels + separators — LABEL STRIP ONLY (below the x-axis) ──
    fig.canvas.draw()
    axb_bot = axb.get_position().y0
    line_bottom = axb_bot - 0.115
    label_y     = axb_bot - 0.075

    def d2f(xd):
        disp = axb.transData.transform((xd, 0))
        return fig.transFigure.inverted().transform(disp)[0]

    glist = list(groups.values())
    sep_x = [(max(glist[i]) + min(glist[i + 1])) / 2.0 for i in range(len(glist) - 1)]
    sep_x.append(x.max() + sep_offset)
    for sx in sep_x:
        xf = d2f(sx)
        fig.add_artist(plt.Line2D([xf, xf], [line_bottom, axb_bot],
                       transform=fig.transFigure, color="#BFBFBF",
                       linewidth=1.0, zorder=0.5))

    xL = axb.get_position().x0
    fig.add_artist(plt.Line2D([xL, xL], [line_bottom, axb_bot],
                   transform=fig.transFigure, color="#666666",
                   linewidth=0.9, zorder=0.5))

    for sc, xs in groups.items():
        fig.text(d2f(float(np.mean(xs))), label_y, sc, ha="center", va="top",
                 fontsize=12, fontweight="bold", color="#1a1a1a")

    for ext in ("png", "pdf"):
        fig.savefig(f"{outfile}.{ext}", bbox_inches="tight", facecolor="white")
    print(f"saved {outfile}  | totals: {[round(t,2) for t in totals]}")
    return fig


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    scenario_figure(COST, BARS, COMP_ORDER, COMP_COLORS, COMP_LABEL,
        ylabel="Annual system cost (B$)",
        title="Combined system cost of 14 Pacific Island Countries by scenario",
        outfile="combined_system_cost", broken=False, ymax=5.1,
        tick_fmt="{:.1f}", total_fmt="{:.2f}", y_tick_step=1,
        legend_title="Cost component")

    plt.show()
#######################Electrolyzer capacities global####################################
"""
Electrolyser (AEL) capacity figure for 14 Pacific Island Countries
====================================================================
Same data as before — restyled to match the shared look of the other
combined figures: same engine (build_x + scenario_figure), same group
separators, scenario-label strip, legend placement, and formatting.
"""
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

plt.rcParams["font.family"] = "DejaVu Sans"

# ── Scenario layout ────────────────────────────────────────────────────────
# (All four scenarios have 2040 and 2050 only.)
BARS = [("IES", 2040), ("IES", 2050),
        ("PES", 2040), ("PES", 2050),
        ("IES-NZ", 2040), ("IES-NZ", 2050),
        ("PES-NZ", 2040), ("PES-NZ", 2050)]
GROUP_GAP = 1.8

# ══════════════════════════════════════════════════════════════════════════════
# DATA — electrolyser (AEL) capacity per scenario / year (GW, values unchanged)
# ══════════════════════════════════════════════════════════════════════════════
RAW_AEL = {
    ("IES", 2040):    0.495275, ("IES", 2050):    4.89527,
    ("PES", 2040):    0.484446, ("PES", 2050):    5.50896,
    ("IES-NZ", 2040): 0.454501, ("IES-NZ", 2050): 2.78129,
    ("PES-NZ", 2040): 0.473620, ("PES-NZ", 2050): 2.61814,
}

AEL_ORDER  = ["AEL"]
AEL_LABEL  = {"AEL": "AEL (Electrolyser)"}
AEL_COLORS = {"AEL": "#E83F94"}   # same AEL identity colour used elsewhere

AEL = {k: {"AEL": v} for k, v in RAW_AEL.items()}

# ══════════════════════════════════════════════════════════════════════════════
# ENGINE  (identical structure to the generation / capacity / heat / cost figures)
# ══════════════════════════════════════════════════════════════════════════════
def build_x(bars, gap=GROUP_GAP):
    x, cur, groups, prev = [], 0.0, {}, None
    for sc, yr in bars:
        if prev is not None and sc != prev:
            cur += gap
        x.append(cur); groups.setdefault(sc, []).append(cur)
        cur += 1.0; prev = sc
    return np.array(x), groups


def scenario_figure(DATA, bars, order, colors, label_map, ylabel, title, outfile,
                     broken=True, brk=None, ymax=None,
                     tick_fmt="{:.0f}", total_fmt="{:,.1f}", y_tick_step=None,
                     legend_title="Technology"):
    x, groups = build_x(bars)
    mat = np.array([[DATA[(sc, yr)].get(t, 0.0) for t in order] for sc, yr in bars])
    totals = mat.sum(axis=1)

    # half-gap offset used for every separator (interior AND closing one),
    # so all separators sit an equal visual distance from their neighbouring bars
    sep_offset = (1.0 + GROUP_GAP) / 2.0

    if broken:
        fig, (axt, axb) = plt.subplots(
            2, 1, sharex=True, figsize=(11, 6.6),
            gridspec_kw={"height_ratios": [2.4, 1.0]})
        axes = (axt, axb)
    else:
        fig, axb = plt.subplots(figsize=(11, 6.6)); axt, axes = None, (axb,)

    for ax in axes:
        bottom = np.zeros(len(bars))
        for ti, t in enumerate(order):
            ax.bar(x, mat[:, ti], bottom=bottom, width=0.82,
                   color=colors[t], label=label_map.get(t, t), zorder=3, edgecolor="none")
            bottom += mat[:, ti]

    ref = totals.max()
    if broken:
        axt.set_ylim(brk, (ymax or ref * 1.06))
        axb.set_ylim(0, brk)
        axt.spines["bottom"].set_visible(False); axb.spines["top"].set_visible(False)
        for ax in axes: ax.spines["right"].set_visible(False)
        axt.spines["top"].set_visible(False); axt.tick_params(bottom=False)
        d = 0.008
        kw = dict(transform=axt.transAxes, color="#666666", clip_on=False, linewidth=1)
        axt.plot((-d, +d), (-d, +d), **kw); axt.plot((1 - d, 1 + d), (-d, +d), **kw)
        kw.update(transform=axb.transAxes)
        axb.plot((-d, +d), (1 - d * 3, 1 + d * 3), **kw)
        axb.plot((1 - d, 1 + d), (1 - d * 3, 1 + d * 3), **kw)
        top_ax = axt
        # avoid the break-point value being drawn twice (once on each axis)
        axt.yaxis.set_major_locator(mticker.MaxNLocator(prune="lower"))
    else:
        axb.set_ylim(0, ymax or ref * 1.10)
        axb.spines["top"].set_visible(False); axb.spines["right"].set_visible(False)
        top_ax = axb

    for ax in axes:
        ax.yaxis.grid(True, color="#E5E5E5", linewidth=0.9, zorder=0)
        ax.set_axisbelow(True); ax.tick_params(axis="x", length=0)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: tick_fmt.format(v)))
        if y_tick_step:
            ax.yaxis.set_major_locator(mticker.MultipleLocator(y_tick_step))

    for xi, t in zip(x, totals):
        ax = axt if (broken and t > brk) else axb
        ax.text(xi, t + ref * 0.012, total_fmt.format(t),
                ha="center", va="bottom", fontsize=8.5, color="#333333")

    axb.set_xlim(x.min() - 0.7, x.max() + sep_offset)
    axb.set_xticks(x); axb.set_xticklabels([str(yr) for _, yr in bars], fontsize=9.5)

    fig.supylabel(ylabel, x=0.03, fontsize=12)
    top_ax.set_title(title, fontsize=13.5, fontweight="bold", pad=12,
                     loc="left", color="#1a1a1a")
    h, l = axes[0].get_legend_handles_labels()
    top_ax.legend(h[::-1], l[::-1], loc="upper left", bbox_to_anchor=(1.01, 1.0),
                  frameon=False, fontsize=10, title=legend_title, title_fontsize=11)

    plt.subplots_adjust(left=0.10, right=0.80, top=0.91, bottom=0.17, hspace=0.06)

    # ── Scenario labels + separators — LABEL STRIP ONLY (below the x-axis) ──
    fig.canvas.draw()
    axb_bot = axb.get_position().y0
    line_bottom = axb_bot - 0.115
    label_y     = axb_bot - 0.075

    def d2f(xd):
        disp = axb.transData.transform((xd, 0))
        return fig.transFigure.inverted().transform(disp)[0]

    glist = list(groups.values())
    sep_x = [(max(glist[i]) + min(glist[i + 1])) / 2.0 for i in range(len(glist) - 1)]
    sep_x.append(x.max() + sep_offset)
    for sx in sep_x:
        xf = d2f(sx)
        fig.add_artist(plt.Line2D([xf, xf], [line_bottom, axb_bot],
                       transform=fig.transFigure, color="#BFBFBF",
                       linewidth=1.0, zorder=0.5))

    xL = axb.get_position().x0
    fig.add_artist(plt.Line2D([xL, xL], [line_bottom, axb_bot],
                   transform=fig.transFigure, color="#666666",
                   linewidth=0.9, zorder=0.5))

    for sc, xs in groups.items():
        fig.text(d2f(float(np.mean(xs))), label_y, sc, ha="center", va="top",
                 fontsize=12, fontweight="bold", color="#1a1a1a")

    for ext in ("png", "pdf"):
        fig.savefig(f"{outfile}.{ext}", bbox_inches="tight", facecolor="white")
    print(f"saved {outfile}  | totals: {[round(t,2) for t in totals]}")
    return fig


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    scenario_figure(AEL, BARS, AEL_ORDER, AEL_COLORS, AEL_LABEL,
        ylabel="Electrolyser capacity (GW)",
        title="Electrolyser (AEL) capacity of 14 Pacific Island Countries by scenario",
        outfile="combined_electrolyser_capacity", broken=False, ymax=6.2,
        tick_fmt="{:.0f}", total_fmt="{:.2f}", y_tick_step=1,
        legend_title="Technology")

    plt.show()
#####################Palau E2P #######################################################
"""
Battery energy-to-power (E2P) ratio — Palau
=============================================
Same data as before — restyled to match the shared visual language of
the other combined figures (font, spine/grid colours, title styling,
total-label formatting). No scenario grouping applies here since this
is a single location, so the group-separator/label-strip machinery
from the other figures isn't needed.
"""
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

plt.rcParams["font.family"]     = "DejaVu Sans"
plt.rcParams["text.color"]      = "#1a1a1a"
plt.rcParams["axes.labelcolor"] = "#1a1a1a"
plt.rcParams["xtick.color"]     = "#333333"
plt.rcParams["ytick.color"]     = "#333333"

# ══════════════════════════════════════════════════════════════════════════════
# DATA — battery energy-to-power (E2P) ratio for Palau (values unchanged)
# ══════════════════════════════════════════════════════════════════════════════
DATA = {
    2030: 20.46332905,
    2040: 5.25034786,
    2050: 5.237465427,
}

BAR_COLOR  = "#2A9D8F"
UNIT_LABEL = "Battery E2P ratio (h)"
TITLE      = "Battery energy-to-power (E2P) ratio \u2014 Palau"

years = sorted(DATA.keys())
vals  = [DATA[y] for y in years]
xpos  = np.arange(len(years), dtype=float)

BAR_W = 0.62
fig, ax = plt.subplots(figsize=(7.5, 6.6), facecolor="white")
ax.set_facecolor("white")

ax.bar(xpos, vals, BAR_W, color=BAR_COLOR, edgecolor="none", zorder=3)

ymax = max(vals)
for x, v in zip(xpos, vals):
    ax.text(x, v + ymax * 0.015, f"{v:.2f}", ha="center", va="bottom",
            fontsize=8.5, color="#333333", zorder=4)

# ── AXES (same conventions as the other combined figures) ──
ax.set_ylim(0, ymax * 1.10)
ax.set_xlim(xpos.min() - 0.7, xpos.max() + 0.7)

ax.set_xticks(xpos)
ax.set_xticklabels([str(y) for y in years], fontsize=9.5)

ax.set_ylabel(UNIT_LABEL, fontsize=12, labelpad=10)

ax.yaxis.grid(True, color="#E5E5E5", linewidth=0.9, zorder=0)
ax.set_axisbelow(True)
ax.tick_params(axis="both", length=0)

for spine in ["top", "right"]:
    ax.spines[spine].set_visible(False)
ax.spines["left"].set_color("#999999")
ax.spines["bottom"].set_color("#999999")

ax.set_title(TITLE, fontsize=13.5, fontweight="bold", pad=16,
             loc="left", color="#1a1a1a")

plt.subplots_adjust(left=0.15, right=0.95, top=0.90, bottom=0.10)

for ext in ("png", "pdf"):
    fig.savefig(f"battery_e2p_ratio_palau.{ext}", bbox_inches="tight", facecolor="white")
print("saved battery_e2p_ratio_palau  | values:", [round(v, 2) for v in vals])

plt.show()
################battery cap and output global##########################################
"""
Battery storage capacity & battery energy output — 14 Pacific Island Countries
=================================================================================
Same data as before — restyled to match the shared look of the other
combined figures: same engine (build_x + scenario_figure), same group
separators, scenario-label strip, legend placement, and formatting.
"""
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

plt.rcParams["font.family"] = "DejaVu Sans"

# ── Scenario layout (all four scenarios have 2030, 2040, 2050) ──────────────
BARS = [("IES", 2030), ("IES", 2040), ("IES", 2050),
        ("PES", 2030), ("PES", 2040), ("PES", 2050),
        ("IES-NZ", 2030), ("IES-NZ", 2040), ("IES-NZ", 2050),
        ("PES-NZ", 2030), ("PES-NZ", 2040), ("PES-NZ", 2050)]
GROUP_GAP = 1.8

# ══════════════════════════════════════════════════════════════════════════════
# DATA — per scenario / year (values unchanged)
# ══════════════════════════════════════════════════════════════════════════════
BATTERY_CAPACITY = {   # "total"  — energy capacity (GWh)
    "IES":    {2030: 10.8836, 2040: 38.297,  2050: 150.864},
    "PES":    {2030: 10.8836, 2040: 38.2817, 2050: 142.291},
    "IES-NZ": {2030: 10.8837, 2040: 37.7837, 2050: 116.728},
    "PES-NZ": {2030: 10.8836, 2040: 38.143,  2050: 114.275},
}

BATTERY_OUTPUT_GWH = {  # "positive" — energy output (GWh, as supplied)
    "IES":    {2030: 2634.05, 2040: 9631.58, 2050: 39646.9},
    "PES":    {2030: 2634.05, 2040: 9582.51, 2050: 36819.0},
    "IES-NZ": {2030: 2633.17, 2040: 9444.95, 2050: 29695.7},
    "PES-NZ": {2030: 2634.21, 2040: 9538.81, 2050: 28992.0},
}

# Convert battery output from GWh to TWh
BATTERY_OUTPUT_TWH = {
    scen: {yr: val / 1000.0 for yr, val in years.items()}
    for scen, years in BATTERY_OUTPUT_GWH.items()
}

CAP_ORDER  = ["Battery"]
CAP_LABEL  = {"Battery": "Battery (storage capacity)"}
CAP_COLORS = {"Battery": "#2A9D8F"}
CAPACITY = {(scen, yr): {"Battery": v} for scen, years in BATTERY_CAPACITY.items() for yr, v in years.items()}

OUT_ORDER  = ["Battery"]
OUT_LABEL  = {"Battery": "Battery (energy output)"}
OUT_COLORS = {"Battery": "#E76F51"}
OUTPUT = {(scen, yr): {"Battery": v} for scen, years in BATTERY_OUTPUT_TWH.items() for yr, v in years.items()}

# ══════════════════════════════════════════════════════════════════════════════
# ENGINE  (identical structure to the generation / capacity / heat / cost figures)
# ══════════════════════════════════════════════════════════════════════════════
def build_x(bars, gap=GROUP_GAP):
    x, cur, groups, prev = [], 0.0, {}, None
    for sc, yr in bars:
        if prev is not None and sc != prev:
            cur += gap
        x.append(cur); groups.setdefault(sc, []).append(cur)
        cur += 1.0; prev = sc
    return np.array(x), groups


def scenario_figure(DATA, bars, order, colors, label_map, ylabel, title, outfile,
                     broken=True, brk=None, ymax=None,
                     tick_fmt="{:.0f}", total_fmt="{:,.1f}", y_tick_step=None,
                     legend_title="Technology"):
    x, groups = build_x(bars)
    mat = np.array([[DATA[(sc, yr)].get(t, 0.0) for t in order] for sc, yr in bars])
    totals = mat.sum(axis=1)

    # half-gap offset used for every separator (interior AND closing one),
    # so all separators sit an equal visual distance from their neighbouring bars
    sep_offset = (1.0 + GROUP_GAP) / 2.0

    if broken:
        fig, (axt, axb) = plt.subplots(
            2, 1, sharex=True, figsize=(11, 6.6),
            gridspec_kw={"height_ratios": [2.4, 1.0]})
        axes = (axt, axb)
    else:
        fig, axb = plt.subplots(figsize=(11, 6.6)); axt, axes = None, (axb,)

    for ax in axes:
        bottom = np.zeros(len(bars))
        for ti, t in enumerate(order):
            ax.bar(x, mat[:, ti], bottom=bottom, width=0.82,
                   color=colors[t], label=label_map.get(t, t), zorder=3, edgecolor="none")
            bottom += mat[:, ti]

    ref = totals.max()
    if broken:
        axt.set_ylim(brk, (ymax or ref * 1.06))
        axb.set_ylim(0, brk)
        axt.spines["bottom"].set_visible(False); axb.spines["top"].set_visible(False)
        for ax in axes: ax.spines["right"].set_visible(False)
        axt.spines["top"].set_visible(False); axt.tick_params(bottom=False)
        d = 0.008
        kw = dict(transform=axt.transAxes, color="#666666", clip_on=False, linewidth=1)
        axt.plot((-d, +d), (-d, +d), **kw); axt.plot((1 - d, 1 + d), (-d, +d), **kw)
        kw.update(transform=axb.transAxes)
        axb.plot((-d, +d), (1 - d * 3, 1 + d * 3), **kw)
        axb.plot((1 - d, 1 + d), (1 - d * 3, 1 + d * 3), **kw)
        top_ax = axt
        # avoid the break-point value being drawn twice (once on each axis)
        axt.yaxis.set_major_locator(mticker.MaxNLocator(prune="lower"))
    else:
        axb.set_ylim(0, ymax or ref * 1.10)
        axb.spines["top"].set_visible(False); axb.spines["right"].set_visible(False)
        top_ax = axb

    for ax in axes:
        ax.yaxis.grid(True, color="#E5E5E5", linewidth=0.9, zorder=0)
        ax.set_axisbelow(True); ax.tick_params(axis="x", length=0)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: tick_fmt.format(v)))
        if y_tick_step:
            ax.yaxis.set_major_locator(mticker.MultipleLocator(y_tick_step))

    for xi, t in zip(x, totals):
        ax = axt if (broken and t > brk) else axb
        ax.text(xi, t + ref * 0.012, total_fmt.format(t),
                ha="center", va="bottom", fontsize=8.5, color="#333333")

    axb.set_xlim(x.min() - 0.7, x.max() + sep_offset)
    axb.set_xticks(x); axb.set_xticklabels([str(yr) for _, yr in bars], fontsize=9.5)

    fig.supylabel(ylabel, x=0.03, fontsize=12)
    top_ax.set_title(title, fontsize=13.5, fontweight="bold", pad=12,
                     loc="left", color="#1a1a1a")
    h, l = axes[0].get_legend_handles_labels()
    top_ax.legend(h[::-1], l[::-1], loc="upper left", bbox_to_anchor=(1.01, 1.0),
                  frameon=False, fontsize=10, title=legend_title, title_fontsize=11)

    plt.subplots_adjust(left=0.10, right=0.80, top=0.91, bottom=0.17, hspace=0.06)

    # ── Scenario labels + separators — LABEL STRIP ONLY (below the x-axis) ──
    fig.canvas.draw()
    axb_bot = axb.get_position().y0
    line_bottom = axb_bot - 0.115
    label_y     = axb_bot - 0.075

    def d2f(xd):
        disp = axb.transData.transform((xd, 0))
        return fig.transFigure.inverted().transform(disp)[0]

    glist = list(groups.values())
    sep_x = [(max(glist[i]) + min(glist[i + 1])) / 2.0 for i in range(len(glist) - 1)]
    sep_x.append(x.max() + sep_offset)
    for sx in sep_x:
        xf = d2f(sx)
        fig.add_artist(plt.Line2D([xf, xf], [line_bottom, axb_bot],
                       transform=fig.transFigure, color="#BFBFBF",
                       linewidth=1.0, zorder=0.5))

    xL = axb.get_position().x0
    fig.add_artist(plt.Line2D([xL, xL], [line_bottom, axb_bot],
                   transform=fig.transFigure, color="#666666",
                   linewidth=0.9, zorder=0.5))

    for sc, xs in groups.items():
        fig.text(d2f(float(np.mean(xs))), label_y, sc, ha="center", va="top",
                 fontsize=12, fontweight="bold", color="#1a1a1a")

    for ext in ("png", "pdf"):
        fig.savefig(f"{outfile}.{ext}", bbox_inches="tight", facecolor="white")
    print(f"saved {outfile}  | totals: {[round(t,2) for t in totals]}")
    return fig


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    scenario_figure(CAPACITY, BARS, CAP_ORDER, CAP_COLORS, CAP_LABEL,
        ylabel="Battery capacity (GWh)",
        title="Battery storage capacity of 14 Pacific Island Countries by scenario",
        outfile="battery_storage_capacity", broken=False, ymax=175,
        tick_fmt="{:.0f}", total_fmt="{:.1f}", y_tick_step=25,
        legend_title="Technology")

    scenario_figure(OUTPUT, BARS, OUT_ORDER, OUT_COLORS, OUT_LABEL,
        ylabel="Battery output (TWh)",
        title="Battery energy output of 14 Pacific Island Countries by scenario",
        outfile="battery_energy_output", broken=False, ymax=50,
        tick_fmt="{:.0f}", total_fmt="{:.2f}", y_tick_step=10,
        legend_title="Technology")

    plt.show()
#################global power and generation and power capacity figure##########################################
#!/usr/bin/env python3
#!/usr/bin/env python3
"""
Combined scenario figures for 14 Pacific Island Countries
=========================================================
  1) Combined generation of 14 PICs by scenario          (TWh, broken y-axis)
  2) Combined generation capacity of 14 PICs by scenario (GW,  broken y-axis)
  3) Combined heat generation of 14 PICs by scenario     (TWh, single y-axis)

Generation & capacity use the exact model data (below).
  • Generation values are model GWh -> divided by 1000 to TWh.
  • Capacity is in GW; the Battery entry is EXCLUDED (generation capacity only).
  • Vintages are merged by technology (PV_B+PV_N -> Solar PV, etc.).
Heat data is still approximate (read from the figure) — replace when available.
"""
"""
Combined scenario figures for 14 Pacific Island Countries
=========================================================
  1) Combined generation of 14 PICs by scenario          (TWh, broken y-axis)
  2) Combined generation capacity of 14 PICs by scenario (GW,  broken y-axis)
  3) Combined heat generation of 14 PICs by scenario     (TWh, single y-axis)

Generation & capacity use the exact model data (below).
  • Generation values are model GWh -> divided by 1000 to TWh.
  • Capacity is in GW; the Battery entry is EXCLUDED (generation capacity only).
  • Vintages are merged by technology (PV_B+PV_N -> Solar PV, etc.).
Heat data is still approximate (read from the figure) — replace when available.
"""
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

plt.rcParams["font.family"] = "DejaVu Sans"

# ── Scenario layout (shared) ──────────────────────────────────────────────────
BARS = [("IES", 2020), ("IES", 2030), ("IES", 2040), ("IES", 2050),
        ("PES", 2040), ("PES", 2050),
        ("IES-NZ", 2040), ("IES-NZ", 2050),
        ("PES-NZ", 2040), ("PES-NZ", 2050)]
GROUP_GAP = 1.8

# ── Palettes ──────────────────────────────────────────────────────────────────
GEN_ORDER  = ["Diesel", "Natural gas", "Bioenergy", "Geothermal", "Hydro",
              "Onshore wind", "Offshore wind", "Wave", "Solar PV"]
GEN_COLORS = {"Diesel": "#4D4D4D", "Natural gas": "#9E9E9E", "Bioenergy": "#4D9221",
              "Geothermal": "#B15928", "Hydro": "#1F6FB2", "Onshore wind": "#7FC7E8",
              "Offshore wind": "#2C7FB8", "Wave": "#3FB8AF", "Solar PV": "#F4C430"}

HEAT_ORDER  = ["Diesel boilers", "Water heating (LPG)", "Cooking (biomass)",
               "Cooking (LPG)", "Water heating (electric)", "Cooking (electric)",
               "Electric boilers", "Heat pumps"]
HEAT_COLORS = {"Diesel boilers": "#4D4D4D", "Water heating (LPG)": "#B0B0B0",
               "Cooking (biomass)": "#B5651D", "Cooking (LPG)": "#F2A900",
               "Water heating (electric)": "#8FD0EE", "Cooking (electric)": "#1F6FB2",
               "Electric boilers": "#2CA6A4", "Heat pumps": "#4A9B2F"}

# raw model tech code -> display technology (Battery excluded from figure)
TECH_MAP = {"DG": "Diesel", "NG_plant": "Natural gas",
            "BG_B": "Bioenergy", "BG_N": "Bioenergy",
            "Geothermal_B": "Geothermal",
            "Hydro_B": "Hydro", "Hydro_N": "Hydro",
            "WindOnshore_B": "Onshore wind", "WindOnshore_N": "Onshore wind",
            "WindOffshore_N": "Offshore wind",
            "Wave_N": "Wave", "PV_B": "Solar PV", "PV_N": "Solar PV"}

def merge(raw, scale=1.0, exclude=("Battery",)):
    out = {t: 0.0 for t in GEN_ORDER}
    for code, v in raw.items():
        if code in exclude:
            continue
        out[TECH_MAP[code]] += v * scale
    return out

# ══════════════════════════════════════════════════════════════════════════════
# RAW MODEL DATA
# ══════════════════════════════════════════════════════════════════════════════
# ── Generation (GWh) ─────────────────────────────────────────────────────────
RAW_GEN = {
 ("IES",2020): {"BG_B":516.53,"DG":2185.24,"Geothermal_B":96.36,"Hydro_B":1618.36,"NG_plant":718.32,"PV_B":100.631,"WindOnshore_B":25.241},
 ("IES",2030): {"BG_B":103.3,"BG_N":928.052,"Geothermal_B":95.1499,"Hydro_B":1466.81,"Hydro_N":56.191,"PV_B":69.3065,"PV_N":5085.22,"Wave_N":0.01119,"WindOffshore_N":15.0456,"WindOnshore_B":15.3072,"WindOnshore_N":753.711},
 ("IES",2040): {"BG_N":497.22,"Geothermal_B":95.6465,"Hydro_B":1596.54,"Hydro_N":58.6977,"PV_N":18290.8,"Wave_N":137.464,"WindOffshore_N":1112.05,"WindOnshore_N":1222.91},
 ("IES",2050): {"BG_N":187.14,"Geothermal_B":95.3561,"Hydro_B":1602.74,"Hydro_N":58.6884,"PV_N":79161.5,"Wave_N":248.433,"WindOffshore_N":3788.67,"WindOnshore_N":2132.61},
 ("PES",2040): {"Hydro_N":58.7235,"Geothermal_B":95.766,"Wave_N":141.802,"BG_N":526.793,"WindOffshore_N":1103.25,"WindOnshore_N":1313.04,"Hydro_B":1594.69,"PV_N":18183.6},
 ("PES",2050): {"Hydro_N":58.2561,"Geothermal_B":95.3242,"BG_N":363.425,"Wave_N":457.636,"Hydro_B":1601.7,"WindOnshore_N":1828.37,"WindOffshore_N":4168.98,"PV_N":78909.9},
 ("IES-NZ",2040): {"PV_N":17908.7,"Hydro_B":1595.15,"WindOnshore_N":1205.39,"WindOffshore_N":1108.19,"BG_N":502.778,"Wave_N":146.353,"Geothermal_B":95.3629,"Hydro_N":58.6003},
 ("IES-NZ",2050): {"PV_N":57354.8,"WindOffshore_N":4338.73,"WindOnshore_N":1948.39,"Hydro_B":1597.09,"Wave_N":520.172,"BG_N":317.414,"Geothermal_B":95.3169,"Hydro_N":58.2868},
 ("PES-NZ",2040): {"PV_N":18089.3,"Hydro_B":1594.64,"WindOnshore_N":1294.36,"WindOffshore_N":1102.13,"BG_N":529.937,"Wave_N":146.064,"Geothermal_B":95.766,"Hydro_N":58.723},
 ("PES-NZ",2050): {"PV_N":55933,"WindOffshore_N":4173.3,"WindOnshore_N":1854.4,"Hydro_B":1601.57,"Wave_N":504.114,"BG_N":358.173,"Geothermal_B":95.4022,"Hydro_N":58.2595},
}
# 2020/2030 are identical across scenarios in the model:
_G2020 = {"BG_B":516.53,"DG":2185.24,"Geothermal_B":96.36,"Hydro_B":1618.36,"NG_plant":718.32,"PV_B":100.631,"WindOnshore_B":25.241}
_G2030 = {"BG_B":103.3,"BG_N":928.052,"Geothermal_B":95.1499,"Hydro_B":1466.81,"Hydro_N":56.191,"PV_B":69.3065,"PV_N":5085.22,"Wave_N":0.01119,"WindOffshore_N":15.0456,"WindOnshore_B":15.3072,"WindOnshore_N":753.711}
for _sc in ("PES", "IES-NZ", "PES-NZ"):
    RAW_GEN[(_sc, 2020)] = dict(_G2020)
    RAW_GEN[(_sc, 2030)] = dict(_G2030)

# ── Capacity (GW) — Battery excluded ─────────────────────────────────────────
RAW_CAP = {
 ("IES",2020): {"BG_B":0.0781,"DG":0.79087,"Geothermal_B":0.011,"Hydro_B":0.184745,"NG_plant":0.082,"PV_B":0.0624,"WindOnshore_B":0.00611},
 ("IES",2030): {"BG_B":0.0781,"BG_N":0.467039,"Battery":1.55468,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_B":0.0624,"PV_N":3.99057,"Wave_N":3.81545e-06,"WindOffshore_N":0.00493734,"WindOnshore_B":0.00611,"WindOnshore_N":0.21234},
 ("IES",2040): {"BG_N":0.542271,"Battery":5.1267,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":13.5133,"Wave_N":0.0320707,"WindOffshore_N":0.318528,"WindOnshore_N":0.362889},
 ("IES",2050): {"BG_N":0.542272,"Battery":22.9879,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":59.4386,"Wave_N":0.0611702,"WindOffshore_N":1.1447,"WindOnshore_N":0.660128},
 ("PES",2040): {"BG_N":0.539255,"Battery":5.12089,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":13.2913,"Wave_N":0.0334634,"WindOffshore_N":0.316653,"WindOnshore_N":0.386788},
 ("PES",2050): {"BG_N":0.598263,"Battery":20.7917,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":54.8627,"Wave_N":0.120423,"WindOffshore_N":1.26539,"WindOnshore_N":0.559352},
 ("IES-NZ",2040): {"BG_N":0.544993,"Battery":5.02325,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":13.2308,"Wave_N":0.0343166,"WindOffshore_N":0.317371,"WindOnshore_N":0.35802},
 ("IES-NZ",2050): {"BG_N":0.549654,"Battery":16.3951,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":42.6795,"Wave_N":0.12123,"WindOffshore_N":1.3152,"WindOnshore_N":0.60386},
 ("PES-NZ",2040): {"BG_N":0.540284,"Battery":5.09655,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":13.2259,"Wave_N":0.0343212,"WindOffshore_N":0.316365,"WindOnshore_N":0.38239},
 ("PES-NZ",2050): {"BG_N":0.599313,"Battery":15.9436,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_N":41.0604,"Wave_N":0.132033,"WindOffshore_N":1.26445,"WindOnshore_N":0.566078},
}
_C2020 = {"BG_B":0.0781,"DG":0.79087,"Geothermal_B":0.011,"Hydro_B":0.184745,"NG_plant":0.082,"PV_B":0.0624,"WindOnshore_B":0.00611}
_C2030 = {"BG_B":0.0781,"BG_N":0.467039,"Battery":1.55468,"Geothermal_B":0.011,"Hydro_B":0.184745,"Hydro_N":0.00675,"PV_B":0.0624,"PV_N":3.99057,"Wave_N":3.81545e-06,"WindOffshore_N":0.00493734,"WindOnshore_B":0.00611,"WindOnshore_N":0.21234}
for _sc in ("PES", "IES-NZ", "PES-NZ"):
    RAW_CAP[(_sc, 2020)] = dict(_C2020)
    RAW_CAP[(_sc, 2030)] = dict(_C2030)

# ── Build merged, display-ready dicts ────────────────────────────────────────
GEN = {k: merge(v, scale=1/1000.0) for k, v in RAW_GEN.items()}   # GWh -> TWh
CAP = {k: merge(v) for k, v in RAW_CAP.items()}                    # GW, Battery dropped

# ── Heat (GWh -> TWh) — exact model data ─────────────────────────────────────
# raw model code -> display technology
# (ST_N = solar thermal, negligible <0.01 GWh in every year/scenario -> excluded)
HEAT_TECH_MAP = {
    "cook_b": "Cooking (biomass)",
    "cook_el": "Cooking (electric)",
    "cook_LPG": "Cooking (LPG)",
    "DW_Electric_converter": "Water heating (electric)",
    "DW_LPG_converter": "Water heating (LPG)",
    "Industry": "Diesel boilers",
    "Industry_EL": "Electric boilers",
    "HP": "Heat pumps",
}

def merge_heat(raw, scale=1/1000.0, exclude=("ST_N",)):
    out = {t: 0.0 for t in HEAT_ORDER}
    for code, v in raw.items():
        if code in exclude:
            continue
        out[HEAT_TECH_MAP[code]] += v * scale
    return out

RAW_HEAT = {
 ("IES",2020): {"cook_b":13140,"DW_Electric_converter":478.296,"DW_LPG_converter":478.296,"Industry":7205.98},
 ("IES",2030): {"cook_b":6757.46,"DW_Electric_converter":1226.4,"DW_LPG_converter":1226.4,"Industry":13366.9},
 ("IES",2040): {"cook_b":1171.21,"cook_el":353.904,"cook_LPG":268.056,"DW_Electric_converter":2279.35,"DW_LPG_converter":173.448,"HP":4519.08,"Industry":8936.95,"Industry_EL":1189.73,"ST_N":0.0014272},
 ("IES",2050): {"cook_el":1226.4,"DW_Electric_converter":2444.92,"HP":24938.9,"Industry_EL":1262.06,"ST_N":0.00512178},

 ("PES",2020): {"cook_b":13140,"DW_Electric_converter":478.296,"DW_LPG_converter":478.296,"Industry":7205.98},
 ("PES",2030): {"cook_b":6757.46,"DW_Electric_converter":1226.4,"DW_LPG_converter":1226.4,"Industry":13366.9},
 ("PES",2040): {"cook_b":1171.21,"cook_el":353.904,"cook_LPG":268.056,"DW_Electric_converter":2279.35,"DW_LPG_converter":173.448,"HP":4508.09,"Industry":8936.95,"Industry_EL":1198.25,"ST_N":0.00255668},
 ("PES",2050): {"cook_el":1226.4,"DW_Electric_converter":2444.92,"HP":24529.8,"Industry_EL":1709.81,"ST_N":0.00598301},

 ("IES-NZ",2020): {"cook_b":13140,"DW_Electric_converter":478.296,"DW_LPG_converter":478.296,"Industry":7205.98},
 ("IES-NZ",2030): {"cook_b":6757.46,"DW_Electric_converter":1226.4,"DW_LPG_converter":1226.4,"Industry":13366.9},
 ("IES-NZ",2040): {"cook_b":1171.21,"cook_el":353.904,"cook_LPG":268.056,"DW_Electric_converter":2279.35,"DW_LPG_converter":173.448,"HP":4401.79,"Industry":8936.95,"Industry_EL":1214.86,"ST_N":0.00100706},
 ("IES-NZ",2050): {"cook_el":1226.4,"DW_Electric_converter":2444.92,"HP":19508.6,"Industry_EL":1776.16,"ST_N":0.00486274},

 ("PES-NZ",2020): {"cook_b":13140,"DW_Electric_converter":478.296,"DW_LPG_converter":478.296,"Industry":7205.98},
 ("PES-NZ",2030): {"cook_b":6757.46,"DW_Electric_converter":1226.4,"DW_LPG_converter":1226.4,"Industry":13366.9},
 ("PES-NZ",2040): {"cook_b":1171.21,"cook_el":353.904,"cook_LPG":268.056,"DW_Electric_converter":2279.35,"DW_LPG_converter":173.448,"HP":4469.78,"Industry":8936.95,"Industry_EL":1209.81,"ST_N":0.00180687},
 ("PES-NZ",2050): {"cook_el":1226.4,"DW_Electric_converter":2444.92,"HP":19134.4,"Industry_EL":1776,"ST_N":0.00637089},
}

HEAT = {k: merge_heat(v) for k, v in RAW_HEAT.items()}

# ══════════════════════════════════════════════════════════════════════════════
# ENGINE
# ══════════════════════════════════════════════════════════════════════════════
def build_x(bars, gap=GROUP_GAP):
    x, cur, groups, prev = [], 0.0, {}, None
    for sc, yr in bars:
        if prev is not None and sc != prev:
            cur += gap
        x.append(cur); groups.setdefault(sc, []).append(cur)
        cur += 1.0; prev = sc
    return np.array(x), groups


def scenario_figure(DATA, order, colors, ylabel, title, outfile,
                    broken=True, brk=None, ymax=None,
                    tick_fmt="{:.0f}", total_fmt="{:,.1f}"):
    x, groups = build_x(BARS)
    mat = np.array([[DATA[(sc, yr)].get(t, 0.0) for t in order] for sc, yr in BARS])
    totals = mat.sum(axis=1)

    # half-gap offset used for every separator (interior AND closing one),
    # so all separators sit an equal visual distance from their neighbouring bars
    sep_offset = (1.0 + GROUP_GAP) / 2.0

    if broken:
        fig, (axt, axb) = plt.subplots(
            2, 1, sharex=True, figsize=(11, 6.6),
            gridspec_kw={"height_ratios": [2.4, 1.0]})
        axes = (axt, axb)
    else:
        fig, axb = plt.subplots(figsize=(11, 6.6)); axt, axes = None, (axb,)

    for ax in axes:
        bottom = np.zeros(len(BARS))
        for ti, t in enumerate(order):
            ax.bar(x, mat[:, ti], bottom=bottom, width=0.82,
                   color=colors[t], label=t, zorder=3, edgecolor="none")
            bottom += mat[:, ti]

    ref = totals.max()
    if broken:
        axt.set_ylim(brk, (ymax or ref * 1.06))
        axb.set_ylim(0, brk)
        axt.spines["bottom"].set_visible(False); axb.spines["top"].set_visible(False)
        for ax in axes: ax.spines["right"].set_visible(False)
        axt.spines["top"].set_visible(False); axt.tick_params(bottom=False)
        d = 0.008
        kw = dict(transform=axt.transAxes, color="#666666", clip_on=False, linewidth=1)
        axt.plot((-d, +d), (-d, +d), **kw); axt.plot((1 - d, 1 + d), (-d, +d), **kw)
        kw.update(transform=axb.transAxes)
        axb.plot((-d, +d), (1 - d * 3, 1 + d * 3), **kw)
        axb.plot((1 - d, 1 + d), (1 - d * 3, 1 + d * 3), **kw)
        top_ax = axt
        # the bottom axis already shows the break value (brk) at its top edge;
        # drop the matching lowest tick on the top axis so it isn't drawn twice
        axt.yaxis.set_major_locator(mticker.MaxNLocator(prune="lower"))
    else:
        axb.set_ylim(0, ymax or ref * 1.10)
        axb.spines["top"].set_visible(False); axb.spines["right"].set_visible(False)
        top_ax = axb

    for ax in axes:
        ax.yaxis.grid(True, color="#E5E5E5", linewidth=0.9, zorder=0)
        ax.set_axisbelow(True); ax.tick_params(axis="x", length=0)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: tick_fmt.format(v)))

    for xi, t in zip(x, totals):
        ax = axt if (broken and t > brk) else axb
        ax.text(xi, t + ref * 0.012, total_fmt.format(t),
                ha="center", va="bottom", fontsize=8.5, color="#333333")

    axb.set_xlim(x.min() - 0.7, x.max() + sep_offset)
    axb.set_xticks(x); axb.set_xticklabels([str(yr) for _, yr in BARS], fontsize=9.5)

    fig.supylabel(ylabel, x=0.03, fontsize=12)
    top_ax.set_title(title, fontsize=13.5, fontweight="bold", pad=12,
                     loc="left", color="#1a1a1a")
    h, l = axes[0].get_legend_handles_labels()
    top_ax.legend(h[::-1], l[::-1], loc="upper left", bbox_to_anchor=(1.01, 1.0),
                  frameon=False, fontsize=10, title="Technology", title_fontsize=11)

    plt.subplots_adjust(left=0.10, right=0.80, top=0.91, bottom=0.17, hspace=0.06)

    # ── Scenario labels + separators — LABEL STRIP ONLY (below the x-axis) ──
    fig.canvas.draw()
    axb_bot = axb.get_position().y0
    line_bottom = axb_bot - 0.115
    label_y     = axb_bot - 0.075

    def d2f(xd):
        disp = axb.transData.transform((xd, 0))
        return fig.transFigure.inverted().transform(disp)[0]

    glist = list(groups.values())
    # interior separators: exact midpoint of each inter-group gap (offset = sep_offset
    # from the bars on either side, since the gap itself equals GROUP_GAP)
    sep_x = [(max(glist[i]) + min(glist[i + 1])) / 2.0 for i in range(len(glist) - 1)]
    # closing separator: same offset (sep_offset) from the last bar as every
    # interior separator uses, so all gaps between separator and bars match
    sep_x.append(x.max() + sep_offset)
    for sx in sep_x:
        xf = d2f(sx)
        fig.add_artist(plt.Line2D([xf, xf], [line_bottom, axb_bot],   # below axis only
                       transform=fig.transFigure, color="#BFBFBF",
                       linewidth=1.0, zorder=0.5))

    # left boundary of the label strip = extended y-axis line
    xL = axb.get_position().x0
    fig.add_artist(plt.Line2D([xL, xL], [line_bottom, axb_bot],
                   transform=fig.transFigure, color="#666666",
                   linewidth=0.9, zorder=0.5))

    for sc, xs in groups.items():
        fig.text(d2f(float(np.mean(xs))), label_y, sc, ha="center", va="top",
                 fontsize=12, fontweight="bold", color="#1a1a1a")

    for ext in ("png", "pdf"):
        fig.savefig(f"{outfile}.{ext}", bbox_inches="tight", facecolor="white")
    print(f"saved {outfile}  | totals: {[round(t,2) for t in totals]}")
    return fig


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    scenario_figure(GEN, GEN_ORDER, GEN_COLORS,
        ylabel="Annual generation (TWh)",
        title="Combined generation of 14 Pacific Island Countries by scenario",
        outfile="combined_generation", broken=True, brk=10, ymax=92)

    scenario_figure(CAP, GEN_ORDER, GEN_COLORS,
        ylabel="Installed capacity (GW)",
        title="Combined generation capacity of 14 Pacific Island Countries by scenario",
        outfile="combined_generation_capacity", broken=True, brk=5, ymax=66)

    scenario_figure(HEAT, HEAT_ORDER, HEAT_COLORS,
        ylabel="Annual heat generation (TWh)",
        title="Combined heat generation of 14 Pacific Island Countries by scenario",
        outfile="combined_heat_generation", broken=False, ymax=33)

    plt.show()   # display all three figures